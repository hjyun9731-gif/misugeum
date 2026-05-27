"""
강원도개인소형화물협회 통합관리시스템 v4
"""
import re, json, math, shutil, os
from pathlib import Path
from datetime import datetime, date
from typing import Optional, List
import tempfile

from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_, text, not_
import pandas as pd

from database import engine, SessionLocal, Base, get_db, IS_SQLITE
from models import (User, UploadBatch, RawImportRow, Member, MonthlyLedger,
                    MemberStatusEvent, WorkQueue, LicenseRecord, BillingPerson,
                    BankTransaction, CollectionTarget, Snapshot, AuditLog, BillingReport,
                    IncomeLedgerDetail)
from auth import hash_pw, verify_pw, ensure_admin, require_user
from core import (
    _s, norm_name, norm_vehicle, veh_last4, normalize_region, parse_amount,
    parse_date_str, detect_status, detect_status_from_mgmt_no,
    build_verify_reasons, phone_clean, address_clean,
    guess_sheet_type, guess_file_year,
    parse_ledger_sheet, parse_status_sheet,
    extract_memo_keys, is_useless_memo, score_bank_match,
    classify_autopay, AUTOPAY_AMOUNTS,
    parse_billing_file, BILLING_WORK_TYPES, BILLING_ACCOUNT_MAP,
)

app = FastAPI(title="강원도개인소형화물협회 v4")
app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SECRET_KEY", "gwf4-secret-2026"),
    max_age=3600 * 12
)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

CURRENT_YEAR = int(os.getenv("BILLING_YEAR", str(datetime.now().year)))
EXCLUDED_STATUSES = {"폐업","양도","이관","탈퇴","사망","말소","확인필요"}

# ── 스타트업 ──────────────────────────────────────────────────────────────────


def income_reason_display(v):
    raw = str(v or "").strip()
    if not raw:
        return ""

    parts = [x.strip() for x in raw.split("/")]

    labels = [
        "\uc218\ub3d9\ubd84\ub958",   # ????
        "\uc0ac\uc720",               # ??
        "\uad00\ub828\ucc28\ub7c9",   # ????
        "\uad00\ub828\uc131\uba85",   # ????
        "\ube44\uace0",               # ??
    ]

    fixed = []
    for i, part in enumerate(parts):
        if not part:
            continue

        # ?? ?? ???? ??? ?
        if any(k in part for k in labels):
            fixed.append(part)
            continue

        # ?? ?? ??: ????: ?, ??: ?
        if "?" in part and ":" in part:
            value = part.split(":", 1)[1].strip()
            label = labels[i] if i < len(labels) else labels[-1]
            fixed.append(label + ": " + value)
            continue

        fixed.append(part)

    return " / ".join(fixed)



@app.middleware("http")
async def redirect_old_work_to_pending_board(request: Request, call_next):
    """
    ?? ?????? /work ???? ????
    ? ????? /work/pending-board ? ?? ??.
    """
    if request.method == "GET" and request.url.path == "/work":
        qs = str(request.url.query)
        target = "/work/pending-board?" + qs if qs else "/work/pending-board"
        return RedirectResponse(target, status_code=302)

    return await call_next(request)


@app.middleware("http")
async def preserve_bank_status_redirect(request: Request, call_next):
    """
    Keep /bank tab filter after POST/redirect actions.
    If user is viewing /bank?status=... and presses a button,
    many routes redirect to /bank only. This middleware appends the
    original status/q query back to the redirect Location.
    """
    from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

    response = await call_next(request)

    try:
        if response.status_code not in (301, 302, 303, 307, 308):
            return response

        location = response.headers.get("location") or ""
        if not location:
            return response

        # Only fix redirects going back to /bank
        loc_parsed = urlparse(location)
        if loc_parsed.path != "/bank":
            return response

        loc_qs = parse_qs(loc_parsed.query)

        # If target already has status, do not touch it
        if "status" in loc_qs:
            return response

        referer = request.headers.get("referer") or ""
        ref_parsed = urlparse(referer)

        if ref_parsed.path != "/bank":
            return response

        ref_qs = parse_qs(ref_parsed.query)

        changed = False

        # Preserve current bank tab
        if "status" in ref_qs and ref_qs["status"]:
            loc_qs["status"] = ref_qs["status"]
            changed = True

        # Preserve search keyword too
        if "q" in ref_qs and ref_qs["q"]:
            loc_qs["q"] = ref_qs["q"]
            changed = True

        if not changed:
            return response

        new_query = urlencode(loc_qs, doseq=True)
        new_location = urlunparse((
            loc_parsed.scheme,
            loc_parsed.netloc,
            loc_parsed.path,
            loc_parsed.params,
            new_query,
            loc_parsed.fragment
        ))

        response.headers["location"] = new_location
        return response

    except Exception:
        return response


@app.on_event("startup")
def startup():
    if os.getenv("RESET_DB") == "1":
        Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    _migrate()
    db = SessionLocal()
    try: ensure_admin(db)
    finally: db.close()

def _migrate():
    with engine.begin() as conn:
        be = engine.url.get_backend_name()
        if be.startswith("postgres"):
            for tbl, col, typ in [
                ("members","is_overpay","BOOLEAN DEFAULT FALSE"),
                ("members","arrears_diff","INTEGER DEFAULT 0"),
                ("members","verify_reason","TEXT"),
                ("members","status_source","VARCHAR(50) DEFAULT 'auto'"),
                ("members","user_confirmed_match","BOOLEAN DEFAULT FALSE"),
                ("upload_batches","file_year","INTEGER"),
                ("license_records","source_sheet","VARCHAR(200)"),
                ("billing_reports","source_file","VARCHAR(300)"),
                ("billing_reports","raw_data","TEXT"),
                ("billing_reports","upload_type","VARCHAR(20) DEFAULT 'manual'"),
                ("billing_persons","billing_report_id","INTEGER REFERENCES billing_reports(id)"),
                ("billing_persons","raw_data","TEXT"),
                ("billing_persons","account","VARCHAR(20)"),
                ("billing_persons","charge_start_month","VARCHAR(10)"),
                ("billing_persons","charge_end_month","VARCHAR(10)"),
                ("billing_persons","from_status","VARCHAR(50)"),
                ("billing_persons","to_status","VARCHAR(50)"),
                ("billing_persons","reflected_at","TIMESTAMP WITH TIME ZONE"),
                ("billing_persons","reflected_by","INTEGER"),
            ]:
                conn.execute(text(f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS {col} {typ}"))
        elif IS_SQLITE:
            def _ac(tbl, col, typ):
                try:
                    cols = [r[1] for r in conn.execute(text(f"PRAGMA table_info({tbl})")).fetchall()]
                    if col not in cols:
                        conn.execute(text(f"ALTER TABLE {tbl} ADD COLUMN {col} {typ}"))
                except: pass
            for tbl, col, typ in [
                ("members","is_overpay","INTEGER DEFAULT 0"),
                ("members","arrears_diff","INTEGER DEFAULT 0"),
                ("members","verify_reason","TEXT"),
                ("members","status_source","TEXT DEFAULT 'auto'"),
                ("members","user_confirmed_match","INTEGER DEFAULT 0"),
                ("upload_batches","file_year","INTEGER"),
                ("license_records","source_sheet","TEXT"),
                ("billing_reports","source_file","TEXT"),
                ("billing_reports","raw_data","TEXT"),
                ("billing_reports","upload_type","TEXT DEFAULT 'manual'"),
                ("billing_persons","billing_report_id","INTEGER"),
                ("billing_persons","raw_data","TEXT"),
                ("billing_persons","account","TEXT"),
                ("billing_persons","charge_start_month","TEXT"),
                ("billing_persons","charge_end_month","TEXT"),
                ("billing_persons","from_status","TEXT"),
                ("billing_persons","to_status","TEXT"),
                ("billing_persons","reflected_at","DATETIME"),
                ("billing_persons","reflected_by","INTEGER"),
            ]:
                _ac(tbl, col, typ)

# ── 유틸 ──────────────────────────────────────────────────────────────────────
@app.get("/health")
def health(): return {"status": "ok"}

def add_log(db, user_id, action, detail=""):
    db.add(AuditLog(user_id=user_id, action=action, detail=str(detail)[:500]))
    db.commit()

def fmt_amt(v) -> str:
    if v is None: return "0"
    try:
        iv = int(v)
        if iv < 0: return f"-{abs(iv):,}"
        return f"{iv:,}"
    except: return str(v)


def fmt_date(v) -> str:
    """날짜를 26.05.18. 형식으로 표시"""
    if not v: return ""
    s = str(v).strip()
    import re as _re
    # 2026-05-18 → 26.05.18.
    m = _re.search(r"(20(\d{2}))-(\d{2})-(\d{2})", s)
    if m: return f"{m.group(2)}.{m.group(3)}.{m.group(4)}."
    # 이미 26.05.18. 형식
    if _re.match(r"\d{2}\.\d{2}\.\d{2}\.", s): return s
    return s

def fmt_acc(a) -> str: return {"협":"협회비","관":"관리비"}.get(a or "", a or "")

def _open_xl(path):
    p = Path(path)
    if p.suffix.lower() in {".xlsx",".xlsm"}:
        return pd.ExcelFile(path, engine="openpyxl")
    return pd.ExcelFile(path, engine="xlrd")

# ── 인증 ──────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    return RedirectResponse("/dashboard" if request.session.get("user_id") else "/login")

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html", {"request": request, "error": ""})

@app.post("/login")
def login(request: Request, username: str = Form(""), password: str = Form(""),
          db: Session = Depends(get_db)):
    u = db.query(User).filter(User.username == username).first()
    if not u or not verify_pw(password, u.password_hash):
        return templates.TemplateResponse(request, "login.html",
            {"request": request, "error": "아이디 또는 비밀번호가 올바르지 않습니다"})
    request.session["user_id"] = u.id
    return RedirectResponse("/dashboard", status_code=302)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login")

# ── 대시보드 ──────────────────────────────────────────────────────────────────
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db),
              user: User = Depends(require_user)):
    # 대시보드는 미수금/대상자 숫자가 자주 바뀌므로 캐시를 쓰지 않고 매번 재계산
    _invalidate_snap(db, "dashboard")
    snap = _build_dashboard_snap(db)
    _set_snap(db, "dashboard", snap)

    return templates.TemplateResponse(request, "dashboard.html", {
        "request": request, "user": user, "snap": snap, "fmt_amt": fmt_amt,
    })


SUM_NAMES_DB = {"합계","총계","소계","계","합산","인원수","입금금액"}

def _real_member_q(db: Session):
    """합계행 제외한 실제 회원 쿼리"""
    return db.query(Member).filter(
        ~Member.vehicle_no.in_(list(SUM_NAMES_DB)),
        ~Member.name_key.in_(list(SUM_NAMES_DB)),
        Member.name != None,
        Member.name != "",
    )

def _clean_filter():
    """합계행 제외 + 상태=정상 필터"""
    return and_(
        Member.status == "정상",
        ~Member.vehicle_no.in_(list(SUM_NAMES_DB)),
        ~Member.name_key.in_(list(SUM_NAMES_DB)),
        Member.name != None,
        Member.name != "",
    )


def _arrears_full_filter():
    """Full arrears/dashboard filter: exclude only strict summary rows."""
    sum_names = list(SUM_NAMES_DB)
    exclude_sum_row = and_(
        Member.name_key != None,
        Member.name_key != "",
        Member.name_key.in_(sum_names),
        or_(Member.vehicle_no == None, Member.vehicle_no == "", Member.vehicle_no.in_(sum_names))
    )
    return and_(Member.id != None, ~exclude_sum_row)

def _build_dashboard_snap(db: Session) -> dict:
    from datetime import date as _date
    now = datetime.now()
    cf = _arrears_full_filter()
    total      = db.query(Member).filter(cf).count()
    total_arr  = db.query(func.sum(Member.excel_arrears)).filter(cf, Member.excel_arrears > 0).scalar() or 0
    overpay_sum= db.query(func.sum(Member.excel_arrears)).filter(cf, Member.excel_arrears < 0).scalar() or 0
    verify_cnt = db.query(Member).filter(cf, Member.arrears_verified == False).count()
    no_lic     = db.query(Member).filter(cf,
                    Member.user_confirmed_match == False,
                    or_(Member.match_license_id == None, Member.match_status == "전체자미확인")).count()
    overpay_cnt= db.query(Member).filter(cf, Member.is_overpay == True).count()
    work_pending = db.query(WorkQueue).filter(WorkQueue.status == "반영대기").count()
    bank_pending = db.query(BankTransaction).filter(
        BankTransaction.applied == False,
        BankTransaction.match_status.in_(["자동매칭","확인필요"])).count()

    # 월별 입금 데이터 (최근 12개월)
    today_d = _date.today()
    monthly_data = []
    for i in range(11, -1, -1):
        y, m = today_d.year, today_d.month - i
        while m <= 0: m += 12; y -= 1
        paid_sum = db.query(func.sum(MonthlyLedger.paid_amount)).filter(
            MonthlyLedger.year == y,
            MonthlyLedger.month == m,
            MonthlyLedger.paid_amount > 0,
        ).scalar() or 0
        dep_cnt = db.query(func.count(func.distinct(MonthlyLedger.member_id))).filter(
            MonthlyLedger.year == y,
            MonthlyLedger.month == m,
            MonthlyLedger.paid_amount > 0,
        ).scalar() or 0
        monthly_data.append({
            "label": f"{y%100:02d}.{m:02d}",
            "paid": int(paid_sum),
            "depositors": int(dep_cnt),
        })

    # 부과대수 최신월 데이터
    latest_br = (db.query(BillingReport)
                 .order_by(BillingReport.year.desc(), BillingReport.month.desc())
                 .first())
    billing_data = {}
    if latest_br:
        billing_data = {
            "year": latest_br.year, "month": latest_br.month,
            "cnt_join":         latest_br.cnt_join or 0,
            "cnt_transfer":     latest_br.cnt_transfer or 0,
            "cnt_cross":        latest_br.cnt_cross or 0,
            "cnt_close":        latest_br.cnt_close or 0,
            "cnt_quit":         latest_br.cnt_quit or 0,
            "cnt_delivery_new": latest_br.cnt_delivery_new or 0,
            "cnt_mgmt_close":   latest_br.cnt_mgmt_close or 0,
            "cnt_age70":        latest_br.cnt_age70 or 0,
            "cnt_base":         latest_br.cnt_base or 0,
            "cnt_total":        latest_br.cnt_total or 0,
            "cnt_delivery":     latest_br.cnt_delivery or 0,
            "source_file":      latest_br.source_file or "",
            "upload_type":      latest_br.upload_type or "manual",
        }
    billing_pending = db.query(BillingPerson).filter(
        BillingPerson.reflect_status == "처리대기").count()
    mgmt_pending = db.query(BillingPerson).filter(
        BillingPerson.reflect_status == "처리대기",
        BillingPerson.account.in_(["관", "택배"])).count()

    return {
        "base_year": now.year, "base_month": now.month,
        "total_members": total, "total_arrears": int(total_arr),
        "overpay_sum": int(abs(overpay_sum)),
        "verify_cnt": verify_cnt, "no_lic": no_lic,
        "overpay_cnt": overpay_cnt, "work_pending": work_pending,
        "bank_pending": bank_pending,
        "monthly_data": monthly_data,
        "billing_data": billing_data,
        "billing_pending": billing_pending,
        "mgmt_pending": mgmt_pending,
    }

def _get_snap(db: Session, key: str):
    s = db.query(Snapshot).filter(Snapshot.snap_key == key).first()
    if not s: return None
    try: return json.loads(s.snap_value)
    except: return None

def _set_snap(db: Session, key: str, data: dict):
    s = db.query(Snapshot).filter(Snapshot.snap_key == key).first()
    val = json.dumps(data, ensure_ascii=False, default=str)
    if s: s.snap_value = val
    else: db.add(Snapshot(snap_key=key, snap_value=val))
    db.commit()

def _invalidate_snap(db: Session, *keys: str):
    for key in keys:
        db.query(Snapshot).filter(Snapshot.snap_key == key).delete()
    db.commit()

# ── 미수금 명단 ────────────────────────────────────────────────────────────────
@app.get("/arrears", response_class=HTMLResponse)
def arrears_page(request: Request, q: str = "", region: str = "",
                 account: str = "", amount_filter: str = "",
                 amount_min: int = 0,
                 status_filter: str = "", contact_filter: str = "",
                 sort: str = "", page: int = 1,
                 db: Session = Depends(get_db), user: User = Depends(require_user)):
    PAGE_SIZE = 200
    page = max(page, 1)
    base_q = db.query(Member).filter(_arrears_full_filter())

    if q:
        like = f"%{q}%"
        l4 = "".join(c for c in q if c.isdigit())[-4:]
        flt = [Member.name.ilike(like), Member.vehicle_no.ilike(like)]
        if l4: flt.append(Member.vehicle_no.ilike(f"%{l4}%"))
        base_q = base_q.filter(or_(*flt))
    if region:
        base_q = base_q.filter(Member.region == region)
    if account:
        acc_map = {
            "협": ["협", "협회비"],
            "관": ["관", "관리비"],
            "택배": ["택배", "택배관리", "택배관리비"],
        }
        base_q = base_q.filter(Member.account.in_(acc_map.get(account, [account])))

    # 금액 필터 (amount_filter 프리셋)
    # 중요: amount_filter가 비어 있으면 "전체 보기"이므로 0원/완납자도 포함한다.
    if amount_filter == "미수만":
        base_q = base_q.filter(Member.excel_arrears > 0)
    elif amount_filter == "완납":
        base_q = base_q.filter(Member.excel_arrears == 0)
    elif amount_filter in ("초과납부", "초과납부선납"):
        base_q = base_q.filter(Member.excel_arrears < 0)
    elif amount_filter == "고액미납":
        # 계정별 고액 기준: 협회비 30만+, 관리비/택배 10만+
        _acc_vals_hyup = ["협", "협회비"]
        _acc_vals_gwan = ["관", "관리비", "택배", "택배관리", "택배관리비"]
        if account in _acc_vals_hyup:
            base_q = base_q.filter(Member.excel_arrears >= 300000)
        elif account in _acc_vals_gwan:
            base_q = base_q.filter(Member.excel_arrears >= 100000)
        else:
            base_q = base_q.filter(
                or_(
                    and_(Member.account.in_(_acc_vals_hyup), Member.excel_arrears >= 300000),
                    and_(Member.account.in_(_acc_vals_gwan), Member.excel_arrears >= 100000),
                )
            )

    # 슬라이더 최소금액 (amount_min > 0 이면 추가 필터)
    if amount_min > 0:
        base_q = base_q.filter(Member.excel_arrears >= amount_min)

    # 상태 필터
    if status_filter == "완납제외":
        base_q = base_q.filter(Member.excel_arrears != 0)
    elif status_filter == "초과제외":
        base_q = base_q.filter(Member.is_overpay == False)
    elif status_filter == "확인필요":
        base_q = base_q.filter(Member.arrears_verified == False)

    # 연락처 필터
    if contact_filter == "있음":
        base_q = base_q.filter(
            or_(
                and_(Member.mobile != None, Member.mobile != ""),
                and_(Member.phone != None, Member.phone != "")
            )
        )
    elif contact_filter == "없음":
        base_q = base_q.filter(
            or_(Member.mobile == None, Member.mobile == ""),
            or_(Member.phone == None, Member.phone == ""),
        )

    # 정렬
    if sort == "amount_desc":
        base_q = base_q.order_by(Member.excel_arrears.desc())
    elif sort == "amount_asc":
        base_q = base_q.order_by(Member.excel_arrears.asc())
    elif sort == "date_asc":
        if IS_SQLITE:
            base_q = base_q.order_by(Member.last_paid_date.asc())
        else:
            base_q = base_q.order_by(Member.last_paid_date.asc().nullsfirst())
    else:
        base_q = base_q.order_by(Member.region, Member.name)

    total_count = base_q.order_by(None).count()
    total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(1, min(page, total_pages))
    items = base_q.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()

    total_arr = base_q.order_by(None).filter(Member.excel_arrears > 0).with_entities(
        func.sum(Member.excel_arrears)
    ).scalar() or 0

    regions = sorted({x[0] for x in db.query(Member.region).distinct().filter(Member.region != None).all() if x[0]})

    return templates.TemplateResponse(request, "arrears.html", {
        "request": request, "user": user, "items": items, "q": q,
        "region": region, "account": account,
        "amount_filter": amount_filter, "amount_min": amount_min,
        "status_filter": status_filter,
        "contact_filter": contact_filter, "sort": sort,
        "total_arr": int(total_arr), "regions": regions,
        "fmt_amt": fmt_amt, "fmt_acc": fmt_acc,
        "msg": request.query_params.get("msg", ""),
        "page": page, "total_pages": total_pages, "total_count": total_count,
    })

# 부과대수 관리으로 보내기
@app.post("/arrears/{mid}/to-work")
def to_work_queue(mid: int, process_type: str = Form("폐업"),
                  reason: str = Form(""), event_date: str = Form(""),
                  db: Session = Depends(get_db), user: User = Depends(require_user)):
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: raise HTTPException(404)
    db.add(WorkQueue(
        member_id=mid, process_type=process_type, status="반영대기",
        source_screen="미수금명단", reason=reason,
        event_date_raw=event_date, arrears_at_submit=m.excel_arrears or 0,
        submitted_by=user.id
    ))
    db.commit()
    _invalidate_snap(db, "dashboard")
    add_log(db, user.id, "업무처리대기등록", f"{m.name}/{m.vehicle_no} → {process_type}")
    return RedirectResponse(f"/arrears?msg={m.name} 부과대수 관리으로 이동", status_code=302)



# ── N월 입금추출 ─────────────────────────────────────────────
@app.get("/payments/export")
def export_monthly_payments(
    year: int = None,
    month: int = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    입금전표 입력용 N월 입금추출.
    출력 컬럼: 지역 / 계정 / 차량번호 / 성명 / 입금액
    """
    from datetime import datetime
    from io import BytesIO
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    now = datetime.now()
    y = int(year or now.year)
    mth = int(month or now.month)

    rows = (
        db.query(Member, MonthlyLedger)
        .join(MonthlyLedger, MonthlyLedger.member_id == Member.id)
        .filter(MonthlyLedger.year == y)
        .filter(MonthlyLedger.month == mth)
        .filter(MonthlyLedger.paid_amount != None)
        .filter(MonthlyLedger.paid_amount > 0)
        .order_by(Member.region, Member.account, Member.vehicle_no, Member.name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}-{mth:02d} 입금"

    headers = ["지역", "계정", "차량번호", "성명", "입금액"]
    ws.append(headers)

    for member, ledger in rows:
        acc = member.account or ""
        if acc == "관":
            acc = "관리비"
        elif acc == "협":
            acc = "협회비"

        ws.append([
            member.region or "",
            acc,
            member.vehicle_no or "",
            member.name or "",
            int(ledger.paid_amount or 0),
        ])

    # 합계행
    total_row = ws.max_row + 1
    ws.cell(total_row, 4).value = "합계"
    ws.cell(total_row, 5).value = f"=SUM(E2:E{total_row-1})"

    header_fill = PatternFill("solid", fgColor="FCE7F3")
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True, color="9D174D")
                cell.fill = header_fill

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = [14, 12, 18, 16, 14][col-1]

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 5).number_format = '#,##0'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = quote(f"{y}년_{mth:02d}월_입금추출.xlsx")
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )




# ── 명단 추가: /member/{mid} 충돌 방지용 /member_add ─────────────
@app.get("/member_add", response_class=HTMLResponse)
def member_add_page(request: Request, db: Session = Depends(get_db), user: User = Depends(require_user)):
    regions = [
        "춘천시","강릉시","원주시","동해시","태백시","속초시","삼척시",
        "홍천군","횡성군","영월군","평창군","정선군","철원군","화천군",
        "양구군","인제군","고성군","양양군"
    ]
    return templates.TemplateResponse(request, "add_member.html", {
        "request": request,
        "user": user,
        "regions": regions,
        "m": None,
    })


@app.post("/member_add")
def member_add_save(
    region: str = Form(""),
    account: str = Form("관리비"),
    vehicle_no: str = Form(""),
    name: str = Form(""),
    mobile: str = Form(""),
    address: str = Form(""),
    excel_arrears: str = Form("0"),
    note: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote

    def _amt(v):
        try:
            return int(str(v or "0").replace(",", "").replace("원", "").strip() or 0)
        except Exception:
            return 0

    amount = _amt(excel_arrears)
    acc = "관" if ("관" in (account or "") or "관리" in (account or "")) else "협"

    m = Member(
        region=normalize_region(region) if region else "",
        account=acc,
        vehicle_no=vehicle_no,
        vehicle_key=norm_vehicle(vehicle_no) if vehicle_no else "",
        name=name,
        name_key=norm_name(name) if name else "",
        mobile=phone_clean(mobile) or mobile,
        address=address,
        excel_arrears=amount,
        calc_arrears=amount,
        arrears_diff=0,
        is_overpay=amount < 0,
        arrears_verified=True,
        status="정상",
        status_source="manual",
        note=note,
        source_file="수기추가",
        source_sheet="명단추가",
    )
    db.add(m)
    db.commit()
    db.refresh(m)

    try:
        add_log(db, user.id, "명단추가", f"{m.name}/{m.vehicle_no}/{m.excel_arrears}원")
    except Exception:
        pass

    _invalidate_snap(db, "dashboard")
    return RedirectResponse("/arrears?msg=" + quote("명단 추가 완료"), status_code=302)


@app.get("/member/new")
def member_new_redirect():
    return RedirectResponse("/member_add", status_code=302)


# ── 회원 상세 ──────────────────────────────────────────────────────────────────
@app.get("/member/{mid}", response_class=HTMLResponse)
def member_detail(mid: int, request: Request, db: Session = Depends(get_db),
                  user: User = Depends(require_user)):
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: raise HTTPException(404)
    ledgers = (db.query(MonthlyLedger).filter(MonthlyLedger.member_id == mid)
               .order_by(MonthlyLedger.year, MonthlyLedger.month).all())
    # 중복 원장 감지
    seen = set(); dups = 0
    deduped = []
    for l in ledgers:
        key = (l.year, l.month, l.source_sheet)
        if key in seen: dups += 1; l.is_duplicate = True
        else: seen.add(key); deduped.append(l)

    events = (db.query(MemberStatusEvent).filter(MemberStatusEvent.member_id == mid)
              .order_by(MemberStatusEvent.created_at).all())
    lic = db.query(LicenseRecord).filter(LicenseRecord.id == m.match_license_id).first() if m.match_license_id else None
    work_history = (db.query(WorkQueue).filter(WorkQueue.member_id == mid)
                    .order_by(WorkQueue.submitted_at.desc()).limit(10).all())

    return templates.TemplateResponse(request, "member_detail.html", {
        "request": request, "user": user, "m": m,
        "ledgers": deduped, "dup_count": dups, "events": events,
        "lic": lic, "work_history": work_history,
        "fmt_amt": fmt_amt, "fmt_acc": fmt_acc,
        "msg": request.query_params.get("msg", ""),
    })

@app.get("/member/{mid}/edit", response_class=HTMLResponse)
def member_edit_page(mid: int, request: Request, db: Session = Depends(get_db),
                     user: User = Depends(require_user)):
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: raise HTTPException(404)
    regions = sorted({x[0] for x in db.query(Member.region).distinct().filter(Member.region != None).all() if x[0]})
    return templates.TemplateResponse(request, "edit_member.html",
        {"request": request, "user": user, "m": m, "regions": regions, "fmt_acc": fmt_acc})

@app.post("/member/{mid}/edit")
def member_edit_save(mid: int,
    name: str = Form(""), vehicle_no: str = Form(""), region: str = Form(""),
    account: str = Form("협"), mobile: str = Form(""), phone: str = Form(""),
    address: str = Form(""), official_address: str = Form(""),
    status: str = Form("정상"), note: str = Form(""),
    join_date_raw: str = Form(""), permit_date_raw: str = Form(""),
    cert_issue_date_raw: str = Form(""), cert_no: str = Form(""),
    db: Session = Depends(get_db), user: User = Depends(require_user)):
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: raise HTTPException(404)
    m.name = name; m.name_key = norm_name(name)
    m.vehicle_no = vehicle_no; m.vehicle_key = norm_vehicle(vehicle_no)
    m.region = region; m.account = account
    m.mobile = phone_clean(mobile) or mobile; m.phone = phone
    m.address = address; m.official_address = official_address
    m.status = status; m.note = note
    m.join_date_raw = join_date_raw; m.permit_date_raw = permit_date_raw
    m.cert_issue_date_raw = cert_issue_date_raw; m.cert_no = cert_no
    db.commit()
    _invalidate_snap(db, "dashboard")
    add_log(db, user.id, "회원수정", f"id={mid} {name}")
    return RedirectResponse(f"/member/{mid}?msg=저장완료", status_code=302)

# ── 엑셀 업로드 ────────────────────────────────────────────────────────────────
@app.get("/upload", response_class=HTMLResponse)
def upload_page(request: Request, db: Session = Depends(get_db),
                user: User = Depends(require_user)):
    batches = db.query(UploadBatch).order_by(UploadBatch.created_at.desc()).limit(20).all()
    latest = {dt: db.query(UploadBatch).filter(UploadBatch.data_type == dt)
              .order_by(UploadBatch.created_at.desc()).first()
              for dt in ["legacy","license","billing","bank","bank_auto"]}
    return templates.TemplateResponse(request, "upload.html", {
        "request": request, "user": user, "batches": batches, "latest": latest,
        "msg": request.query_params.get("msg", ""),
    })


def _guess_name_from_raw_row(db: Session, batch_id: int, src_sheet: str, src_row: int, current_name: str = ""):
    """
    parse_ledger_sheet가 성명을 못 읽은 경우 RawImportRow 원본 행에서 이름 후보를 찾는다.
    차량번호/지역/계정/금액/합계행 단어는 제외하고, 한글 2~8자 또는 업체명 후보를 사용한다.
    """
    if current_name and str(current_name).strip():
        return current_name

    import re, json as _json
    raw = (db.query(RawImportRow)
             .filter(RawImportRow.batch_id == batch_id,
                     RawImportRow.source_sheet == src_sheet,
                     RawImportRow.source_row == src_row)
             .first())
    if not raw or not raw.raw_data:
        return current_name or ""

    try:
        vals = _json.loads(raw.raw_data)
    except Exception:
        return current_name or ""

    bad_words = {
        "강릉시","춘천시","원주시","동해시","태백시","속초시","삼척시",
        "홍천군","횡성군","영월군","평창군","정선군","철원군","화천군","양구군","인제군","고성군","양양군",
        "협회비","관리비","택배","협","관","합계","총계","소계","계","입금금액","미수금","잔액",
        "차량번호","성명","이름","지역","전화번호","연락처"
    }

    candidates = []
    for v in vals:
        t = str(v or "").strip()
        if not t:
            continue
        t = t.replace(" ", "")

        if t in bad_words:
            continue
        if re.search(r"\d", t):
            continue
        if len(t) < 2 or len(t) > 20:
            continue

        # 일반 성명, 외국인 영문명, 법인/업체명 허용
        if re.fullmatch(r"[가-힣]{2,8}", t):
            candidates.append(t)
        elif re.search(r"[㈜주식회사협동조합]", t) and len(t) <= 20:
            candidates.append(t)
        elif re.fullmatch(r"[A-Za-z]{3,20}", t):
            candidates.append(t)

    return candidates[0] if candidates else (current_name or "")



def _get_name_from_col_e(db: Session, batch_id: int, src_sheet: str, src_row: int, current_name: str = ""):
    """
    미수금 원본에서 성명 보정.
    기존 name이 비어 있으면 raw row에서 이름 후보를 찾는다.
    단, '이체' 같은 거래구분/메모값은 절대 성명으로 쓰지 않는다.
    """
    if current_name and str(current_name).strip() and str(current_name).strip() not in ("이체", "입금", "출금"):
        return str(current_name).strip().replace(" ", "")

    import json as _json
    raw = (db.query(RawImportRow)
             .filter(RawImportRow.batch_id == batch_id,
                     RawImportRow.source_sheet == src_sheet,
                     RawImportRow.source_row == src_row)
             .first())
    if not raw or not raw.raw_data:
        return ""

    try:
        vals = _json.loads(raw.raw_data)
    except Exception:
        return ""

    bad_words = {
        "이체","입금","출금","현금","카드","대체","자동이체",
        "관리비","협회비","택배","협","관",
        "강릉시","춘천시","원주시","동해시","태백시","속초시","삼척시",
        "홍천군","횡성군","영월군","평창군","정선군","철원군","화천군","양구군","인제군","고성군","양양군",
        "합계","총계","소계","계","합산","인원수","입금금액","미수금","잔액",
        "차량번호","성명","이름","지역","전화번호","연락처","비고"
    }

    candidates = []

    for v in vals:
        t = str(v or "").strip().replace(" ", "")
        if not t:
            continue
        if t in bad_words:
            continue
        if len(t) < 2 or len(t) > 25:
            continue
        if re.search(r"\d", t):
            continue

        # 일반 한글 성명
        if re.fullmatch(r"[가-힣]{2,8}", t):
            candidates.append(t)
            continue

        # 외국인 영문명
        if re.fullmatch(r"[A-Za-z]{3,25}", t):
            candidates.append(t)
            continue

        # 업체명/법인명
        if ("㈜" in t or "(주)" in t or "주식회사" in t or "협동조합" in t) and len(t) <= 25:
            candidates.append(t)
            continue

    return candidates[0] if candidates else ""



def _get_name_from_excel_col_e(db: Session, batch_id: int, src_sheet: str, src_row: int, current_name: str = ""):
    """
    [사용]미수금2026 원본 기준:
    A 지역 / B 계정 / C 비고 또는 구분 / D 차량번호 / E 성명.
    name이 비었거나 '이체' 같은 구분값이면 E열(index 4)을 성명으로 사용한다.
    """
    import json as _json

    bad = {"", "이체", "입금", "출금", "현금", "자동이체", "협회비", "관리비", "택배", "협", "관"}

    cur = str(current_name or "").strip().replace(" ", "")
    if cur and cur not in bad:
        return cur

    raw = (db.query(RawImportRow)
             .filter(RawImportRow.batch_id == batch_id,
                     RawImportRow.source_sheet == src_sheet,
                     RawImportRow.source_row == src_row)
             .first())

    if not raw or not raw.raw_data:
        return "" if cur in bad else cur

    try:
        vals = _json.loads(raw.raw_data)
    except Exception:
        return "" if cur in bad else cur

    if len(vals) >= 5:
        name = str(vals[4] or "").strip().replace(" ", "")
        if name not in bad:
            return name

    return "" if cur in bad else cur



def _find_or_create_member(db, veh, name, region, account, note, batch_id, src_file, src_sheet, src_row, force_create=False):
    from core import is_sum_row
    if is_sum_row(name, veh): return None, "sum_row"
    vkey = norm_vehicle(veh) if veh else ""
    nkey = norm_name(name) if name else ""
    reg  = normalize_region(region) if region else ""
    acc  = "관" if ("관" in (account or "") or "관리" in (account or "")) else "협"
    is_auto = False  # 자동이체는 전용통장 업로드 후에만 확정

    # 미수금 원장 업로드에서는 엑셀 원본 행을 그대로 보존해야 한다.
    # 차량번호/성명 중복이어도 같은 사람으로 병합하지 않고 새 Member로 저장한다.
    if not force_create:
        if vkey and nkey:
            m = db.query(Member).filter(Member.vehicle_key == vkey, Member.name_key == nkey).first()
            if m: return m, "exact"
        if vkey:
            cs = db.query(Member).filter(Member.vehicle_key == vkey).all()
            if len(cs) == 1: return cs[0], "vehicle"
            if len(cs) > 1:  return None, "dup_vehicle"
        if nkey and reg:
            cs = db.query(Member).filter(Member.name_key == nkey, Member.region == reg, Member.account == acc).all()
            if len(cs) == 1: return cs[0], "name_region"
            if len(cs) > 1:  return None, "dup_name"

    m = Member(vehicle_no=veh, vehicle_key=vkey, name=name, name_key=nkey,
               region=reg, account=acc, note=note, is_auto_transfer=is_auto,
               status="정상", status_source="auto",
               source_file=src_file, source_sheet=src_sheet,
               source_row=src_row, source_batch_id=batch_id)
    db.add(m); db.flush()
    return m, "created"

def _recalc_member(db: Session, m: Member):
    """
    회원 미수금 갱신.
    핵심 규칙:
    - excel_arrears = 엑셀 기재 기준월 미수금 (우선값, 음수 그대로 유지)
    - calc_arrears  = excel_arrears (재계산 누적 완전 중단)
    - arrears_diff  = 0
    - is_overpay    = excel_arrears < 0
    - verify_reason = 실제 사유 있을 때만 (차이 0원 금지)
    """
    ledgers = (db.query(MonthlyLedger).filter(MonthlyLedger.member_id == m.id)
               .order_by(MonthlyLedger.year, MonthlyLedger.month).all())

    if ledgers:
        # 기준: 최신월 원장의 arrears_amount (음수 그대로)
        latest = ledgers[-1]
        excel_arr = latest.arrears_amount
        if excel_arr is None:
            excel_arr = m.excel_arrears or 0
        m.excel_arrears = excel_arr
        # 최초 미납월
        m.first_unpaid_month = ""
        for l in ledgers:
            if (l.arrears_amount or 0) > 0:
                m.first_unpaid_month = f"{l.year}-{l.month:02d}"
                break
        # 최근 입금일
        paid_ls = [l for l in reversed(ledgers) if (l.paid_amount or 0) > 0 and l.paid_date]
        if paid_ls:
            m.last_paid_date = paid_ls[0].paid_date
    else:
        excel_arr = m.excel_arrears or 0

    # 재계산 누적 완전 중단 — calc = excel
    m.calc_arrears = excel_arr
    m.arrears_diff = 0
    m.is_overpay   = excel_arr < 0

    # 검증사유: 실제 문제 있을 때만 (차이 0원이므로 금액차이는 없음)
    reasons = build_verify_reasons(
        excel_arr, excel_arr,  # diff = 0 → 금액차이 사유 절대 생성 안 됨
        m.name or "", m.vehicle_no or "", m.account or "",
        m.match_status or "", m.region or "",
        m.mobile or "", m.address or "", m.note or ""
    )
    if reasons:
        m.verify_reason = " | ".join(reasons)
        m.arrears_verified = False
    else:
        m.verify_reason = None
        m.arrears_verified = True

def _full_recalc(db: Session):
    """전체 회원 재계산 후 스냅샷 갱신"""
    for m in db.query(Member).all():
        _recalc_member(db, m)
    db.commit()
    _invalidate_snap(db, "dashboard")

@app.post("/upload/legacy")
async def upload_legacy(request: Request, file: UploadFile = File(...),
                        db: Session = Depends(get_db), user: User = Depends(require_user)):
    tmpdir = Path(tempfile.mkdtemp())
    try:
        dest = tmpdir / (file.filename or "upload.xlsx")
        with open(dest, "wb") as f: shutil.copyfileobj(file.file, f)
        if dest.stat().st_size == 0:
            return RedirectResponse("/upload?msg=파일이 비어 있습니다", status_code=302)

        file_year = guess_file_year(file.filename or "")
        try: xl = _open_xl(dest)
        except Exception as e:
            return RedirectResponse(f"/upload?msg=파일 열기 실패: {e}", status_code=302)

        batch = UploadBatch(file_name=file.filename, data_type="legacy",
                            file_year=file_year, created_by=user.id)
        db.add(batch); db.commit(); db.refresh(batch)

        nm = nl = ne = 0; warns = []
        for sname in xl.sheet_names:
            try: raw = pd.read_excel(xl, sheet_name=sname, header=None, dtype=object)
            except: continue
            if raw is None or raw.empty: continue

            rows_list = raw.values.tolist()
            h = [_s(v) for v in rows_list[0]] if rows_list else []
            stype = guess_sheet_type(sname, h)

            for ridx, row in enumerate(rows_list):
                vals = [_s(v) for v in row]
                if not any(v for v in vals if v): continue
                db.add(RawImportRow(batch_id=batch.id, source_file=file.filename,
                    source_sheet=sname, source_row=ridx+1, sheet_type=stype,
                    raw_data=json.dumps(vals, ensure_ascii=False, default=str)))

            if stype == "ledger":
                parsed = parse_ledger_sheet(raw, sname, file_year)
                for veh,name,region,account,note,carry,monthly,src_row in parsed:
                    name = _get_name_from_excel_col_e(db, batch.id, sname, src_row, name)
                    m, how = _find_or_create_member(db, veh, name, region, account, note,
                                                    batch.id, file.filename, sname, src_row, force_create=True)
                    if m is None: warns.append(f"{sname}/{src_row}: 중복후보"); continue
                    if how == "created": nm += 1

                    for mo,(charge,paid,arrears,pdate) in monthly.items():
                        exist = db.query(MonthlyLedger).filter(
                            MonthlyLedger.member_id == m.id,
                            MonthlyLedger.year == file_year,
                            MonthlyLedger.month == mo,
                            MonthlyLedger.source_sheet == sname,
                            MonthlyLedger.source_row == src_row,
                        ).first()
                        if exist: continue
                        carry_v = carry if mo == min(monthly.keys()) else 0
                        calc = carry_v + charge - paid
                        db.add(MonthlyLedger(
                            member_id=m.id, batch_id=batch.id,
                            source_file=file.filename, source_sheet=sname, source_row=src_row,
                            raw_vehicle_no=veh, raw_name=name,
                            raw_region=region, raw_account=account, raw_note=note,
                            year=file_year, month=mo,
                            carry_over=carry_v, charge_amount=charge,
                            paid_amount=paid, arrears_amount=arrears,
                            paid_date=pdate,
                            calc_arrears=arrears,   # 재계산 누적 중단: calc = excel
                            verified=True,          # 차이 0이므로 항상 verified
                        ))
                        nl += 1
                    db.flush()

            elif stype == "status":
                parsed = parse_status_sheet(raw, file_year)
                for veh,name,region,account,reason,arrears,src_row,em in parsed:
                    evtype = detect_status(reason)
                    if not evtype: continue
                    m, how = _find_or_create_member(db, veh, name, region,
                        account or "협", "", batch.id, file.filename, sname, src_row)
                    if m:
                        if how == "created": nm += 1
                        if evtype in EXCLUDED_STATUSES:
                            m.status = evtype; m.status_source = "column"
                        db.add(MemberStatusEvent(
                            member_id=m.id, batch_id=batch.id,
                            source_file=file.filename, source_sheet=sname, source_row=src_row,
                            raw_vehicle_no=veh, raw_name=name,
                            raw_region=normalize_region(region), raw_account=account,
                            event_type=evtype, event_date_raw=parse_date_str(reason),
                            reason_raw=reason, arrears_at_event=arrears, event_month=em or "",
                        ))
                        ne += 1
                    db.flush()
            db.commit()

        # 재계산 (excel_arrears 동기화 + is_overpay 갱신)
        # 한 번에 Member 객체 전체를 오래 붙잡으면 reset/upload와 데드락이 날 수 있어 id만 짧게 조회
        member_ids = [x[0] for x in db.query(Member.id).filter(Member.source_batch_id == batch.id).all()]
        for mid in member_ids:
            m_obj = db.get(Member, mid)
            if m_obj:
                _recalc_member(db, m_obj)
        db.commit()
        _invalidate_snap(db, "dashboard")
        batch.total_rows = db.query(RawImportRow).filter(RawImportRow.batch_id == batch.id).count()
        batch.saved_rows = nl; batch.warn_rows = len(warns); db.commit()
        add_log(db, user.id, "미수금업로드",
                f"{file.filename}({file_year}): 회원{nm}명 원장{nl}행 이벤트{ne}건")
        msg = f"업로드완료: 신규회원 {nm}명, 원장 {nl}행"
        if warns: msg += f" (경고 {len(warns)}건)"
    except Exception as e:
        msg = f"오류: {e}"
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    return RedirectResponse(f"/upload?msg={msg}", status_code=302)

@app.post("/upload/license")
async def upload_license(request: Request, file: UploadFile = File(...),
                          db: Session = Depends(get_db), user: User = Depends(require_user)):
    tmpdir = Path(tempfile.mkdtemp())
    try:
        dest = tmpdir / (file.filename or "license.xlsx")
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)

        if dest.stat().st_size == 0:
            return RedirectResponse("/upload?msg=파일이 비어 있습니다", status_code=302)

        from utils.excel_parser import (
            read_excel_sheets, choose,
            get_region as _gr, clean_name as _cn,
            normalize_vehicle as _nv
        )

        db.query(LicenseRecord).delete(synchronize_session=False)
        db.query(UploadBatch).filter(UploadBatch.data_type == "license").delete(synchronize_session=False)
        db.commit()

        batch = UploadBatch(file_name=file.filename, data_type="license", created_by=user.id)
        db.add(batch)
        db.commit()
        db.refresh(batch)

        try:
            sheets = read_excel_sheets(str(dest))
        except ValueError as e:
            return RedirectResponse(f"/upload?msg={e}", status_code=302)

        GANGWON_REGIONS = [
            "춘천시","강릉시","원주시","동해시","태백시","속초시","삼척시",
            "홍천군","횡성군","영월군","평창군","정선군","철원군","화천군",
            "양구군","인제군","고성군","양양군"
        ]

        SUM_WORDS = {"합계","총계","소계","계","합산","인원수","입금금액","성명","이름","차량번호","지역"}

        def _txt(v):
            if v is None:
                return ""
            try:
                if str(v).lower() == "nan":
                    return ""
            except Exception:
                pass
            return str(v).strip()

        def _row_values(row):
            return [_txt(v) for v in list(row.values)]

        def _by_header(row, aliases):
            for col, val in row.items():
                c = _txt(col).replace(" ", "")
                for a in aliases:
                    if a.replace(" ", "") in c:
                        return _txt(val)
            return ""

        def _fallback_region(row):
            v = _by_header(row, ["지역", "관할", "시군", "시·군"])
            if v:
                for r in GANGWON_REGIONS:
                    if r in v:
                        return r
                return v
            for v in _row_values(row):
                for r in GANGWON_REGIONS:
                    if r in v:
                        return r
            return ""

        def _fallback_vehicle(row):
            v = _by_header(row, ["차량번호", "자동차등록번호", "등록번호", "차번", "차량"])
            if v:
                return v

            for v in _row_values(row):
                t = v.replace(" ", "")
                if not t:
                    continue
                if any(x in t for x in ["배", "바", "아", "자"]):
                    if any(ch.isdigit() for ch in t):
                        return v
                if re.fullmatch(r"\d{2,3}[- ]?\d{4}", t):
                    return v
            return ""

        def _fallback_name(row):
            v = _by_header(row, ["성명", "대표자", "이름", "성 명", "차주명", "소유자"])
            v = _cn(v)
            if v and v not in SUM_WORDS and not any(ch.isdigit() for ch in v):
                return v

            bad = set(GANGWON_REGIONS) | SUM_WORDS | {
                "협회비","관리비","택배","주소","전화번호","휴대폰","핸드폰",
                "면허번호","자격증명","인가일자","가입일자"
            }

            for v in _row_values(row):
                t = _cn(v)
                if not t:
                    continue
                if t in bad:
                    continue
                if any(ch.isdigit() for ch in t):
                    continue
                if len(t) < 2 or len(t) > 25:
                    continue
                if re.fullmatch(r"[가-힣]{2,8}", t):
                    return t
                if "㈜" in t or "(주)" in t or "주식회사" in t or "협동조합" in t:
                    return t
            return ""

        def _field(row, colmap, key, aliases):
            try:
                v = choose(row, colmap, key)
                if _txt(v):
                    return _txt(v)
            except Exception:
                pass
            return _by_header(row, aliases)

        saved = 0
        skipped = 0

        for sname, df, hrow, colmap in sheets:
            for idx, row in df.iterrows():
                vals = _row_values(row)
                if not any(vals):
                    continue

                nm_ = _cn(_field(row, colmap, "name", ["성명", "대표자", "이름", "차주명", "소유자"]))
                veh_ = _field(row, colmap, "vehicle_no", ["차량번호", "자동차등록번호", "등록번호", "차번"])

                if not nm_:
                    nm_ = _fallback_name(row)
                if not veh_:
                    veh_ = _fallback_vehicle(row)

                region_ = _field(row, colmap, "region", ["지역", "관할", "시군", "시·군"])
                if not region_:
                    region_ = _fallback_region(row)

                # 진짜 합계/헤더행 제외
                if (nm_ in SUM_WORDS) or (veh_ in SUM_WORDS):
                    skipped += 1
                    continue

                # 이름/차량번호 둘 다 없으면 제외
                if not nm_ and not veh_:
                    skipped += 1
                    continue

                mobile_ = _field(row, colmap, "mobile", ["휴대폰", "핸드폰", "휴대전화", "연락처"])
                phone_ = _field(row, colmap, "phone", ["전화번호", "일반전화"])
                address_ = _field(row, colmap, "address", ["주소", "소재지"])
                official_address_ = _field(row, colmap, "official_address", ["공문주소", "공식주소", "주소"])
                resident_no_ = _field(row, colmap, "resident_no", ["주민등록번호", "생년월일", "법인번호"])
                join_date_ = _field(row, colmap, "join_date", ["가입일자", "협회가입일"])
                permit_date_ = _field(row, colmap, "permit_date", ["인가일자", "허가일자"])
                cert_issue_date_ = _field(row, colmap, "cert_issue_date", ["자격증명발급일자", "자격증명일자"])
                cert_no_ = _field(row, colmap, "cert_no", ["자격증명발급번호", "자격증명번호"])
                note_ = _field(row, colmap, "note", ["비고", "메모"])

                db.add(LicenseRecord(
                    batch_id=batch.id,
                    source_file=file.filename,
                    source_sheet=sname,
                    region=region_,
                    name=nm_,
                    name_key=norm_name(nm_),
                    vehicle_no=veh_,
                    vehicle_key=_nv(veh_) if veh_ else "",
                    resident_no=resident_no_,
                    mobile=mobile_,
                    phone=phone_,
                    address=address_,
                    official_address=official_address_,
                    join_date_raw=join_date_,
                    permit_date_raw=permit_date_,
                    cert_issue_date_raw=cert_issue_date_,
                    cert_no=cert_no_,
                    note=note_,
                ))
                saved += 1

            db.flush()

        batch.saved_rows = saved
        batch.warn_rows = skipped
        db.commit()

        _reconcile_license(db)
        _full_recalc(db)

        try:
            add_log(db, user.id, "전체자업로드", f"{file.filename}: {saved}건 저장, 제외 {skipped}건")
        except Exception:
            pass

        msg = f"전체자명단 {saved}건 저장, 대조완료"
        if skipped:
            msg += f" (제외 {skipped}건)"

    except Exception as e:
        db.rollback()
        msg = f"오류: {str(e)[:200]}"
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    return RedirectResponse(f"/upload?msg={msg}", status_code=302)


def _reconcile_license(db: Session):
    """
    전체자명단 ↔ 미수금명단 대조.
    정규화: norm_vehicle(강원·공백·호 제거), norm_name(공백제거)
    우선순위:
    1) vkey + 성명 완전일치 → 정상매칭
    2) 뒷자리4 + 성명 일치 → 정상매칭
    3) vkey 단독 단일후보 → 성명확인필요
    4) 뒷자리 단독 단일후보 → 뒷자리매칭
    5) 성명 단독 단일후보 → 차량번호확인필요
    """
    from core import is_sum_row
    lics = db.query(LicenseRecord).all()
    by_vkey: dict = {}
    by_l4: dict = {}
    by_nkey: dict = {}

    for lic in lics:
        # 전체자명단도 현재 정규화 기준으로 재인덱싱
        vk = norm_vehicle(lic.vehicle_no or "")
        if vk: by_vkey.setdefault(vk, []).append(lic)
        l4 = veh_last4(lic.vehicle_no or "")
        if l4 and len(l4) >= 3: by_l4.setdefault(l4, []).append(lic)
        nk = norm_name(lic.name or "")
        if nk: by_nkey.setdefault(nk, []).append(lic)

    for m in db.query(Member).all():
        # 합계행은 대조 생략
        if is_sum_row(m.name or "", m.vehicle_no or ""):
            m.match_status = "해당없음"
            m.match_fail_reason = "합계행 제외"
            continue

        vkey = norm_vehicle(m.vehicle_no or "")
        l4   = veh_last4(m.vehicle_no or "")
        nkey = norm_name(m.name or "")

        vc = list({x.id: x for x in by_vkey.get(vkey, [])}.values())
        lc = list({x.id: x for x in (by_l4.get(l4, []) if l4 and len(l4) >= 3 else [])}.values())
        nc = list({x.id: x for x in by_nkey.get(nkey, [])}.values())

        best = None; status = "전체자미확인"; fail = ""

        # 1순위: vkey + 성명
        exact_vn = [x for x in vc if norm_name(x.name or "") == nkey]
        # 2순위: 뒷자리 + 성명
        exact_l4n = [x for x in lc if norm_name(x.name or "") == nkey]

        if exact_vn:
            best = exact_vn[0]; status = "정상매칭"
        elif exact_l4n:
            best = exact_l4n[0]; status = "정상매칭"
        elif vkey and len(vc) == 1:
            best = vc[0]; status = "성명확인필요"
        elif l4 and len(l4) >= 3 and len(lc) == 1:
            best = lc[0]; status = "뒷자리매칭"
        elif len(nc) == 1:
            best = nc[0]; status = "차량번호확인필요"
        elif len(nc) > 1:
            fail = f"동명이인 {len(nc)}명"
        elif not (m.vehicle_no or "").strip():
            fail = "차량번호 없음"
        elif l4 and len(lc) > 1:
            fail = f"차량뒷자리 후보 {len(lc)}명"
        else:
            fail = "전체자명단에 없음"

        if best:
            m.match_license_id = best.id; m.match_status = status
            m.match_fail_reason = None
            for attr, lattr in [("mobile","mobile"),("phone","phone"),("address","address"),
                                 ("official_address","official_address"),
                                 ("join_date_raw","join_date_raw"),("permit_date_raw","permit_date_raw"),
                                 ("cert_issue_date_raw","cert_issue_date_raw"),("cert_no","cert_no")]:
                if not getattr(m, attr, None) and getattr(best, lattr, None):
                    setattr(m, attr, getattr(best, lattr))
        else:
            m.match_license_id = None; m.match_status = "전체자미확인"
            m.match_fail_reason = fail
    db.commit()

# ── 전체자 대조 미확인 ────────────────────────────────────────────────────────
@app.get("/license-check", response_class=HTMLResponse)
def license_check(request: Request, db: Session = Depends(get_db),
                  user: User = Depends(require_user)):
    """
    전체자 대조 미확인 화면.
    DB 컬럼/데이터 문제로 화면 전체가 Internal Server Error로 죽지 않도록 안전 처리.
    """
    msg = request.query_params.get("msg", "")
    items = []
    lic_count = 0

    try:
        lic_count = db.query(LicenseRecord).count()
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("license_check lic_count error:", repr(e))
        lic_count = 0
        msg = (msg + " / " if msg else "") + "전체자명단 건수 확인 중 오류"

    try:
        items = (db.query(Member)
                 .filter(_clean_filter(),
                         Member.user_confirmed_match == False,
                         or_(Member.match_license_id == None,
                             Member.match_status == "전체자미확인"))
                 .order_by(Member.region, Member.name)
                 .limit(500)
                 .all())
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("license_check items error:", repr(e))
        items = []
        msg = (msg + " / " if msg else "") + "전체자대조 목록 조회 오류"

    return templates.TemplateResponse(request, "license_check.html", {
        "request": request,
        "user": user,
        "items": items or [],
        "lic_count": lic_count or 0,
        "fmt_amt": fmt_amt,
        "msg": msg,
        "quote": quote,
    })


@app.post("/license-check/{mid}/confirm")
def lic_confirm(mid: int, db: Session = Depends(get_db), user: User = Depends(require_user)):
    try:
        m = db.query(Member).filter(Member.id == mid).first()
        if not m:
            return RedirectResponse("/license-check?msg=회원 없음", status_code=302)
        if hasattr(m, "user_confirmed_match"):
            m.user_confirmed_match = True
        db.add(m)
        db.commit()
        return RedirectResponse("/license-check?msg=확인완료", status_code=302)
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("license confirm error:", repr(e))
        return RedirectResponse("/license-check?msg=확인완료 처리 오류", status_code=302)


@app.post("/license-check/{mid}/link")
def lic_link(mid: int, license_id: int = Form(...),
             db: Session = Depends(get_db), user: User = Depends(require_user)):
    try:
        m = db.query(Member).filter(Member.id == mid).first()
        lic = db.query(LicenseRecord).filter(LicenseRecord.id == license_id).first()
        if not m or not lic:
            return RedirectResponse("/license-check?msg=레코드 없음", status_code=302)

        if hasattr(m, "match_license_id"):
            m.match_license_id = lic.id
        if hasattr(m, "match_status"):
            m.match_status = "수동매칭"
        if hasattr(m, "match_fail_reason"):
            m.match_fail_reason = None
        if hasattr(m, "user_confirmed_match"):
            m.user_confirmed_match = True

        for attr, lattr in [
            ("mobile", "mobile"),
            ("address", "address"),
            ("official_address", "official_address"),
        ]:
            try:
                if hasattr(m, attr) and not getattr(m, attr, None) and getattr(lic, lattr, None):
                    setattr(m, attr, getattr(lic, lattr))
            except Exception:
                pass

        db.add(m)
        db.commit()

        try:
            _full_recalc(db)
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            print("license link recalc error:", repr(e))

        return RedirectResponse(f"/license-check?msg={getattr(m, 'name', '')} 연결완료", status_code=302)
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("license link error:", repr(e))
        return RedirectResponse("/license-check?msg=전체자 연결 처리 오류", status_code=302)


@app.post("/license-check/{mid}/to-work")
def lic_to_work(mid: int, process_type: str = Form("폐업"),
                db: Session = Depends(get_db), user: User = Depends(require_user)):
    try:
        m = db.query(Member).filter(Member.id == mid).first()
        if not m:
            return RedirectResponse("/license-check?msg=회원 없음", status_code=302)

        allowed = ["폐업", "폐지", "양도", "이관", "탈퇴", "사망", "말소"]
        if process_type not in allowed:
            process_type = "폐업"

        w = WorkQueue(
            member_id=mid,
            process_type=process_type,
            status="반영대기",
            arrears_at_submit=getattr(m, "excel_arrears", 0) or 0,
        )

        # 컬럼이 있는 경우에만 안전하게 세팅
        for attr, val in [
            ("source_screen", "전체자대조"),
            ("submitted_by", getattr(user, "id", None)),
            ("note", "전체자대조에서 업무처리 등록"),
        ]:
            try:
                if hasattr(w, attr):
                    setattr(w, attr, val)
            except Exception:
                pass

        db.add(w)
        db.commit()

        try:
            _invalidate_snap(db, "dashboard")
        except Exception:
            pass

        return RedirectResponse("/license-check?msg=처리대기목록에 등록완료", status_code=302)
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("license to-work error:", repr(e))
        return RedirectResponse("/license-check?msg=처리대기 등록 오류", status_code=302)


@app.get("/api/license-search")
def api_lic_search(q: str = "", db: Session = Depends(get_db),
                   user: User = Depends(require_user)):
    try:
        if not q:
            return {"items": []}
        l4 = "".join(c for c in q if c.isdigit())[-4:]
        nk = norm_name(q)
        flt = []
        if nk:
            flt.append(LicenseRecord.name_key.ilike(f"%{nk}%"))
        if l4 and len(l4) >= 3:
            flt.append(LicenseRecord.vehicle_no.ilike(f"%{l4}%"))
        if not flt:
            return {"items": []}
        recs = db.query(LicenseRecord).filter(or_(*flt)).limit(10).all()
        return {"items": [{
            "id": r.id,
            "name": getattr(r, "name", "") or "",
            "vehicle_no": getattr(r, "vehicle_no", "") or "",
            "mobile": getattr(r, "mobile", "") or "",
            "address": (getattr(r, "official_address", "") or getattr(r, "address", "") or "")
        } for r in recs]}
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("api license search error:", repr(e))
        return {"items": [], "error": "검색 중 오류"}



@app.get("/debug/billing-upload-stats")
def debug_billing_upload_stats(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    운영 DB에서 부과대수 업로드가 월/시트별로 들어갔는지 확인하는 임시 진단용.
    """
    from sqlalchemy import text, inspect

    result = {
        "db": "",
        "tables": [],
        "target_table": None,
        "columns": [],
        "total": 0,
        "by_month_sheet_process": [],
        "by_process": [],
        "error": None,
    }

    try:
        bind = db.get_bind()
        dialect = getattr(getattr(bind, "dialect", None), "name", "")
        result["db"] = dialect

        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result["tables"] = [t for t in tables if "billing" in t.lower() or "work" in t.lower()]

        target = None
        for cand in ["billing_people", "billing_persons", "billing_person", "work_queue", "work_queues"]:
            if cand in tables:
                target = cand
                break

        if not target:
            result["error"] = "billing 관련 대상 테이블을 찾지 못했습니다."
            return result

        result["target_table"] = target

        cols = [c["name"] for c in inspector.get_columns(target)]
        result["columns"] = cols

        result["total"] = db.execute(text(f"SELECT COUNT(*) FROM {target}")).scalar() or 0

        source_year_col = "source_year" if "source_year" in cols else "''"
        source_month_col = "source_month" if "source_month" in cols else "''"
        source_sheet_col = "source_sheet" if "source_sheet" in cols else "''"
        process_col = "process_type" if "process_type" in cols else ("account" if "account" in cols else "''")

        rows = db.execute(text(f"""
            SELECT 
                COALESCE(CAST({source_year_col} AS TEXT),'') AS source_year,
                COALESCE(CAST({source_month_col} AS TEXT),'') AS source_month,
                COALESCE(CAST({source_sheet_col} AS TEXT),'') AS source_sheet,
                COALESCE(CAST({process_col} AS TEXT),'') AS process_type,
                COUNT(*) AS cnt
            FROM {target}
            GROUP BY source_year, source_month, source_sheet, process_type
            ORDER BY source_year, source_month, source_sheet, process_type
        """)).mappings().all()

        result["by_month_sheet_process"] = [dict(r) for r in rows]

        rows2 = db.execute(text(f"""
            SELECT COALESCE(CAST({process_col} AS TEXT),'') AS process_type, COUNT(*) AS cnt
            FROM {target}
            GROUP BY process_type
            ORDER BY process_type
        """)).mappings().all()

        result["by_process"] = [dict(r) for r in rows2]

        return result

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        result["error"] = repr(e)
        return result





@app.get("/api/debug/billing-upload-stats")
def api_debug_billing_upload_stats(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text, inspect

    result = {
        "ok": True,
        "db": "",
        "tables": [],
        "target_table": None,
        "columns": [],
        "total": 0,
        "by_month_sheet_process": [],
        "by_process": [],
        "error": None,
    }

    try:
        bind = db.get_bind()
        dialect = getattr(getattr(bind, "dialect", None), "name", "")
        result["db"] = dialect

        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result["tables"] = [t for t in tables if "billing" in t.lower() or "work" in t.lower()]

        target = None
        for cand in [
            "billing_people",
            "billing_persons",
            "billing_person",
            "work_queue",
            "work_queues"
        ]:
            if cand in tables:
                target = cand
                break

        if not target:
            result["ok"] = False
            result["error"] = "billing/work 관련 대상 테이블을 찾지 못했습니다."
            return result

        result["target_table"] = target

        cols = [c["name"] for c in inspector.get_columns(target)]
        result["columns"] = cols

        result["total"] = db.execute(text(f"SELECT COUNT(*) FROM {target}")).scalar() or 0

        source_year_col = "source_year" if "source_year" in cols else "''"
        source_month_col = "source_month" if "source_month" in cols else "''"
        source_sheet_col = "source_sheet" if "source_sheet" in cols else "''"
        process_col = "process_type" if "process_type" in cols else ("account" if "account" in cols else "''")

        rows = db.execute(text(f"""
            SELECT 
                COALESCE(CAST({source_year_col} AS TEXT),'') AS source_year,
                COALESCE(CAST({source_month_col} AS TEXT),'') AS source_month,
                COALESCE(CAST({source_sheet_col} AS TEXT),'') AS source_sheet,
                COALESCE(CAST({process_col} AS TEXT),'') AS process_type,
                COUNT(*) AS cnt
            FROM {target}
            GROUP BY source_year, source_month, source_sheet, process_type
            ORDER BY source_year, source_month, source_sheet, process_type
        """)).mappings().all()

        result["by_month_sheet_process"] = [dict(r) for r in rows]

        rows2 = db.execute(text(f"""
            SELECT COALESCE(CAST({process_col} AS TEXT),'') AS process_type, COUNT(*) AS cnt
            FROM {target}
            GROUP BY process_type
            ORDER BY process_type
        """)).mappings().all()

        result["by_process"] = [dict(r) for r in rows2]

        return result

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        result["ok"] = False
        result["error"] = repr(e)
        return result




# ── 부과대수 관리 ──────────────────────────────────────────────────────────────


# ── 업무처리/부과대수 반영 로직 ─────────────────────────────
def _safe_set(obj, name, value):
    try:
        if hasattr(obj, name):
            setattr(obj, name, value)
            return True
    except Exception:
        pass
    return False


def _append_member_note(member, text):
    try:
        old = member.note or ""
        if text and text not in old:
            member.note = (old + (" / " if old else "") + text)[-1000:]
    except Exception:
        pass


def _add_status_event_safe(db, member, wq, before_status, after_status, user_id):
    """
    MemberStatusEvent 컬럼이 프로젝트마다 달라도 최대한 안전하게 기록.
    실패해도 반영 자체는 막지 않음.
    """
    try:
        cols = set(MemberStatusEvent.__table__.columns.keys())
        kwargs = {}

        if "member_id" in cols:
            kwargs["member_id"] = member.id
        if "event_type" in cols:
            kwargs["event_type"] = getattr(wq, "process_type", "") or ""
        if "from_status" in cols:
            kwargs["from_status"] = before_status or ""
        if "to_status" in cols:
            kwargs["to_status"] = after_status or ""
        if "reason" in cols:
            kwargs["reason"] = getattr(wq, "reason", "") or ""
        if "note" in cols:
            kwargs["note"] = getattr(wq, "note", "") or ""
        if "created_by" in cols:
            kwargs["created_by"] = user_id
        if "source" in cols:
            kwargs["source"] = "업무처리반영"
        if "source_id" in cols:
            kwargs["source_id"] = getattr(wq, "id", None)

        if kwargs:
            db.add(MemberStatusEvent(**kwargs))
    except Exception:
        pass


def _reflect_workqueue_item(db: Session, wq, user: User):
    """
    WorkQueue 반영대기 1건을 실제 Member에 반영.
    반영 전까지는 회원 상태를 바꾸지 않고, 여기서 최종 적용한다.
    """
    if not wq:
        return False, "처리대기 항목 없음"

    member = db.query(Member).filter(Member.id == wq.member_id).first()
    if not member:
        return False, "회원 없음"

    if getattr(wq, "status", "") == "반영완료":
        return True, "이미 반영완료"

    ptype_raw = (getattr(wq, "process_type", "") or "").strip()
    ptype = ptype_raw.replace(" ", "")
    before_status = member.status or ""
    before_account = member.account or ""
    before_arrears = int(member.excel_arrears or 0)

    after_status = before_status
    msg = ""

    # 폐업/폐지/관리비폐지 계열
    if ptype in ("폐업", "폐지"):
        member.status = "폐업"
        member.status_source = "work"
        after_status = "폐업"
        msg = "폐업 처리"

    elif ptype in ("관리비폐지", "관리비폐업"):
        member.status = "관리비폐지"
        member.status_source = "work"
        after_status = "관리비폐지"
        msg = "관리비폐지 처리"

    # 양도
    elif ptype in ("양도", "폐업양도", "폐지양도"):
        member.status = "양도"
        member.status_source = "work"
        after_status = "양도"
        msg = "양도 처리"

    # 이관/타도
    elif ptype in ("이관", "타도", "타도이전", "타도전출"):
        member.status = "이관"
        member.status_source = "work"
        after_status = "이관"
        msg = "이관/타도 처리"

    # 탈퇴
    elif ptype == "탈퇴":
        member.status = "탈퇴"
        member.status_source = "work"
        after_status = "탈퇴"
        msg = "탈퇴 처리"

    # 사망
    elif ptype == "사망":
        member.status = "사망"
        member.status_source = "work"
        after_status = "사망"
        msg = "사망 처리"

    # 말소
    elif ptype == "말소":
        member.status = "말소"
        member.status_source = "work"
        after_status = "말소"
        msg = "말소 처리"

    # 택배신규: 관리비/택배관리 대상으로 편입
    elif ptype == "택배신규":
        member.status = "정상"
        member.status_source = "work"
        member.account = "관"
        after_status = "정상"
        msg = "택배신규/관리비 대상 편입"

    # 70세: 부과대상 제외 아님. 협회비 10,000 → 5,000 감액 대상.
    elif ptype in ("70세", "70세감액", "70세이상"):
        member.status = "정상"
        member.status_source = "work"
        _safe_set(member, "age70", True)
        _append_member_note(member, "70세 감액대상: 협회비 10,000원 → 5,000원")
        after_status = "정상"
        msg = "70세 감액 처리"

    # 현역복구/검증해제
    elif ptype in ("현역복구", "검증해제", "정상복구"):
        member.status = "정상"
        member.status_source = "work"
        after_status = "정상"
        msg = "정상 복구"

    # 금액수정은 기존 별도 라우트가 있으면 그쪽 사용. 여기서는 막음.
    elif ptype == "금액수정":
        return False, "금액수정은 금액수정 반영 버튼을 사용하세요"

    else:
        return False, f"알 수 없는 처리구분: {ptype_raw}"

    # 미수금은 삭제하지 않음. 기존 미수금은 정산용으로 보존.
    # 다만 상태가 정상 외로 바뀌면 _clean_filter 기준에서 부과대상 제외됨.
    try:
        member.is_overpay = (member.excel_arrears or 0) < 0
    except Exception:
        pass

    db.add(member)

    # WorkQueue 완료 처리
    wq.status = "반영완료"
    _safe_set(wq, "reflected_by", user.id)
    _safe_set(wq, "reflected_at", datetime.now())
    _safe_set(wq, "result_note", msg)
    db.add(wq)

    _add_status_event_safe(db, member, wq, before_status, after_status, user.id)

    try:
        add_log(
            db,
            user.id,
            "업무처리반영",
            f"{member.name}/{member.vehicle_no}: {ptype_raw} / 상태 {before_status}→{after_status} / 계정 {before_account}→{member.account} / 미수금 {before_arrears:,}원 보존"
        )
    except Exception:
        pass

    return True, msg


@app.post("/workorder/{wid}/reflect")
def workorder_reflect_api(
    wid: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    work.html의 JS가 호출하는 API:
    fetch('/workorder/' + woId + '/reflect', {method:'POST'})
    """
    try:
        wq = db.query(WorkQueue).filter(WorkQueue.id == wid).first()
        ok, msg = _reflect_workqueue_item(db, wq, user)

        if ok:
            db.commit()
            _invalidate_snap(db, "dashboard")
            return {"ok": True, "msg": msg}

        db.rollback()
        return {"ok": False, "msg": msg}

    except Exception as e:
        db.rollback()
        return {"ok": False, "msg": "반영 오류: " + str(e)[:180]}


@app.post("/workorder/reflect_all")
def workorder_reflect_all_api(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    현재 반영대기 전체 반영.
    금액수정은 별도 반영 버튼이 있으므로 제외.
    """
    try:
        items = (
            db.query(WorkQueue)
            .filter(WorkQueue.status == "반영대기")
            .filter(WorkQueue.process_type != "금액수정")
            .order_by(WorkQueue.id.asc())
            .all()
        )

        ok_count = 0
        fail_count = 0
        messages = []

        for wq in items:
            ok, msg = _reflect_workqueue_item(db, wq, user)
            if ok:
                ok_count += 1
            else:
                fail_count += 1
                messages.append(f"{wq.id}:{msg}")

        db.commit()
        _invalidate_snap(db, "dashboard")

        return {
            "ok": True,
            "msg": f"전체반영 완료 {ok_count}건 / 실패 {fail_count}건" + ((" / " + "; ".join(messages[:5])) if messages else "")
        }

    except Exception as e:
        db.rollback()
        return {"ok": False, "msg": "전체반영 오류: " + str(e)[:180]}


@app.post("/work/{wid}/reflect")
def work_reflect_redirect(
    wid: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    혹시 form 방식 버튼이 있는 경우를 위한 redirect 라우트.
    """
    from urllib.parse import quote

    try:
        wq = db.query(WorkQueue).filter(WorkQueue.id == wid).first()
        ok, msg = _reflect_workqueue_item(db, wq, user)

        if ok:
            db.commit()
            _invalidate_snap(db, "dashboard")
            return RedirectResponse("/work?tab=반영완료&msg=" + quote("반영완료: " + msg), status_code=302)

        db.rollback()
        return RedirectResponse("/work?tab=반영대기&msg=" + quote("반영실패: " + msg), status_code=302)

    except Exception as e:
        db.rollback()
        return RedirectResponse("/work?tab=반영대기&msg=" + quote("반영 오류: " + str(e)[:180]), status_code=302)



@app.post("/admin/fix-billing-pending-latest-only")
def admin_fix_billing_pending_latest_only(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    잘못 들어간 과거월 부과대수 처리대기 정리.
    최신 연월 BillingPerson만 처리대기 유지하고,
    과거월은 부과대수상세로 되돌린다.
    잡수입/가수금 IncomeLedgerDetail은 건드리지 않는다.
    """
    try:
        latest = (
            db.query(BillingPerson.year, BillingPerson.month)
            .order_by(BillingPerson.year.desc(), BillingPerson.month.desc())
            .first()
        )
        if not latest:
            return {"ok": True, "msg": "BillingPerson 없음", "changed": 0}

        latest_y, latest_m = int(latest[0]), int(latest[1])

        changed = (
            db.query(BillingPerson)
            .filter(
                BillingPerson.reflect_status.in_(["처리대기", "반영대기"]),
                or_(BillingPerson.year != latest_y, BillingPerson.month != latest_m)
            )
            .update(
                {BillingPerson.reflect_status: "부과대수상세"},
                synchronize_session=False
            )
        )

        db.commit()

        return {
            "ok": True,
            "latest_year": latest_y,
            "latest_month": latest_m,
            "changed": changed,
            "msg": "과거월 처리대기를 부과대수상세로 정리했습니다."
        }

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        return {"ok": False, "error": repr(e)}





@app.post("/admin/rebuild-latest-billing-pending")
def admin_rebuild_latest_billing_pending(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    부과대수 엑셀 전체월 중 최신월만 처리대기목록에 올린다.
    기존 잡수입/가수금으로 만든 협회비/관리비 후보는 유지하고,
    최신월 엑셀 사람 중 중복 아닌 사람만 처리대기로 만든다.
    """
    def _k(pt, vehicle_no, name):
        pt = str(pt or "").strip()
        pt = PROCESS_NORM.get(pt, pt)
        pt = BILLING_TO_PENDING_PT.get(pt, pt)

        v = str(vehicle_no or "").replace(" ", "").replace("-", "").replace("호", "").strip()
        n = str(name or "").replace(" ", "").strip()
        return (pt, v, n)

    try:
        latest = (
            db.query(BillingPerson.year, BillingPerson.month)
            .order_by(BillingPerson.year.desc(), BillingPerson.month.desc())
            .first()
        )

        if not latest:
            return {"ok": False, "msg": "BillingPerson 자료 없음"}

        latest_y = int(latest[0])
        latest_m = int(latest[1])

        # 1) 전체 부과대수 자료는 우선 상세로 돌림
        db.query(BillingPerson).filter(
            BillingPerson.reflect_status.in_(["처리대기", "반영대기", "부과대수상세"])
        ).update(
            {BillingPerson.reflect_status: "부과대수상세"},
            synchronize_session=False
        )

        # 2) 기존 잡수입/가수금 후보 키 수집
        existing = set()

        try:
            _ensure_income_ledger_details(db)
            details = db.query(IncomeLedgerDetail).filter(
                IncomeLedgerDetail.pending_target.in_(["협회비", "관리비"])
            ).all()

            for d in details:
                existing.add(_k(
                    getattr(d, "pending_target", "") or "",
                    getattr(d, "related_vehicle_no", "") or "",
                    getattr(d, "related_name", "") or "",
                ))
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            print("income duplicate key load error:", e)

        try:
            rows = db.execute(_text("SELECT * FROM bank_income_pending_queue")).mappings().all()
            for r in rows:
                existing.add(_k(
                    r.get("process_type", "") or r.get("income_kind", "") or "",
                    r.get("related_vehicle_no", "") or "",
                    r.get("related_name", "") or "",
                ))
        except Exception:
            pass

        existing = {x for x in existing if x[1] or x[2]}

        # 3) 최신월의 협회가입/택배신규만 처리대기 후보
        latest_people = db.query(BillingPerson).filter(
            BillingPerson.year == latest_y,
            BillingPerson.month == latest_m
        ).all()

        made_pending = 0
        skipped_duplicate = 0
        ignored = 0

        for bp in latest_people:
            raw_pt = str(getattr(bp, "process_type", "") or "")
            mapped_pt = BILLING_TO_PENDING_PT.get(PROCESS_NORM.get(raw_pt, raw_pt), PROCESS_NORM.get(raw_pt, raw_pt))

            # 반영대기 탭에는 협회비/관리비 신규 부과 후보만 올림
            if mapped_pt not in ["협회비", "관리비"]:
                ignored += 1
                continue

            key = _k(mapped_pt, getattr(bp, "vehicle_no", ""), getattr(bp, "name", ""))

            if key in existing:
                bp.reflect_status = "부과대수상세"
                skipped_duplicate += 1
                continue

            bp.reflect_status = "처리대기"
            existing.add(key)
            made_pending += 1

        db.commit()

        return {
            "ok": True,
            "latest_year": latest_y,
            "latest_month": latest_m,
            "made_pending": made_pending,
            "skipped_duplicate": skipped_duplicate,
            "ignored_not_new_billing": ignored,
            "msg": "최신월 부과대수 신규 부과 후보만 처리대기로 재구성했습니다."
        }

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        return {"ok": False, "error": repr(e)}




@app.get("/work", response_class=HTMLResponse)
def work_page(request: Request, tab: str = "전체", q: str = "",
              db: Session = Depends(get_db), user: User = Depends(require_user)):
    TABS = ["전체","반영대기","폐업","양도","이관","탈퇴","사망","말소","현역복구","반영완료","부과반영대기"]
    billing_tab = (tab == "부과반영대기")
    items = []
    billing_items = []

    if billing_tab:
        # BillingPerson 처리대기 항목
        bq = db.query(BillingPerson).filter(BillingPerson.reflect_status == "처리대기")
        if q:
            like = f"%{q}%"
            bq = bq.filter(or_(BillingPerson.name.ilike(like), BillingPerson.vehicle_no.ilike(like)))
        billing_items = bq.order_by(BillingPerson.year.desc(), BillingPerson.month.desc(),
                                     BillingPerson.process_type).limit(500).all()
    else:
        wq = db.query(WorkQueue, Member).join(Member)
        if tab == "반영대기": wq = wq.filter(WorkQueue.status == "반영대기")
        elif tab == "반영완료": wq = wq.filter(WorkQueue.status == "반영완료")
        elif tab != "전체": wq = wq.filter(WorkQueue.process_type == tab, WorkQueue.status == "반영대기")
        if q:
            like = f"%{q}%"
            wq = wq.filter(or_(Member.name.ilike(like), Member.vehicle_no.ilike(like)))
        items = wq.order_by(WorkQueue.submitted_at.desc()).limit(500).all()

    def cnt(t):
        if t == "부과반영대기":
            return db.query(BillingPerson).filter(BillingPerson.reflect_status == "처리대기").count()
        q2 = db.query(WorkQueue)
        if t == "반영대기": return q2.filter(WorkQueue.status == "반영대기").count()
        if t == "반영완료": return q2.filter(WorkQueue.status == "반영완료").count()
        if t == "전체":    return q2.count()
        return q2.filter(WorkQueue.process_type == t, WorkQueue.status == "반영대기").count()
    tab_counts = {t: cnt(t) for t in TABS}

    return templates.TemplateResponse(request, "work.html", {
        "request": request, "user": user,
        "items": items, "billing_items": billing_items,
        "tab": tab, "q": q, "billing_tab": billing_tab,
        "tab_counts": tab_counts, "TABS": TABS, "fmt_amt": fmt_amt,
        "msg": request.query_params.get("msg", ""),
    })

@app.post("/work/{wid}/reflect")
def work_reflect(wid: int, db: Session = Depends(get_db), user: User = Depends(require_user)):
    wq = db.query(WorkQueue).filter(WorkQueue.id == wid).first()
    if not wq: raise HTTPException(404)
    m = db.query(Member).filter(Member.id == wq.member_id).first()
    if not m: raise HTTPException(404)
    m.status = wq.process_type; m.status_source = "manual"
    wq.status = "반영완료"; wq.reflected_by = user.id
    wq.reflected_at = datetime.now()
    db.commit()
    _full_recalc(db)
    add_log(db, user.id, "업무처리반영", f"{m.name}/{m.vehicle_no} → {wq.process_type}")
    return RedirectResponse(f"/work?msg={m.name} {wq.process_type} 반영완료", status_code=302)

@app.post("/work/{wid}/cancel")
def work_cancel(wid: int, db: Session = Depends(get_db), user: User = Depends(require_user)):
    wq = db.query(WorkQueue).filter(WorkQueue.id == wid).first()
    if not wq: raise HTTPException(404)
    wq.status = "취소"; db.commit()
    _invalidate_snap(db, "dashboard")
    return RedirectResponse("/work?msg=취소완료", status_code=302)

# ── 검증필요 ──────────────────────────────────────────────────────────────────
@app.get("/verify", response_class=HTMLResponse)
def verify_page(request: Request, db: Session = Depends(get_db),
                user: User = Depends(require_user)):
    items = (db.query(Member).filter(Member.status == "정상",
                                     Member.arrears_verified == False)
             .order_by(Member.region, Member.name).limit(500).all())
    return templates.TemplateResponse(request, "verify.html", {
        "request": request, "user": user, "items": items, "fmt_amt": fmt_amt,
        "msg": request.query_params.get("msg", ""),
    })

@app.post("/verify/{mid}/clear")
def verify_clear(mid: int, db: Session = Depends(get_db), user: User = Depends(require_user)):
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: raise HTTPException(404)
    m.arrears_verified = True; m.verify_reason = None; db.commit()
    _invalidate_snap(db, "dashboard")
    return RedirectResponse("/verify?msg=검증해제완료", status_code=302)

# ── 문자대상 ──────────────────────────────────────────────────────────────────
def _collection_member_q(db, q="", region="", account="", min_amt=0,
                          contact_filter="", long_term=""):
    """문자대상 공통 쿼리: Member 직접 조회, 필터 적용"""
    from datetime import date as _date
    base = db.query(Member).filter(_clean_filter(), Member.excel_arrears > 0)

    if q:
        like = f"%{q}%"
        base = base.filter(or_(Member.name.ilike(like), Member.vehicle_no.ilike(like)))
    if region:
        base = base.filter(Member.region == region)
    if account:
        acc_map = {
            "협": ["협", "협회비"],
            "관": ["관", "관리비"],
            "택배": ["택배", "택배관리", "택배관리비"],
        }
        base = base.filter(Member.account.in_(acc_map.get(account, [account])))
    if min_amt > 0:
        base = base.filter(Member.excel_arrears >= min_amt)
    if contact_filter == "있음":
        base = base.filter(
            or_(
                and_(Member.mobile != None, Member.mobile != ""),
                and_(Member.phone != None, Member.phone != "")
            )
        )
    elif contact_filter == "없음":
        base = base.filter(
            or_(Member.mobile == None, Member.mobile == ""),
            or_(Member.phone == None, Member.phone == ""),
        )
    if long_term:
        months_map = {"3개월+": 3, "6개월+": 6, "12개월+": 12}
        mo = months_map.get(long_term, 3)
        today = _date.today()
        cutoff_year = today.year - (1 if today.month <= mo else 0)
        cutoff_month = (today.month - mo - 1) % 12 + 1
        cutoff = f"{cutoff_year}-{cutoff_month:02d}"
        base = base.filter(
            Member.first_unpaid_month != None,
            Member.first_unpaid_month != "",
            Member.first_unpaid_month <= cutoff,
        )
    return base


@app.get("/collection", response_class=HTMLResponse)
def collection_page(request: Request, q: str = "", region: str = "",
                    account: str = "", min_amt: int = 0,
                    contact_filter: str = "", long_term: str = "",
                    sort: str = "", page: int = 1,
                    db: Session = Depends(get_db), user: User = Depends(require_user)):
    PAGE_SIZE = 200
    page = max(page, 1)

    base_q = _collection_member_q(db, q, region, account, min_amt, contact_filter, long_term)

    if sort == "amount_desc":
        base_q = base_q.order_by(Member.excel_arrears.desc())
    elif sort == "amount_asc":
        base_q = base_q.order_by(Member.excel_arrears.asc())
    else:
        base_q = base_q.order_by(Member.region, Member.name)

    total_count = base_q.order_by(None).count()
    total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(1, min(page, total_pages))
    items = base_q.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
    total = base_q.order_by(None).with_entities(func.sum(Member.excel_arrears)).scalar() or 0

    regions = sorted({x[0] for x in db.query(Member.region).distinct().filter(Member.region != None).all() if x[0]})

    return templates.TemplateResponse(request, "collection.html", {
        "request": request, "user": user, "items": items,
        "q": q, "region": region, "account": account,
        "min_amt": min_amt, "contact_filter": contact_filter,
        "long_term": long_term, "sort": sort,
        "total": int(total), "regions": regions,
        "page": page, "total_pages": total_pages, "total_count": total_count,
        "fmt_amt": fmt_amt, "fmt_acc": fmt_acc,
        "msg": request.query_params.get("msg", ""),
    })


@app.get("/collection/extract-phones")
def extract_phones(request: Request,
                   q: str = "", region: str = "", account: str = "",
                   min_amt: int = 0, contact_filter: str = "", long_term: str = "",
                   fmt: str = "tsv",
                   db: Session = Depends(get_db), user: User = Depends(require_user)):
    from fastapi.responses import StreamingResponse
    members = (_collection_member_q(db, q, region, account, min_amt, contact_filter, long_term)
               .order_by(Member.region, Member.name).all())
    sep = "," if fmt == "csv" else "\t"
    ext = "csv" if fmt == "csv" else "tsv"
    rows = [sep.join(["지역", "계정", "차량번호", "성명", "번호", "미수금"]) + "\n"]
    for m in members:
        mob = phone_clean(m.mobile or "")
        pho = phone_clean(m.phone or "")
        number = mob or pho or "연락처없음"
        rows.append(sep.join([
            m.region or "", fmt_acc(m.account),
            m.vehicle_no or "", m.name or "",
            number, str(m.excel_arrears or 0),
        ]) + "\n")
    add_log(db, user.id, "번호추출", f"{len(members)}명")
    content = "".join(rows).encode("utf-8-sig")
    return StreamingResponse(
        iter([content]),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=phones.{ext}"},
    )


@app.get("/collection/extract-addresses")
def extract_addresses(request: Request,
                      q: str = "", region: str = "", account: str = "",
                      min_amt: int = 0, contact_filter: str = "", long_term: str = "",
                      fmt: str = "tsv",
                      db: Session = Depends(get_db), user: User = Depends(require_user)):
    from fastapi.responses import StreamingResponse
    members = (_collection_member_q(db, q, region, account, min_amt, contact_filter, long_term)
               .order_by(Member.region, Member.name).all())
    sep = "," if fmt == "csv" else "\t"
    ext = "csv" if fmt == "csv" else "tsv"
    rows = [sep.join(["지역", "계정", "차량번호", "성명", "주소", "미수금"]) + "\n"]
    for m in members:
        addr = address_clean(m) or "주소없음"
        rows.append(sep.join([
            m.region or "", fmt_acc(m.account),
            m.vehicle_no or "", m.name or "",
            addr, str(m.excel_arrears or 0),
        ]) + "\n")
    add_log(db, user.id, "주소추출", f"{len(members)}명")
    content = "".join(rows).encode("utf-8-sig")
    return StreamingResponse(
        iter([content]),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=addresses.{ext}"},
    )

@app.post("/collection/generate")
def collection_generate(db: Session = Depends(get_db), user: User = Depends(require_user)):
    db.query(CollectionTarget).delete(synchronize_session=False); db.commit()
    members = db.query(Member).filter(
        Member.status == "정상",
        Member.excel_arrears > 0,
        Member.is_overpay == False,
    ).all()
    count = 0
    for m in members:
        # 자동이체는 전용통장 업로드 후에만 확정 → 현재는 일반으로 처리
        no_mobile = not (m.mobile or "").strip()
        no_address = not (m.official_address or m.address or "").strip()
        excluded = False
        exc_reason = ""
        # 음수(초과납부) 제외
        if (m.excel_arrears or 0) < 0:
            excluded = True; exc_reason = "초과납부/선납"
        # 검증필요 제외
        elif m.arrears_verified == False and m.verify_reason:
            excluded = True; exc_reason = "검증필요"
        mob_clean = phone_clean(m.mobile or "")
        addr = address_clean(m)
        cat = "연락처없음" if no_mobile else "일반"
        now_ = datetime.now()
        db.add(CollectionTarget(
            member_id=m.id, generated_by=user.id,
            base_year=now_.year, base_month=now_.month,
            arrears=m.excel_arrears, category=cat,
            excluded=excluded, exclude_reason=exc_reason,
            mobile_clean=mob_clean, address_clean=addr,
        ))
        count += 1
    db.commit()
    _invalidate_snap(db, "dashboard")
    add_log(db, user.id, "문자대상생성", f"{count}명")
    return RedirectResponse(f"/collection?msg={count}명 생성완료", status_code=302)

# ── 통장매칭 ──────────────────────────────────────────────────────────────────


# ── 자동매칭 전체반영 ─────────────────────────────────────

@app.get("/bank/{tid}/match", response_class=HTMLResponse)
def bank_manual_match_page(
    tid: int,
    request: Request,
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    import re as _re
    from sqlalchemy import or_, func

    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        raise HTTPException(404)

    def _compact(v):
        return _re.sub(r"\\s+", "", str(v or "")).strip()

    def _digits(v):
        return _re.sub(r"\\D", "", str(v or ""))

    def _last4(v):
        d = _digits(v)
        return d[-4:] if len(d) >= 4 else d

    search_text = q or getattr(tx, "memo", "") or ""
    search_compact = _compact(search_text)
    search_digits = _digits(search_text)
    search_last4 = search_digits[-4:] if len(search_digits) >= 4 else search_digits

    stop_words = {
        "??","??","??","??","??","???","??",
        "??","??","??","??","??","??","??","???"
    }

    name_tokens = [
        x for x in _re.findall(r"[?-?]{2,5}", search_compact)
        if x not in stop_words
    ]

    base = (
        db.query(Member)
        .filter(Member.name != None, Member.name != "")
        .filter(Member.vehicle_no != None, Member.vehicle_no != "")
    )

    try:
        base = base.filter(~Member.name_key.in_(list(SUM_NAMES_DB)))
        base = base.filter(~Member.vehicle_no.in_(list(SUM_NAMES_DB)))
    except Exception:
        pass

    filters = []

    if q:
        like = f"%{q}%"
        filters.append(Member.name.ilike(like))
        filters.append(Member.vehicle_no.ilike(like))

    for nt in name_tokens:
        filters.append(func.replace(Member.name, " ", "").ilike(f"%{nt}%"))

    if search_last4 and len(search_last4) >= 3:
        filters.append(Member.vehicle_no.ilike(f"%{search_last4}%"))

    if filters:
        base = base.filter(or_(*filters))

    raw_members = base.limit(500).all()

    def _score(m):
        score = 0
        nm = _compact(m.name)
        veh = str(m.vehicle_no or "")
        l4 = _last4(veh)

        if nm and nm in search_compact:
            score += 80
        if l4 and l4 in search_digits:
            score += 60
        if nm and l4 and nm in search_compact and l4 in search_digits:
            score += 200
        return score

    members = sorted(
        raw_members,
        key=lambda m: (-_score(m), m.region or "", m.name or "", m.vehicle_no or "")
    )[:100]

    return templates.TemplateResponse(request, "bank_manual_match.html", {
        "request": request,
        "user": user,
        "tx": tx,
        "members": members,
        "q": q,
        "fmt_amt": fmt_amt,
        "fmt_acc": fmt_acc,
    })

@app.post("/bank/{tid}/manual-match")
def bank_manual_match_save(
    tid: int,
    member_id: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    수동매칭 즉시 반영.
    핵심: 같은 회원/같은 년월의 기존 원장이 있으면 새 줄을 만들지 않고 그 줄에 입금액을 반영한다.
    """
    from urllib.parse import quote
    from datetime import datetime
    import re as _re

    def _num(v):
        try:
            if v is None:
                return 0
            if isinstance(v, (int, float)):
                return int(v)
            t = str(v).replace(",", "").replace("원", "").strip()
            nums = _re.findall(r"\d[\d,]*", t)
            if not nums:
                return 0
            return int(nums[0].replace(",", ""))
        except Exception:
            return 0

    def _get_amount(tx):
        for attr in ["amount", "deposit_amount", "in_amount", "paid_amount", "txn_amount", "money"]:
            if hasattr(tx, attr):
                v = _num(getattr(tx, attr))
                if v:
                    return abs(v)

        for attr in ["memo", "raw_data", "description"]:
            if hasattr(tx, attr):
                t = str(getattr(tx, attr) or "")
                nums = [_num(x) for x in _re.findall(r"\d[\d,]*", t)]
                nums = [x for x in nums if x > 0]
                if nums:
                    return max(nums)
        return 0

    def _get_date(tx):
        raw_date = getattr(tx, "txn_date", None) or getattr(tx, "date", None) or getattr(tx, "paid_date", None)
        y = datetime.now().year
        m = datetime.now().month
        pdate = ""
        if raw_date:
            try:
                if hasattr(raw_date, "year"):
                    y = raw_date.year
                    m = raw_date.month
                    pdate = raw_date.isoformat()[:10]
                else:
                    dt = datetime.fromisoformat(str(raw_date)[:10])
                    y = dt.year
                    m = dt.month
                    pdate = dt.date().isoformat()
            except Exception:
                pdate = str(raw_date)[:10]
        return y, m, pdate

    try:
        tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
        member = db.query(Member).filter(Member.id == member_id).first()

        if not tx or not member:
            return RedirectResponse("/bank?status=미매칭&msg=" + quote("수동매칭 대상 없음"), status_code=302)

        amount = _get_amount(tx)
        if amount <= 0:
            tx.matched_member_id = member.id
            tx.match_status = "확인필요"
            tx.match_reason = f"수동매칭 금액확인필요: {member.name}/{member.vehicle_no}"
            db.add(tx)
            db.commit()
            return RedirectResponse("/bank?status=확인필요&msg=" + quote("금액을 읽지 못해 확인필요로 이동했습니다."), status_code=302)

        year, month, paid_date = _get_date(tx)
        ledger_cols = set(MonthlyLedger.__table__.columns.keys())

        # 1) 같은 회원/같은 월의 기존 원장 찾기
        ledger = (
            db.query(MonthlyLedger)
            .filter(MonthlyLedger.member_id == member.id)
            .filter(MonthlyLedger.year == year)
            .filter(MonthlyLedger.month == month)
            .order_by(MonthlyLedger.charge_amount.desc(), MonthlyLedger.id.asc())
            .first()
        )

        # 2) 기존 원장이 있으면 그 줄에 입금액 반영
        if ledger:
            old_paid = int(ledger.paid_amount or 0)
            ledger.paid_amount = old_paid + amount
            ledger.paid_date = paid_date

            charge = int(ledger.charge_amount or 0)
            carry = int(ledger.carry_over or 0)
            total_due = carry + charge
            ledger.arrears_amount = max(total_due - int(ledger.paid_amount or 0), 0)
            if hasattr(ledger, "calc_arrears"):
                ledger.calc_arrears = ledger.arrears_amount

            # 추적용 메모
            if hasattr(ledger, "raw_note"):
                old_note = ledger.raw_note or ""
                add_note = f" / 통장반영:{getattr(tx, 'memo', '') or ''}"
                ledger.raw_note = (old_note + add_note)[-500:]

            db.add(ledger)

        # 3) 기존 원장이 없을 때만 새 원장 생성
        else:
            kwargs = {
                "member_id": member.id,
                "batch_id": None,
                "source_file": "통장반영",
                "source_sheet": "수동매칭",
                "source_row": tx.id,
                "raw_vehicle_no": member.vehicle_no or "",
                "raw_name": member.name or "",
                "raw_region": member.region or "",
                "raw_account": member.account or "",
                "raw_note": getattr(tx, "memo", "") or "",
                "year": year,
                "month": month,
                "carry_over": 0,
                "charge_amount": 0,
                "paid_amount": amount,
                "arrears_amount": 0,
                "paid_date": paid_date,
                "calc_arrears": 0,
            }
            kwargs = {k: v for k, v in kwargs.items() if k in ledger_cols}
            db.add(MonthlyLedger(**kwargs))

        # 4) 통장거래는 바로 반영완료
        tx.matched_member_id = member.id
        tx.match_status = "반영완료"
        tx.match_reason = f"수동매칭 즉시반영: {member.name}/{member.vehicle_no} / {amount}원"
        tx.applied = True
        db.add(tx)

        # 5) 회원 미수금은 원장 기준으로 재계산
        db.flush()
        try:
            _recalc_member(db, member)
        except Exception:
            # 재계산 실패 시 최소한 현재 미수금에서 차감
            before = int(member.excel_arrears or 0)
            member.excel_arrears = before - amount
            member.calc_arrears = member.excel_arrears
            member.is_overpay = member.excel_arrears < 0
            db.add(member)

        db.commit()
        _invalidate_snap(db, "dashboard")

        try:
            add_log(db, user.id, "통장수동매칭반영", f"{getattr(tx, 'memo', '')} → {member.name}/{member.vehicle_no} / {amount}원")
        except Exception:
            pass

        return RedirectResponse(
            "/bank?status=반영완료&msg=" + quote(f"수동매칭 및 반영완료: {member.name} {amount:,}원"),
            status_code=302
        )

    except Exception as e:
        db.rollback()
        return RedirectResponse(
            "/bank?status=미매칭&msg=" + quote("수동매칭 반영 오류: " + str(e)[:180]),
            status_code=302
        )


@app.get("/sms/export")
def sms_export_excel(
    export_type: str = "phone",
    region: str = "",
    account: str = "",
    status: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from io import BytesIO
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(Member).filter(_arrears_full_filter())

    if region:
        q = q.filter(Member.region == region)

    if account:
        if account in ("관리비", "관"):
            q = q.filter(Member.account.in_(["관", "관리비"]))
        elif account in ("협회비", "협"):
            q = q.filter(Member.account.in_(["협", "협회비"]))
        else:
            q = q.filter(Member.account == account)

    if status:
        q = q.filter(Member.status == status)

    if export_type == "address":
        q = q.filter(or_(Member.official_address != None, Member.address != None))
    else:
        q = q.filter(or_(Member.mobile != None, Member.phone != None))

    members = q.order_by(Member.region, Member.account, Member.name, Member.vehicle_no).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "주소추출" if export_type == "address" else "번호추출"
    ws.append(["지역", "계정", "차량번호", "성명", "연락처", "주소", "현재미수금", "상태"])

    for m in members:
        contact = (m.mobile or "").strip() or (m.phone or "").strip()
        addr = (m.official_address or "").strip() or (m.address or "").strip()

        if export_type != "address" and not contact:
            continue
        if export_type == "address" and not addr:
            continue

        acc = m.account or ""
        if acc == "관":
            acc = "관리비"
        elif acc == "협":
            acc = "협회비"

        ws.append([
            m.region or "",
            acc,
            m.vehicle_no or "",
            m.name or "",
            contact,
            addr,
            int(m.excel_arrears or 0),
            m.status or "",
        ])

    header_fill = PatternFill("solid", fgColor="FCE7F3")
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True, color="9D174D")
                cell.fill = header_fill

    widths = [14, 12, 18, 16, 18, 42, 14, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 7).number_format = '#,##0'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    title = "주소추출" if export_type == "address" else "번호추출"
    filename = quote(f"문자관리_{title}.xlsx")

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )




# ── 중복 월 원장 정리 ─────────────────────────────────────
@app.post("/admin/merge-monthly-ledgers")
def admin_merge_monthly_ledgers(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    같은 회원 + 같은 연도 + 같은 월 원장이 여러 줄로 쪼개진 경우 한 줄로 합친다.
    예:
    5월 부과 5,000 / 입금 0 / 미수 5,000
    5월 부과 0 / 입금 5,000 / 미수 0
    → 5월 부과 5,000 / 입금 5,000 / 미수 0
    """
    from urllib.parse import quote
    from sqlalchemy import func

    try:
        groups = (
            db.query(
                MonthlyLedger.member_id,
                MonthlyLedger.year,
                MonthlyLedger.month,
                func.count(MonthlyLedger.id).label("cnt")
            )
            .filter(MonthlyLedger.member_id != None)
            .filter(MonthlyLedger.year != None)
            .filter(MonthlyLedger.month != None)
            .group_by(MonthlyLedger.member_id, MonthlyLedger.year, MonthlyLedger.month)
            .having(func.count(MonthlyLedger.id) > 1)
            .all()
        )

        merged_groups = 0
        deleted_rows = 0
        touched_members = set()

        for member_id, year, month, cnt in groups:
            rows = (
                db.query(MonthlyLedger)
                .filter(MonthlyLedger.member_id == member_id)
                .filter(MonthlyLedger.year == year)
                .filter(MonthlyLedger.month == month)
                .order_by(
                    MonthlyLedger.charge_amount.desc(),
                    MonthlyLedger.carry_over.desc(),
                    MonthlyLedger.id.asc()
                )
                .all()
            )

            if len(rows) <= 1:
                continue

            # 기준행: 부과금/이월금이 있는 행 우선, 없으면 첫 행
            base = rows[0]

            total_carry = 0
            total_charge = 0
            total_paid = 0
            paid_dates = []
            notes = []

            for r in rows:
                total_carry += int(r.carry_over or 0)
                total_charge += int(r.charge_amount or 0)
                total_paid += int(r.paid_amount or 0)

                if getattr(r, "paid_date", None):
                    paid_dates.append(str(r.paid_date)[:10])

                if getattr(r, "raw_note", None):
                    notes.append(str(r.raw_note))

            # 부과금/이월금은 중복으로 합산하면 안 되는 경우가 있어 보정
            # 같은 월 원장 중 부과금이 있는 행이 하나면 그 값을 사용.
            charge_values = [int(r.charge_amount or 0) for r in rows if int(r.charge_amount or 0) > 0]
            carry_values = [int(r.carry_over or 0) for r in rows if int(r.carry_over or 0) > 0]

            if charge_values:
                total_charge = max(charge_values)
            if carry_values:
                total_carry = max(carry_values)

            total_due = total_carry + total_charge
            arrears = max(total_due - total_paid, 0)

            base.carry_over = total_carry
            base.charge_amount = total_charge
            base.paid_amount = total_paid
            base.arrears_amount = arrears

            if hasattr(base, "calc_arrears"):
                base.calc_arrears = arrears

            if paid_dates:
                base.paid_date = sorted(paid_dates)[-1]

            if hasattr(base, "raw_note"):
                joined = " / ".join(dict.fromkeys([n for n in notes if n]))
                base.raw_note = joined[-500:]

            db.add(base)

            for r in rows[1:]:
                db.delete(r)
                deleted_rows += 1

            touched_members.add(member_id)
            merged_groups += 1

        db.flush()

        # 회원별 미수금 재계산
        for mid in touched_members:
            m = db.query(Member).filter(Member.id == mid).first()
            if m:
                try:
                    _recalc_member(db, m)
                except Exception:
                    pass

        db.commit()
        _invalidate_snap(db, "dashboard")

        try:
            add_log(db, user.id, "중복월원장정리", f"{merged_groups}개 월 정리, {deleted_rows}행 삭제")
        except Exception:
            pass

        msg = quote(f"중복 월 원장 정리 완료: {merged_groups}개 월 정리, {deleted_rows}행 삭제")
        return RedirectResponse("/settings?msg=" + msg, status_code=302)

    except Exception as e:
        db.rollback()
        msg = quote("중복 월 원장 정리 오류: " + str(e)[:180])
        return RedirectResponse("/settings?msg=" + msg, status_code=302)


@app.get("/bank", response_class=HTMLResponse)
def bank_page(request: Request, status: str = "", q: str = "",
              db: Session = Depends(get_db), user: User = Depends(require_user)):

    # SAFE DEFAULT BANK STATUS TABS
    # /bank ??? ? status_tabs? ??? ????? ???? 500 ?? ?? ??
    status_tabs = [
        ("", "전체"),
        ("자동매칭", "자동매칭"),
        ("확인필요", "확인필요"),
        ("미매칭", "미매칭"),
        ("반영완료", "반영완료"),
    ]

    bq = db.query(BankTransaction)
    if status: bq = bq.filter(BankTransaction.match_status == status)
    if q:
        like = f"%{q}%"
        bq = bq.filter(or_(BankTransaction.memo.ilike(like)))
    txs = bq.order_by(BankTransaction.id.desc()).limit(500).all()
    import json as _json
    for tx in txs:
        try: tx._candidates = _json.loads(tx.match_candidates_json) if tx.match_candidates_json else []
        except: tx._candidates = []

    _count_keys = ["자동매칭", "확인필요", "미매칭", "반영완료"]
    counts = {s: db.query(BankTransaction).filter(BankTransaction.match_status == s).count() for s in _count_keys}
    counts["전체"] = db.query(BankTransaction).count()

    return templates.TemplateResponse(request, "bank.html", {
        "request": request, "user": user, "txs": txs, "status": status, "q": q,
        "counts": counts, "status_tabs": status_tabs, "fmt_amt": fmt_amt,
        "msg": request.query_params.get("msg", ""),
    })

@app.post("/bank/paste")
def bank_paste(request: Request, pasted_text: str = Form(""),
               db: Session = Depends(get_db), user: User = Depends(require_user)):
    count = _parse_and_match_bank_lines(db, pasted_text.splitlines(), "paste", None)
    add_log(db, user.id, "통장붙여넣기", f"{count}건")
    return RedirectResponse(f"/bank?status=자동매칭&msg={count}건 파싱/매칭", status_code=302)

def _parse_bank_line(line: str) -> Optional[dict]:
    """
    통장 한 줄 파싱
    컬럼 순서: 날짜 [시간] 입금액 [출금액] 잔액 거래구분 메모
    핵심: 2026같은 4자리 연도를 금액으로 오인하지 않도록 처리
    """
    raw = line.rstrip()
    if not raw.strip(): return None
    if "\t" in raw: parts = [p.strip() for p in raw.split("\t")]
    else: parts = re.split(r"\s{2,}", raw.strip())
    parts = [p for p in parts if p]
    if not parts: return None

    # 1) 날짜 추출
    tx_date = ""
    date_idx = -1
    for i, p in enumerate(parts):
        m = re.match(r"^(20\d{2}[-/.]\d{1,2}[-/.]\d{1,2})", p)
        if m:
            tx_date = m.group(1).replace(".", "-").replace("/", "-")
            date_idx = i; break
    remaining_parts = [p for i, p in enumerate(parts) if i != date_idx] if date_idx >= 0 else parts[:]

    # 2) 금액 컬럼 추출 (쉼표 포함 or 5자리 이상 숫자)
    #    쉼표 없는 4자리 이하 순수 숫자(2026, 1234 등)는 제외
    def is_amount_token(s):
        plain = s.replace(",", "").replace(".", "")
        if not plain.isdigit(): return False
        if "," in s: return True          # 1,000 형태
        if len(plain) >= 5: return True    # 10000 이상
        return False

    amounts = []
    text_parts = []
    for p in remaining_parts:
        if is_amount_token(p):
            try:
                a = parse_amount(p, allow_negative=False)
                if a >= 1000: amounts.append((a, p)); continue
            except: pass
        text_parts.append(p)

    # 컬럼 순서: 첫 번째가 입금액, 두 번째가 잔액 (일반적)
    # 단, 금액이 1개만 있으면 그게 잔액일 수 있으므로 주의
    if len(amounts) >= 2:
        # 첫 번째가 잔액보다 작으면 입금액, 아니면 역순 가능
        deposit = amounts[0][0]
        balance = amounts[-1][0]
    elif len(amounts) == 1:
        # 금액 1개 = 잔액만 있거나 입금액만 있음
        # 잔액이 더 크면 잔액으로 간주 (deposit=0)
        deposit = amounts[0][0]
        balance = 0
    else:
        deposit = balance = 0

    # 3) 나머지에서 거래구분/메모 추출
    tokens = [t for t in text_parts if t and not re.match(r"^\d{1,4}$", t)]
    method = tokens[0] if tokens else ""
    memo   = " ".join(tokens[1:]) if len(tokens) > 1 else ""

    return {"txn_date": tx_date, "deposit_amount": deposit, "balance_amount": balance,
            "method": method, "memo": memo, "raw_text": line}

def _parse_and_match_bank_lines(db: Session, lines, source_type: str, batch_id) -> int:
    count = 0
    for line in lines:
        parsed = _parse_bank_line(str(line))
        if not parsed or parsed["deposit_amount"] <= 0: continue
        if is_useless_memo(parsed["memo"]): continue
        tx = BankTransaction(
            txn_date=parsed["txn_date"], deposit_amount=parsed["deposit_amount"],
            balance_amount=parsed["balance_amount"], method=parsed["method"],
            memo=parsed["memo"], source_type=source_type,
            batch_id=batch_id, raw_text=parsed["raw_text"],
        )
        _do_bank_match(db, tx)
        count += 1
    return count

def _do_bank_match(db: Session, tx: BankTransaction):
    import json as _json
    memo = tx.memo or ""; deposit = tx.deposit_amount or 0
    name_c, vkey, l4 = extract_memo_keys(memo)

    flt = []
    if name_c and len(name_c) >= 2:
        flt.append(func.replace(Member.name," ","").ilike(f"%{norm_name(name_c)}%"))
    if vkey and len(vkey) >= 5: flt.append(Member.vehicle_key == vkey)
    if l4 and len(l4) >= 3: flt.append(Member.vehicle_no.ilike(f"%{l4}%"))

    if not flt:
        tx.match_status = "미매칭"; db.add(tx); db.commit(); db.refresh(tx); return tx

    raw_members = db.query(Member).filter(Member.status == "정상", or_(*flt)).limit(500).all()
    seen: dict = {}
    for m in raw_members:
        sc, rs, tier = score_bank_match(memo, deposit, m)
        if sc <= 0: continue
        if m.id not in seen or sc > seen[m.id][0]:
            seen[m.id] = (sc, rs, tier, {
                "id":m.id,"name":m.name,"vehicle_no":m.vehicle_no,
                "vehicle_last4":veh_last4(m.vehicle_no),
                "region":m.region or "","account":fmt_acc(m.account),
                "amount":m.excel_arrears or 0,"mobile":m.mobile or "",
                "score":sc,"reason":rs,"tier":tier,
            })

    if not seen:
        tx.match_status = "미매칭"; db.add(tx); db.commit(); db.refresh(tx); return tx

    ranked = sorted(seen.values(), key=lambda x: x[0], reverse=True)
    candidates = [x[3] for x in ranked]
    best = ranked[0]
    best_id, best_sc, best_tier = best[3]["id"], best[0], best[2]
    best_member = db.query(Member).filter(Member.id == best_id).first()
    n = len(candidates)
    reasons_str = best[1]

    # ── 자동매칭 판정 ──────────────────────────────────────────────────────────
    # 조건 A: 성명 정확일치 + 차량번호 전체일치 → 후보 수 무관 자동
    cond_A = best_tier == "exact" and "차량번호전체일치" in reasons_str and "성명일치" in reasons_str
    # 조건 B: 성명 정확일치 + 차량번호 뒷4자리 일치 → 후보 수 무관 자동
    cond_B = best_tier == "exact" and "차량번호뒷자리일치" in reasons_str and "성명일치" in reasons_str
    # 조건 C: 성명 정확일치 + 금액=미수금 일치 + 후보 1명
    cond_C = "성명일치" in reasons_str and "금액=미수금일치" in reasons_str and n == 1
    # 조건 D: 성명 정확일치 + 후보 1명 (점수 80점 이상)
    cond_D = best_tier == "name_exact" and n == 1 and best_sc >= 80
    # 조건 E: 차량번호 전체일치 + 후보 1명
    cond_E = "차량번호전체일치" in reasons_str and n == 1
    # 조건 F: 차량번호 뒷4자리 + 금액=미수금 일치 + 후보 1명
    cond_F = "차량번호뒷자리일치" in reasons_str and "금액=미수금일치" in reasons_str and n == 1

    # 동명이인(후보 2명 이상이고 성명만 일치) → 절대 자동매칭 금지
    only_name = reasons_str == "성명일치" or (best_tier == "name_exact" and "차량번호" not in reasons_str and "금액=미수금일치" not in reasons_str)
    no_auto_dupe = only_name and n >= 2

    auto_match = (cond_A or cond_B or cond_C or cond_D or cond_E or cond_F) and not no_auto_dupe
    status = "자동매칭" if auto_match else "확인필요"

    tx.matched_member_id = best_id if status == "자동매칭" else None
    tx.match_score = best_sc; tx.match_reason = best[1]
    tx.match_status = status
    tx.match_candidates_json = _json.dumps(candidates, ensure_ascii=False)
    db.add(tx); db.commit(); db.refresh(tx)
    return tx

@app.post("/bank/{txid}/apply")
def bank_apply(txid: int, member_id: Optional[int] = None,
               note: str = Form(""),
               db: Session = Depends(get_db), user: User = Depends(require_user)):
    tx = db.query(BankTransaction).filter(BankTransaction.id == txid).first()
    if not tx: raise HTTPException(404)
    if tx.applied:
        return RedirectResponse("/bank?status=자동매칭&msg=이미 반영된 건입니다", status_code=302)
    mid = member_id or tx.matched_member_id
    if not mid: return RedirectResponse("/bank?status=자동매칭&msg=대상자를 선택해주세요", status_code=302)
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: return RedirectResponse("/bank?status=자동매칭&msg=회원을 찾을 수 없습니다", status_code=302)

    amt = tx.deposit_amount or 0
    before = m.excel_arrears or 0
    after  = before - amt
    overpay = max(0, -after)  # 0 미만이면 초과납부

    # 1. excel_arrears 직접 차감
    m.excel_arrears = after
    m.is_overpay = after < 0
    m.last_paid_date = tx.txn_date or m.last_paid_date

    # 2. monthly_ledgers에 입금 기록 추가 (최신월 원장에 paid_amount 반영)
    if tx.txn_date:
        try:
            tx_year  = int(tx.txn_date[:4])
            tx_month = int(tx.txn_date[5:7])
            # 최신월 원장 찾아서 입금액 추가
            existing_ledger = (db.query(MonthlyLedger)
                .filter(MonthlyLedger.member_id == mid,
                        MonthlyLedger.year == tx_year,
                        MonthlyLedger.month == tx_month)
                .order_by(MonthlyLedger.id.desc()).first())
            if existing_ledger:
                existing_ledger.paid_amount = (existing_ledger.paid_amount or 0) + amt
                existing_ledger.arrears_amount = after
                existing_ledger.paid_date = tx.txn_date
            else:
                # 해당월 원장 없으면 새로 생성
                latest_l = (db.query(MonthlyLedger)
                    .filter(MonthlyLedger.member_id == mid)
                    .order_by(MonthlyLedger.year.desc(), MonthlyLedger.month.desc())
                    .first())
                db.add(MonthlyLedger(
                    member_id=mid, batch_id=None,
                    source_file="통장매칭", source_sheet="bank", source_row=tx.id,
                    year=tx_year, month=tx_month,
                    carry_over=before, charge_amount=0,
                    paid_amount=amt, arrears_amount=after,
                    paid_date=tx.txn_date, calc_arrears=after,
                    verified=True,
                ))
        except Exception as _e:
            pass  # 날짜 파싱 실패 시 무시

    # 3. BankTransaction 상태 업데이트
    tx.applied = True
    tx.applied_amount = amt
    tx.overpay_amount = overpay
    tx.match_status = "반영완료"
    tx.matched_member_id = mid
    tx.applied_at = datetime.now()
    tx.note = note
    db.commit()

    # 4. 해당 회원 재계산 (검증필요, 초과납부, 문자대상 갱신)
    _recalc_member(db, m)
    db.commit()
    _invalidate_snap(db, "dashboard")

    add_log(db, user.id, "통장반영",
            f"{m.name}/{m.vehicle_no}: {before:,}→{after:,}원 (입금 {amt:,}원, 날짜 {tx.txn_date})")
    return RedirectResponse(f"/bank?status=자동매칭&msg={m.name} 반영완료 ({before:,}→{after:,}원)", status_code=302)

@app.get("/bank/{txid}/preview")
def bank_preview(txid: int, member_id: Optional[int] = None,
                 db: Session = Depends(get_db), user: User = Depends(require_user)):
    tx = db.query(BankTransaction).filter(BankTransaction.id == txid).first()
    if not tx: return JSONResponse({"ok": False, "msg": "거래 없음"})
    mid = member_id or tx.matched_member_id
    if not mid: return JSONResponse({"ok": False, "msg": "대상자 없음"})
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: return JSONResponse({"ok": False, "msg": "회원 없음"})
    amt = tx.deposit_amount or 0
    before = m.excel_arrears or 0
    after = before - amt
    return JSONResponse({"ok":True,"member_id":m.id,"name":m.name,"vehicle_no":m.vehicle_no,
        "account":fmt_acc(m.account),"deposit":amt,"before":before,"after":after,
        "overpay":max(0,amt-before)})

@app.post("/bank/{txid}/hold")
def bank_hold(txid: int, db: Session = Depends(get_db), user: User = Depends(require_user)):
    tx = db.query(BankTransaction).filter(BankTransaction.id == txid).first()
    if not tx: raise HTTPException(404)
    tx.match_status = "보류"; db.commit()
    return RedirectResponse("/bank?status=자동매칭&msg=보류처리", status_code=302)


# ── 통장매칭 초기화 ─────────────────────────────────────────────

@app.post("/bank/reset-unapplied")
def bank_reset_unapplied(db: Session = Depends(get_db), user: User = Depends(require_user)):
    """미반영 통장내역만 삭제. applied False/NULL 대상. 반영완료 건 보호."""
    from urllib.parse import quote

    try:
        unapplied = or_(BankTransaction.applied == False, BankTransaction.applied == None)
        cnt = db.query(BankTransaction).filter(unapplied).count()
        db.query(BankTransaction).filter(unapplied).delete(synchronize_session=False)
        db.commit()
        _invalidate_snap(db, "dashboard")

        try:
            add_log(db, user.id, "통장초기화-미반영삭제", f"{cnt}건 삭제")
        except Exception:
            pass

        return RedirectResponse("/bank?status=자동매칭&msg=" + quote(f"미반영 통장내역 {cnt}건 삭제완료"), status_code=302)

    except Exception as e:
        db.rollback()
        return RedirectResponse("/bank?status=자동매칭&msg=" + quote("초기화 오류: " + str(e)[:100]), status_code=302)



@app.post("/bank/reset-all")
def bank_reset_all(include_applied: str = Form(""),
                   db: Session = Depends(get_db), user: User = Depends(require_user)):
    """
    통장내역 전체 초기화.
    기본: 반영완료 보호, 미반영만 삭제.
    체크 시: 반영완료 포함 전체 삭제.
    """
    from urllib.parse import quote

    try:
        include_all = str(include_applied).lower() in ("yes", "on", "true", "1")

        if include_all:
            cnt = db.query(BankTransaction).count()
            db.query(BankTransaction).delete(synchronize_session=False)
            scope = "반영완료 포함 전체"
        else:
            unapplied = or_(BankTransaction.applied == False, BankTransaction.applied == None)
            cnt = db.query(BankTransaction).filter(unapplied).count()
            db.query(BankTransaction).filter(unapplied).delete(synchronize_session=False)
            scope = "미반영"

        db.commit()
        _invalidate_snap(db, "dashboard")

        try:
            add_log(db, user.id, "통장초기화-전체", f"{scope} {cnt}건 삭제")
        except Exception:
            pass

        return RedirectResponse("/bank?status=자동매칭&msg=" + quote(f"{scope} 통장내역 {cnt}건 삭제완료"), status_code=302)

    except Exception as e:
        db.rollback()
        return RedirectResponse("/bank?status=자동매칭&msg=" + quote("전체초기화 오류: " + str(e)[:100]), status_code=302)





# ── 설정 ─────────────────────────────────────────────────────
@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request,
                  db: Session = Depends(get_db),
                  user: User = Depends(require_user)):
    return templates.TemplateResponse(request, "settings.html", {
        "request": request,
        "user": user,
    })


@app.post("/admin/reset")
def admin_reset(db: Session = Depends(get_db), user: User = Depends(require_user)):
    from urllib.parse import quote
    try:
        before_members = db.query(Member).count()
        before_bank = db.query(BankTransaction).count()
        deleted_total = 0
        for table in reversed(Base.metadata.sorted_tables):
            if table.name in {"users", "user"}:
                continue
            result = db.execute(table.delete())
            if result.rowcount and result.rowcount > 0:
                deleted_total += result.rowcount
        db.commit()
        try: _invalidate_snap(db, "dashboard")
        except Exception: pass
        msg = f"전체 초기화 완료: 회원 {before_members}명, 통장 {before_bank}건 포함 총 {deleted_total}건 삭제"
        return RedirectResponse("/settings?msg=" + quote(msg), status_code=302)
    except Exception as e:
        db.rollback()
        return RedirectResponse("/settings?msg=" + quote("전체초기화 오류: " + str(e)[:180]), status_code=302)


@app.get("/member/new", response_class=HTMLResponse)
def add_member_page(request: Request, db: Session = Depends(get_db),
                    user: User = Depends(require_user)):
    regions = sorted({x[0] for x in db.query(Member.region).distinct().filter(Member.region != None).all() if x[0]})
    return templates.TemplateResponse(request, "add_member.html",
        {"request": request, "user": user, "regions": regions})

@app.post("/member/new")
def add_member_save(request: Request,
    name: str = Form(""), vehicle_no: str = Form(""), region: str = Form(""),
    account: str = Form("협"), mobile: str = Form(""),
    amount: str = Form("0"), address: str = Form(""), note: str = Form(""),
    db: Session = Depends(get_db), user: User = Depends(require_user)):
    m = Member(name=name, name_key=norm_name(name),
               vehicle_no=vehicle_no, vehicle_key=norm_vehicle(vehicle_no) if vehicle_no else "",
               region=normalize_region(region), account=account,
               mobile=phone_clean(mobile) or mobile,
               address=address, note=note,
               excel_arrears=parse_amount(amount),
               calc_arrears=parse_amount(amount),
               status="정상", status_source="manual",
               source_file="명단추가")
    db.add(m); db.commit()
    _full_recalc(db)
    add_log(db, user.id, "회원추가", f"{name}/{vehicle_no}")
    return RedirectResponse(f"/member/{m.id}", status_code=302)

@app.get("/api/member-search")
def api_member_search(q: str = "", db: Session = Depends(get_db),
                      user: User = Depends(require_user)):
    if not q: return {"items": []}
    l4 = "".join(c for c in q if c.isdigit())[-4:]
    nk = norm_name(q)
    flt = []
    if nk and len(nk) >= 2: flt.append(func.replace(Member.name," ","").ilike(f"%{nk}%"))
    if l4 and len(l4) >= 3: flt.append(Member.vehicle_no.ilike(f"%{l4}%"))
    if not flt: return {"items": []}
    members = db.query(Member).filter(Member.status == "정상", or_(*flt)).limit(10).all()
    return {"items": [{"id":m.id,"name":m.name,"vehicle_no":m.vehicle_no or "",
        "amount":m.excel_arrears or 0} for m in members]}


@app.get("/bank/auto-analysis", response_class=HTMLResponse)
def bank_auto_analysis(request: Request, db: Session = Depends(get_db),
                       user: User = Depends(require_user)):
    """자동이체 전용통장 1년치 분석: 동일 금액 반복 입금 패턴 감지"""
    import json as _json
    from datetime import date, timedelta

    # 최근 1년치 통장 내역
    txs = (db.query(BankTransaction)
           .filter(BankTransaction.is_auto_account == True)
           .order_by(BankTransaction.txn_date, BankTransaction.memo)
           .all())

    if not txs:
        txs = db.query(BankTransaction).order_by(BankTransaction.txn_date).all()

    # 메모별 입금패턴 분석
    from collections import defaultdict
    memo_groups: dict = defaultdict(list)
    for tx in txs:
        key = tx.memo or "(메모없음)"
        memo_groups[key].append({
            "date": tx.txn_date,
            "amount": tx.deposit_amount or 0,
            "applied": tx.applied,
        })

    # 자동이체/정기납부 후보 분류
    results = []
    for memo, entries in memo_groups.items():
        if len(entries) < 2: continue
        amounts = [e["amount"] for e in entries]
        unique_amts = set(amounts)
        valid_amts = unique_amts & AUTOPAY_AMOUNTS
        if not valid_amts and max(amounts, default=0) > 60000: continue
        pat = classify_autopay([(e["amount"], e["date"] or "") for e in entries])
        if pat == "해당없음": continue
        results.append({
            "memo": memo,
            "pattern": pat,
            "count": len(entries),
            "amounts": sorted(valid_amts or unique_amts),
            "last_date": max(e["date"] or "" for e in entries),
            "entries": entries[-6:],  # 최근 6건만 표시
        })

    results.sort(key=lambda x: (-x["count"], x["memo"]))

    return templates.TemplateResponse(request, "bank_auto_analysis.html", {
        "request": request, "user": user, "results": results,
        "total_txs": len(txs), "fmt_amt": fmt_amt,
        "msg": request.query_params.get("msg", ""),
    })


@app.post("/settings/cleanup-sumrows")
def cleanup_sumrows(db: Session = Depends(get_db), user: User = Depends(require_user)):
    """기존 DB에 잘못 저장된 합계행 삭제"""
    from core import is_sum_row
    members = db.query(Member).all()
    deleted = 0
    for m in members:
        if is_sum_row(m.name or "", m.vehicle_no or ""):
            # 연관 데이터 삭제
            db.query(MonthlyLedger).filter(MonthlyLedger.member_id == m.id).delete(synchronize_session=False)
            db.query(MemberStatusEvent).filter(MemberStatusEvent.member_id == m.id).delete(synchronize_session=False)
            db.query(WorkQueue).filter(WorkQueue.member_id == m.id).delete(synchronize_session=False)
            db.delete(m)
            deleted += 1
    db.commit()
    _invalidate_snap(db, "dashboard")
    add_log(db, user.id, "합계행정리", f"{deleted}건 삭제")
    return RedirectResponse(f"/settings?msg=합계행 {deleted}건 정리완료", status_code=302)


@app.post("/member/{mid}/amount-adjust")
def member_amount_adjust(mid: int,
    adjust_amount: int = Form(0),
    adjust_reason: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db), user: User = Depends(require_user)):
    """금액수정 → 부과대수 반영대기로 등록"""
    m = db.query(Member).filter(Member.id == mid).first()
    if not m: raise HTTPException(404)
    before = m.excel_arrears or 0
    db.add(WorkQueue(
        member_id=mid, process_type="금액수정", status="반영대기",
        source_screen="미수금명단",
        reason=f"수정 전:{before:,}원 → 수정 후:{adjust_amount:,}원 / 사유:{adjust_reason}",
        note=note,
        arrears_at_submit=before,
        submitted_by=user.id,
    ))
    db.commit()
    _invalidate_snap(db, "dashboard")
    add_log(db, user.id, "금액수정요청", f"{m.name}: {before:,}→{adjust_amount:,}원")
    return RedirectResponse(f"/member/{mid}?msg=금액수정 반영대기 등록완료", status_code=302)

@app.post("/work/{wid}/reflect-amount")
def work_reflect_amount(wid: int, db: Session = Depends(get_db),
                        user: User = Depends(require_user)):
    """금액수정 반영대기 → 실제 반영"""
    wq = db.query(WorkQueue).filter(WorkQueue.id == wid).first()
    if not wq or wq.process_type != "금액수정": raise HTTPException(404)
    m = db.query(Member).filter(Member.id == wq.member_id).first()
    if not m: raise HTTPException(404)
    # reason에서 수정 후 금액 파싱
    import re
    match = re.search(r"수정 후:(-?[\d,]+)원", wq.reason or "")
    if match:
        new_amt = int(match.group(1).replace(",",""))
        before = m.excel_arrears or 0
        m.excel_arrears = new_amt
        m.is_overpay = new_amt < 0
        wq.status = "반영완료"
        wq.reflected_by = user.id
        wq.reflected_at = datetime.now()
        db.commit()
        _recalc_member(db, m)
        db.commit()
        _invalidate_snap(db, "dashboard")
        add_log(db, user.id, "금액수정반영", f"{m.name}: {before:,}→{new_amt:,}원")
        return RedirectResponse(f"/work?msg=금액수정 반영완료", status_code=302)
    return RedirectResponse(f"/work?msg=금액 파싱 실패", status_code=302)



@app.get("/debug/billing-upload-stats")
def debug_billing_upload_stats(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    운영 DB에서 부과대수 업로드가 월/시트별로 들어갔는지 확인하는 임시 진단용.
    """
    from sqlalchemy import text, inspect

    result = {
        "db": "",
        "tables": [],
        "target_table": None,
        "columns": [],
        "total": 0,
        "by_month_sheet_process": [],
        "by_process": [],
        "error": None,
    }

    try:
        bind = db.get_bind()
        dialect = getattr(getattr(bind, "dialect", None), "name", "")
        result["db"] = dialect

        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result["tables"] = [t for t in tables if "billing" in t.lower() or "work" in t.lower()]

        target = None
        for cand in ["billing_people", "billing_persons", "billing_person", "work_queue", "work_queues"]:
            if cand in tables:
                target = cand
                break

        if not target:
            result["error"] = "billing 관련 대상 테이블을 찾지 못했습니다."
            return result

        result["target_table"] = target

        cols = [c["name"] for c in inspector.get_columns(target)]
        result["columns"] = cols

        result["total"] = db.execute(text(f"SELECT COUNT(*) FROM {target}")).scalar() or 0

        source_year_col = "source_year" if "source_year" in cols else "''"
        source_month_col = "source_month" if "source_month" in cols else "''"
        source_sheet_col = "source_sheet" if "source_sheet" in cols else "''"
        process_col = "process_type" if "process_type" in cols else ("account" if "account" in cols else "''")

        rows = db.execute(text(f"""
            SELECT 
                COALESCE(CAST({source_year_col} AS TEXT),'') AS source_year,
                COALESCE(CAST({source_month_col} AS TEXT),'') AS source_month,
                COALESCE(CAST({source_sheet_col} AS TEXT),'') AS source_sheet,
                COALESCE(CAST({process_col} AS TEXT),'') AS process_type,
                COUNT(*) AS cnt
            FROM {target}
            GROUP BY source_year, source_month, source_sheet, process_type
            ORDER BY source_year, source_month, source_sheet, process_type
        """)).mappings().all()

        result["by_month_sheet_process"] = [dict(r) for r in rows]

        rows2 = db.execute(text(f"""
            SELECT COALESCE(CAST({process_col} AS TEXT),'') AS process_type, COUNT(*) AS cnt
            FROM {target}
            GROUP BY process_type
            ORDER BY process_type
        """)).mappings().all()

        result["by_process"] = [dict(r) for r in rows2]

        return result

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        result["error"] = repr(e)
        return result





@app.get("/api/debug/billing-upload-stats")
def api_debug_billing_upload_stats(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text, inspect

    result = {
        "ok": True,
        "db": "",
        "tables": [],
        "target_table": None,
        "columns": [],
        "total": 0,
        "by_month_sheet_process": [],
        "by_process": [],
        "error": None,
    }

    try:
        bind = db.get_bind()
        dialect = getattr(getattr(bind, "dialect", None), "name", "")
        result["db"] = dialect

        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result["tables"] = [t for t in tables if "billing" in t.lower() or "work" in t.lower()]

        target = None
        for cand in [
            "billing_people",
            "billing_persons",
            "billing_person",
            "work_queue",
            "work_queues"
        ]:
            if cand in tables:
                target = cand
                break

        if not target:
            result["ok"] = False
            result["error"] = "billing/work 관련 대상 테이블을 찾지 못했습니다."
            return result

        result["target_table"] = target

        cols = [c["name"] for c in inspector.get_columns(target)]
        result["columns"] = cols

        result["total"] = db.execute(text(f"SELECT COUNT(*) FROM {target}")).scalar() or 0

        source_year_col = "source_year" if "source_year" in cols else "''"
        source_month_col = "source_month" if "source_month" in cols else "''"
        source_sheet_col = "source_sheet" if "source_sheet" in cols else "''"
        process_col = "process_type" if "process_type" in cols else ("account" if "account" in cols else "''")

        rows = db.execute(text(f"""
            SELECT 
                COALESCE(CAST({source_year_col} AS TEXT),'') AS source_year,
                COALESCE(CAST({source_month_col} AS TEXT),'') AS source_month,
                COALESCE(CAST({source_sheet_col} AS TEXT),'') AS source_sheet,
                COALESCE(CAST({process_col} AS TEXT),'') AS process_type,
                COUNT(*) AS cnt
            FROM {target}
            GROUP BY source_year, source_month, source_sheet, process_type
            ORDER BY source_year, source_month, source_sheet, process_type
        """)).mappings().all()

        result["by_month_sheet_process"] = [dict(r) for r in rows]

        rows2 = db.execute(text(f"""
            SELECT COALESCE(CAST({process_col} AS TEXT),'') AS process_type, COUNT(*) AS cnt
            FROM {target}
            GROUP BY process_type
            ORDER BY process_type
        """)).mappings().all()

        result["by_process"] = [dict(r) for r in rows2]

        return result

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        result["ok"] = False
        result["error"] = repr(e)
        return result




# ── 부과대수 관리 ──────────────────────────────────────────────────────────────



@app.get("/debug/billing-upload-stats")
def debug_billing_upload_stats(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    운영 DB에서 부과대수 업로드가 월/시트별로 들어갔는지 확인하는 임시 진단용.
    """
    from sqlalchemy import text, inspect

    result = {
        "db": "",
        "tables": [],
        "target_table": None,
        "columns": [],
        "total": 0,
        "by_month_sheet_process": [],
        "by_process": [],
        "error": None,
    }

    try:
        bind = db.get_bind()
        dialect = getattr(getattr(bind, "dialect", None), "name", "")
        result["db"] = dialect

        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result["tables"] = [t for t in tables if "billing" in t.lower() or "work" in t.lower()]

        target = None
        for cand in ["billing_people", "billing_persons", "billing_person", "work_queue", "work_queues"]:
            if cand in tables:
                target = cand
                break

        if not target:
            result["error"] = "billing 관련 대상 테이블을 찾지 못했습니다."
            return result

        result["target_table"] = target

        cols = [c["name"] for c in inspector.get_columns(target)]
        result["columns"] = cols

        result["total"] = db.execute(text(f"SELECT COUNT(*) FROM {target}")).scalar() or 0

        source_year_col = "source_year" if "source_year" in cols else "''"
        source_month_col = "source_month" if "source_month" in cols else "''"
        source_sheet_col = "source_sheet" if "source_sheet" in cols else "''"
        process_col = "process_type" if "process_type" in cols else ("account" if "account" in cols else "''")

        rows = db.execute(text(f"""
            SELECT 
                COALESCE(CAST({source_year_col} AS TEXT),'') AS source_year,
                COALESCE(CAST({source_month_col} AS TEXT),'') AS source_month,
                COALESCE(CAST({source_sheet_col} AS TEXT),'') AS source_sheet,
                COALESCE(CAST({process_col} AS TEXT),'') AS process_type,
                COUNT(*) AS cnt
            FROM {target}
            GROUP BY source_year, source_month, source_sheet, process_type
            ORDER BY source_year, source_month, source_sheet, process_type
        """)).mappings().all()

        result["by_month_sheet_process"] = [dict(r) for r in rows]

        rows2 = db.execute(text(f"""
            SELECT COALESCE(CAST({process_col} AS TEXT),'') AS process_type, COUNT(*) AS cnt
            FROM {target}
            GROUP BY process_type
            ORDER BY process_type
        """)).mappings().all()

        result["by_process"] = [dict(r) for r in rows2]

        return result

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        result["error"] = repr(e)
        return result





@app.get("/api/debug/billing-upload-stats")
def api_debug_billing_upload_stats(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text, inspect

    result = {
        "ok": True,
        "db": "",
        "tables": [],
        "target_table": None,
        "columns": [],
        "total": 0,
        "by_month_sheet_process": [],
        "by_process": [],
        "error": None,
    }

    try:
        bind = db.get_bind()
        dialect = getattr(getattr(bind, "dialect", None), "name", "")
        result["db"] = dialect

        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result["tables"] = [t for t in tables if "billing" in t.lower() or "work" in t.lower()]

        target = None
        for cand in [
            "billing_people",
            "billing_persons",
            "billing_person",
            "work_queue",
            "work_queues"
        ]:
            if cand in tables:
                target = cand
                break

        if not target:
            result["ok"] = False
            result["error"] = "billing/work 관련 대상 테이블을 찾지 못했습니다."
            return result

        result["target_table"] = target

        cols = [c["name"] for c in inspector.get_columns(target)]
        result["columns"] = cols

        result["total"] = db.execute(text(f"SELECT COUNT(*) FROM {target}")).scalar() or 0

        source_year_col = "source_year" if "source_year" in cols else "''"
        source_month_col = "source_month" if "source_month" in cols else "''"
        source_sheet_col = "source_sheet" if "source_sheet" in cols else "''"
        process_col = "process_type" if "process_type" in cols else ("account" if "account" in cols else "''")

        rows = db.execute(text(f"""
            SELECT 
                COALESCE(CAST({source_year_col} AS TEXT),'') AS source_year,
                COALESCE(CAST({source_month_col} AS TEXT),'') AS source_month,
                COALESCE(CAST({source_sheet_col} AS TEXT),'') AS source_sheet,
                COALESCE(CAST({process_col} AS TEXT),'') AS process_type,
                COUNT(*) AS cnt
            FROM {target}
            GROUP BY source_year, source_month, source_sheet, process_type
            ORDER BY source_year, source_month, source_sheet, process_type
        """)).mappings().all()

        result["by_month_sheet_process"] = [dict(r) for r in rows]

        rows2 = db.execute(text(f"""
            SELECT COALESCE(CAST({process_col} AS TEXT),'') AS process_type, COUNT(*) AS cnt
            FROM {target}
            GROUP BY process_type
            ORDER BY process_type
        """)).mappings().all()

        result["by_process"] = [dict(r) for r in rows2]

        return result

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        result["ok"] = False
        result["error"] = repr(e)
        return result




# ── 부과대수 관리 ─────────────────────────────────────────────
BILLING_COUNT_ITEMS = ["협회가입", "양도", "타도", "폐지", "탈퇴", "택배신규", "관리비폐지", "70세"]

def _billing_count_norm_text(v):
    if v is None:
        return ""
    t = str(v).strip()
    if t.lower() == "nan":
        return ""
    return t.replace("\n", " ").replace("\r", " ").strip()

def _billing_count_int(v):
    import re
    try:
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return int(v)
        t = str(v).replace(",", "").replace("대", "").replace("명", "").strip()
        m = re.search(r"-?\d+", t)
        return int(m.group(0)) if m else 0
    except Exception:
        return 0

def _billing_count_item(v):
    t = _billing_count_norm_text(v).replace(" ", "")
    if not t:
        return ""
    if "협회가입원" in t or "협회가입" in t:
        return "협회가입"
    if "택배신규" in t:
        return "택배신규"
    if "관리비폐지" in t or "관리비폐업" in t:
        return "관리비폐지"
    if "70세" in t or "70세이상" in t:
        return "70세"
    if "양도" in t:
        return "양도"
    if "타도" in t:
        return "타도"
    if "탈퇴" in t:
        return "탈퇴"
    if "폐지" in t or "폐업" in t:
        return "폐지"
    return ""

def _billing_count_sheet_year_month(sheet_name, default_year=None):
    import re
    now = datetime.now()
    year = int(default_year or now.year)
    month = now.month

    m = re.search(r"(\d{1,2})월", str(sheet_name))
    if m:
        month = int(m.group(1))

    y = re.search(r"\((\d{2,4})년\)", str(sheet_name))
    if y:
        yy = int(y.group(1))
        year = 2000 + yy if yy < 100 else yy

    return year, month

def _billing_count_parse_xlsx(path, default_year=None):
    """
    [사용]부과대수 엑셀 파서.
    2025년 6월~2026년 6월 전체 시트 대응.
    - 최신 항목 컬럼형
    - 과거 구역 제목형
    둘 다 읽어서 BillingReport/BillingPerson용 데이터로 반환한다.
    """
    import re
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    all_details = []
    summary_by_month = {}

    sheet_names = [ws.title for ws in wb.worksheets]
    skip_sheet_names = set()

    # 3월부과(1)이 있으면 일반 3월부과는 중복 방지를 위해 스킵
    if "3월부과(1)" in sheet_names and "3월부과" in sheet_names:
        skip_sheet_names.add("3월부과")

    header_aliases = {
        "region": ["지역", "시군", "시·군"],
        "vehicle_no": ["차량번호", "자동차등록번호", "등록번호"],
        "name": ["성명", "이름", "대표자"],
        "resident_no": ["주민번호", "주민등록번호", "생년월일"],
        "address": ["주소", "소재지"],
        "permit_date": ["인가일자", "허가일자"],
        "join_date": ["가입일자", "협회가입일"],
        "cert_issue_date": ["자격증명발급일자", "자격증명 발급일자", "자격증명일자"],
        "cert_no": ["자격증명발급번호", "자격증명 발급번호", "자격증명번호"],
        "amount": ["입금액", "금액"],
        "item": ["항목", "구분", "처리구분"],
        "note": ["비고", "메모"],
    }

    regions = [
        "춘천시","강릉시","원주시","동해시","태백시","속초시","삼척시",
        "홍천군","횡성군","영월군","평창군","정선군","철원군","화천군",
        "양구군","인제군","고성군","양양군"
    ]

    vehicle_re = re.compile(r"(?:[가-힣]{2,3})?\s*\d{2,3}\s*[가-힣]\s*\d{3,4}\s*호?")

    def cell_text(v):
        return _billing_count_norm_text(v)

    def clean_vehicle(v):
        t = cell_text(v)
        return t.replace(" ", "").replace("-", "")

    def is_noise(t):
        t2 = cell_text(t).replace(" ", "")
        if not t2:
            return True
        bad = [
            "합계","총계","소계","누계","구분","항목","비고","성명","차량번호",
            "지역","인가일자","가입일자","주소","주민","부과차량","폐지차량",
            "관리비","협회비","양도양수","자료","대수","월부과"
        ]
        return any(x in t2 for x in bad)

    def section_item(text):
        t = cell_text(text).replace(" ", "")
        if not t:
            return ""

        if "70세" in t:
            return "70세"

        if "관리비폐지" in t or "관리비폐업" in t:
            return "관리비폐지"

        if "택배관리비폐지" in t:
            return "관리비폐지"

        if "관리비" in t and ("폐지" in t or "폐업" in t):
            return "관리비폐지"

        if "택배신규" in t:
            return "택배신규"

        if ("택배" in t and "관리비" in t and ("부과" in t or "신규" in t)) or "관리비부과차량" in t:
            return "택배신규"

        if "협회가입원" in t or "협회가입" in t:
            return "협회가입"

        if "협회비" in t and ("부과" in t or "가입" in t or "신규" in t):
            return "협회가입"

        if "양도양수" in t or "양도" in t:
            return "양도"

        if "타도" in t or "이관" in t or "전입" in t:
            return "타도"

        if "탈퇴" in t:
            return "탈퇴"

        if "폐지" in t or "폐업" in t:
            return "폐지"

        return _billing_count_item(t)

    def find_header(ws):
        max_scan = min(ws.max_row, 60)

        for r in range(1, max_scan + 1):
            vals = [cell_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            joined = " ".join(vals)
            score = 0
            if "지역" in joined or "시군" in joined:
                score += 1
            if "차량번호" in joined or "등록번호" in joined:
                score += 1
            if "성명" in joined or "대표자" in joined or "이름" in joined:
                score += 1
            if "항목" in joined or "처리구분" in joined or "구분" in joined:
                score += 1

            if score >= 2:
                colmap = {}
                for idx, txt in enumerate(vals, start=1):
                    key_txt = txt.replace(" ", "")
                    for key, aliases in header_aliases.items():
                        for a in aliases:
                            if a.replace(" ", "") in key_txt:
                                colmap[key] = idx
                return r, colmap

        return None, {}

    def get(row, colmap, key):
        c = colmap.get(key)
        if not c:
            return ""
        if c - 1 >= len(row):
            return ""
        return cell_text(row[c - 1])

    def infer_from_row(vals):
        texts = [cell_text(v) for v in vals]
        joined = " ".join(texts)

        region = ""
        for t in texts:
            t2 = t.strip()
            if t2 in regions:
                region = t2
                break
            if any(rg in t2 for rg in regions):
                for rg in regions:
                    if rg in t2:
                        region = rg
                        break
            if region:
                break

        vehicle_no = ""
        m = vehicle_re.search(joined)
        if m:
            vehicle_no = clean_vehicle(m.group(0))

        name = ""
        # 차량번호 근처 또는 한국 이름처럼 보이는 값 추정
        for t in texts:
            tt = t.strip()
            compact = tt.replace(" ", "")
            if not compact:
                continue
            if compact == region:
                continue
            if vehicle_no and vehicle_no.replace("호", "") in compact.replace(" ", ""):
                continue
            if is_noise(compact):
                continue
            if re.fullmatch(r"[가-힣]{2,5}", compact):
                name = compact
                break

        return region, vehicle_no, name

    def add_detail(year, month, item, region, vehicle_no, name, row, colmap, sheet_name, source_row):
        if not item:
            return

        vehicle_no = clean_vehicle(vehicle_no)
        name = cell_text(name)
        region = cell_text(region)

        joined = " ".join(cell_text(x) for x in row)
        if any(x in joined for x in ["합계", "총계", "소계"]):
            return

        if not vehicle_no and not name:
            return

        ym = (year, month)
        if ym not in summary_by_month:
            summary_by_month[ym] = {k: 0 for k in BILLING_COUNT_ITEMS}
            summary_by_month[ym]["협회기본대수"] = 0
            summary_by_month[ym]["총부과대수"] = 0
            summary_by_month[ym]["택배관리"] = 0

        summary_by_month[ym][item] = summary_by_month[ym].get(item, 0) + 1

        detail = {
            "year": year,
            "month": month,
            "item": item,
            "region": region,
            "vehicle_no": vehicle_no,
            "name": name,
            "resident_no": get(row, colmap, "resident_no"),
            "address": get(row, colmap, "address"),
            "permit_date": get(row, colmap, "permit_date"),
            "join_date": get(row, colmap, "join_date"),
            "cert_issue_date": get(row, colmap, "cert_issue_date"),
            "cert_no": get(row, colmap, "cert_no"),
            "amount": _billing_count_int(get(row, colmap, "amount")),
            "note": get(row, colmap, "note") or joined,
            "source_sheet": sheet_name,
            "source_row": source_row,
            "raw_data": {str(i + 1): cell_text(v) for i, v in enumerate(row)},
        }
        all_details.append(detail)

    seen_keys = set()

    for ws in wb.worksheets:
        sheet_name = ws.title

        if sheet_name in skip_sheet_names:
            continue

        year, month = _billing_count_sheet_year_month(sheet_name, default_year)
        ym = (year, month)

        if ym not in summary_by_month:
            summary_by_month[ym] = {k: 0 for k in BILLING_COUNT_ITEMS}
            summary_by_month[ym]["협회기본대수"] = 0
            summary_by_month[ym]["총부과대수"] = 0
            summary_by_month[ym]["택배관리"] = 0

        header_row, colmap = find_header(ws)
        current_section = ""

        # 1) 표/구역 혼합 파싱
        start_row = header_row + 1 if header_row else 1

        for r in range(start_row, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            vals = [cell_text(x) for x in row]
            row_text = " ".join(x for x in vals if x)

            if not row_text:
                continue

            inferred_region, inferred_vehicle, inferred_name = infer_from_row(vals)
            has_person_like = bool(inferred_vehicle or inferred_name)

            sec = section_item(row_text)
            # 구역 제목은 사람 행이 아닐 때만 current_section으로 잡음
            if sec and not has_person_like:
                current_section = sec
                continue

            item = ""
            if colmap:
                item = section_item(get(row, colmap, "item"))
            if not item:
                item = current_section
            if not item:
                item = section_item(row_text)

            if not item:
                continue

            region = get(row, colmap, "region") if colmap else ""
            vehicle_no = get(row, colmap, "vehicle_no") if colmap else ""
            name = get(row, colmap, "name") if colmap else ""

            if not region:
                region = inferred_region
            if not vehicle_no:
                vehicle_no = inferred_vehicle
            if not name:
                name = inferred_name

            key = (year, month, item, clean_vehicle(vehicle_no), cell_text(name))
            if key in seen_keys:
                continue
            seen_keys.add(key)

            add_detail(year, month, item, region, vehicle_no, name, row, colmap, sheet_name, r)

        # 2) 요약문 보조 파싱
        # 중요:
        # 최신 시트는 총부과대수/n월 택배관리 문구가 사람 행 오른쪽 끝에 붙어 있는 경우가 있어
        # 앞 10칸만 보면 놓치므로 전체 열을 다 훑는다.
        for r in range(1, ws.max_row + 1):
            vals = [cell_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            text = " ".join(x for x in vals if x)

            if not text:
                continue

            nums = [int(x.replace(",", "")) for x in re.findall(r"\d[\d,]*", text)]

            if "70세" in text and nums:
                summary_by_month[ym]["70세"] = max(summary_by_month[ym].get("70세", 0), nums[-1])

            # 예:
            # 협회비 총부과대수 1,095 + 168 = 1,263
            # 총부과대수 1262
            # 총부과대수 1,100 + 139 = 1,239
            if "총부과대수" in text and nums:
                total_val = nums[-1]
                summary_by_month[ym]["총부과대수"] = max(summary_by_month[ym].get("총부과대수", 0), total_val)

                # 협회기본대수는 별도 값이 없으면 총부과대수와 같은 값으로 우선 표시
                # 예전처럼 DB 전체 회원수 3198로 고정되는 것보다 엑셀 월별 값이 맞음
                if not summary_by_month[ym].get("협회기본대수"):
                    summary_by_month[ym]["협회기본대수"] = total_val

            # 예:
            # 6월 택배관리 1,894 - 29 + 26 = 1,891
            # n월 택배관리 1981
            if "택배" in text and "관리" in text and nums:
                summary_by_month[ym]["택배관리"] = max(summary_by_month[ym].get("택배관리", 0), nums[-1])

    return summary_by_month, all_details




@app.get("/billing-counts", response_class=HTMLResponse)
def billing_counts_page(
    request: Request,
    year: int = None,
    month: int = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    now = datetime.now()
    y = int(year or now.year)
    m = int(month or now.month)

    reports = (
        db.query(BillingReport)
        .order_by(BillingReport.year.desc(), BillingReport.month.desc())
        .limit(24)
        .all()
    )

    current = (
        db.query(BillingReport)
        .filter(BillingReport.year == y, BillingReport.month == m)
        .first()
    )

    persons = []
    if current:
        persons = (
            db.query(BillingPerson)
            .filter(BillingPerson.billing_report_id == current.id)
            .order_by(BillingPerson.process_type, BillingPerson.region, BillingPerson.name)
            .limit(500)
            .all()
        )

    return templates.TemplateResponse(request, "billing_counts.html", {
        "request": request,
        "user": user,
        "year": y,
        "month": m,
        "reports": reports,
        "current": current,
        "persons": persons,
        "fmt_amt": fmt_amt,
        "msg": request.query_params.get("msg", ""),
    })


@app.post("/billing-counts/upload")
async def billing_counts_upload(
    file: UploadFile = File(...),
    year: int = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    import tempfile, os, json
    from urllib.parse import quote

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    try:
        with open(tmp_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        summary_by_month, details = _billing_count_parse_xlsx(tmp_path, year)

        saved_reports = 0
        saved_details = 0

        # 여러 월 시트가 들어있는 부과대수 파일은 최신월만 처리대기목록에 올린다.
        # 과거월은 부과대수 이력/상세로만 보관한다.
        target_pending_ym = None
        if details:
            target_pending_ym = max((int(d.get("year", 0) or 0), int(d.get("month", 0) or 0)) for d in details)


        # 월별 BillingReport 저장
        for (yy, mm), counts in summary_by_month.items():
            cnt_base = db.query(Member).filter(_clean_filter()).count()
            cnt_delivery_db = db.query(Member).filter(_clean_filter(), Member.account.in_(["관", "택배"])).count()

            r = (
                db.query(BillingReport)
                .filter(BillingReport.year == yy, BillingReport.month == mm)
                .first()
            )

            if not r:
                r = BillingReport(year=yy, month=mm, created_by=user.id)
                db.add(r)
                db.flush()

            r.cnt_join = counts.get("협회가입", 0)
            r.cnt_transfer = counts.get("양도", 0)
            r.cnt_cross = counts.get("타도", 0)
            r.cnt_close = counts.get("폐지", 0)
            r.cnt_quit = counts.get("탈퇴", 0)
            r.cnt_delivery_new = counts.get("택배신규", 0)
            r.cnt_mgmt_close = counts.get("관리비폐지", 0)
            r.cnt_age70 = counts.get("70세", 0)
            r.cnt_base = _billing_count_int(counts.get("협회기본대수"))
            r.cnt_total = _billing_count_int(counts.get("총부과대수"))
            r.cnt_delivery = _billing_count_int(counts.get("택배관리"))
            r.source_file = file.filename
            r.upload_type = "billing_counts"
            r.raw_data = json.dumps(counts, ensure_ascii=False, default=str)[:65000]
            db.add(r)
            saved_reports += 1

            # 해당 월 상세 기존 것 정리 후 재생성
            db.query(BillingPerson).filter(
                BillingPerson.billing_report_id == r.id,
                BillingPerson.reflect_status.in_(["부과대수상세", "처리대기"]),
            ).delete(synchronize_session=False)

            month_details = [d for d in details if d["year"] == yy and d["month"] == mm]

            existing_keys = _existing_pending_keys_for_billing(db)

            for d in month_details:
                pending_key = _norm_pending_key(
                    d.get("item", ""),
                    d.get("vehicle_no", ""),
                    d.get("name", "")
                )
                if pending_key in existing_keys:
                    continue
                existing_keys.add(pending_key)

                db.add(BillingPerson(
                    billing_report_id=r.id,
                    source_file=file.filename,
                    source_sheet=d.get("source_sheet", ""),
                    source_row=d.get("source_row", 0),
                    raw_data=json.dumps(d, ensure_ascii=False, default=str),
                    year=yy,
                    month=mm,
                    process_type=d.get("item", ""),
                    account="",
                    name=d.get("name", ""),
                    vehicle_no=d.get("vehicle_no", ""),
                    region=d.get("region", ""),
                    from_status="",
                    to_status="",
                    reflect_status=("처리대기" if target_pending_ym == (yy, mm) else "부과대수상세"),
                ))
                saved_details += 1

        db.commit()

        try:
            add_log(db, user.id, "부과대수업로드", f"{file.filename}: 월 {saved_reports}개, 상세 {saved_details}건")
        except Exception:
            pass

        msg = quote(f"부과대수 업로드 완료: 월 {saved_reports}개, 상세 {saved_details}건")
        return RedirectResponse(f"/billing-counts?msg={msg}", status_code=302)

    except Exception as e:
        db.rollback()
        msg = quote("부과대수 업로드 오류: " + str(e)[:180])
        return RedirectResponse(f"/billing-counts?msg={msg}", status_code=302)

    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


@app.get("/billing-counts/export")
def billing_counts_export(
    year: int = None,
    month: int = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from io import BytesIO
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    now = datetime.now()
    y = int(year or now.year)
    m = int(month or now.month)

    r = (
        db.query(BillingReport)
        .filter(BillingReport.year == y, BillingReport.month == m)
        .first()
    )

    if not r:
        return RedirectResponse(f"/billing-counts?year={y}&month={m}&msg=해당 월 부과대수 자료가 없습니다", status_code=302)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}-{m:02d} 부과대수"

    headers = [
        "기준년월", "협회가입", "양도", "타도", "폐지", "탈퇴",
        "택배신규", "관리비폐지", "70세",
        "협회기본대수", "총부과대수", f"{m}월 택배관리", "원본파일"
    ]
    ws.append(headers)

    ws.append([
        f"{y}-{m:02d}",
        r.cnt_join or 0,
        r.cnt_transfer or 0,
        r.cnt_cross or 0,
        r.cnt_close or 0,
        r.cnt_quit or 0,
        r.cnt_delivery_new or 0,
        r.cnt_mgmt_close or 0,
        r.cnt_age70 or 0,
        r.cnt_base or 0,
        r.cnt_total or 0,
        r.cnt_delivery or 0,
        r.source_file or "",
    ])

    ws2 = wb.create_sheet("상세내역")
    ws2.append(["항목", "지역", "차량번호", "성명", "원본시트", "원본행"])
    persons = (
        db.query(BillingPerson)
        .filter(BillingPerson.billing_report_id == r.id)
        .order_by(BillingPerson.process_type, BillingPerson.region, BillingPerson.name)
        .all()
    )
    for p in persons:
        ws2.append([
            p.process_type or "",
            p.region or "",
            p.vehicle_no or "",
            p.name or "",
            p.source_sheet or "",
            p.source_row or "",
        ])

    for sh in [ws, ws2]:
        header_fill = PatternFill("solid", fgColor="FCE7F3")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in sh.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row == 1:
                    cell.font = Font(bold=True, color="9D174D")
                    cell.fill = header_fill
        for col in range(1, sh.max_column + 1):
            sh.column_dimensions[get_column_letter(col)].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = quote(f"{y}년_{m:02d}월_부과대수.xlsx")
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@app.get("/billing-report", response_class=HTMLResponse)
def billing_report_page(request: Request, year: Optional[int] = None, month: Optional[int] = None,
                         db: Session = Depends(get_db), user: User = Depends(require_user)):
    now = datetime.now()
    sel_y = year or now.year; sel_m = month or now.month
    report = db.query(BillingReport).filter(
        BillingReport.year == sel_y, BillingReport.month == sel_m).first()
    history = (db.query(BillingReport)
               .order_by(BillingReport.year.desc(), BillingReport.month.desc())
               .limit(24).all())
    # 처리 대기 항목 (현재 선택 월 기준)
    pending_items = (db.query(BillingPerson)
                     .filter(BillingPerson.year == sel_y,
                             BillingPerson.month == sel_m,
                             BillingPerson.reflect_status == "처리대기")
                     .order_by(BillingPerson.process_type, BillingPerson.id)
                     .all())
    all_pending_cnt = db.query(BillingPerson).filter(
        BillingPerson.reflect_status == "처리대기").count()
    years = list(range(now.year - 2, now.year + 2))
    return templates.TemplateResponse(request, "billing_report.html", {
        "request": request, "user": user, "report": report,
        "sel_y": sel_y, "sel_m": sel_m, "history": history,
        "pending_items": pending_items, "all_pending_cnt": all_pending_cnt,
        "years": years, "months": list(range(1, 13)),
        "msg": request.query_params.get("msg", ""),
        "fmt_amt": fmt_amt,
    })


@app.post("/billing-report")
def billing_report_save(
    request: Request,
    year: int = Form(...), month: int = Form(...),
    cnt_join: int = Form(0), cnt_transfer: int = Form(0),
    cnt_cross: int = Form(0), cnt_close: int = Form(0),
    cnt_quit: int = Form(0), cnt_delivery_new: int = Form(0),
    cnt_mgmt_close: int = Form(0), cnt_age70: int = Form(0),
    memo: str = Form(""),
    db: Session = Depends(get_db), user: User = Depends(require_user)):
    cnt_base = db.query(Member).filter(_clean_filter()).count()
    cnt_delivery = db.query(Member).filter(_clean_filter(), Member.account.in_(["관", "택배"])).count()
    cnt_total = cnt_base

    r = db.query(BillingReport).filter(
        BillingReport.year == year, BillingReport.month == month).first()
    if r:
        r.cnt_join=cnt_join; r.cnt_transfer=cnt_transfer
        r.cnt_cross=cnt_cross; r.cnt_close=cnt_close
        r.cnt_quit=cnt_quit; r.cnt_delivery_new=cnt_delivery_new
        r.cnt_mgmt_close=cnt_mgmt_close; r.cnt_age70=cnt_age70
        r.cnt_base=cnt_base; r.cnt_total=cnt_total
        r.cnt_delivery=cnt_delivery; r.memo=memo
        r.upload_type = r.upload_type or "manual"
    else:
        db.add(BillingReport(
            year=year, month=month,
            cnt_join=cnt_join, cnt_transfer=cnt_transfer,
            cnt_cross=cnt_cross, cnt_close=cnt_close,
            cnt_quit=cnt_quit, cnt_delivery_new=cnt_delivery_new,
            cnt_mgmt_close=cnt_mgmt_close, cnt_age70=cnt_age70,
            cnt_base=cnt_base, cnt_total=cnt_total,
            cnt_delivery=cnt_delivery, memo=memo,
            upload_type="manual", created_by=user.id,
        ))
    db.commit()
    _invalidate_snap(db, "dashboard")
    add_log(db, user.id, "부과대수저장", f"{year}년{month}월")
    return RedirectResponse(f"/billing-report?year={year}&month={month}&msg=저장완료", status_code=302)


@app.post("/billing-report/upload")
async def billing_report_upload(
    request: Request,
    file: UploadFile = File(...),
    year: int = Form(...),
    month: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    """부과대수 엑셀 파일 업로드 — 자동 파싱 후 BillingReport + BillingPerson 저장"""
    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        return RedirectResponse(
            f"/billing-report?year={year}&month={month}&msg=엑셀파일만 업로드 가능합니다",
            status_code=302)

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = parse_billing_file(tmp_path, year, month)
    except Exception as e:
        import os as _os; _os.unlink(tmp_path)
        return RedirectResponse(
            f"/billing-report?year={year}&month={month}&msg=파일 파싱 실패: {str(e)[:80]}",
            status_code=302)
    finally:
        try:
            import os as _os; _os.unlink(tmp_path)
        except: pass

    counts = result["counts"]
    persons = result["persons"]

    # 자동계산
    cnt_base = db.query(Member).filter(_clean_filter()).count()
    cnt_delivery_db = db.query(Member).filter(_clean_filter(), Member.account.in_(["관","택배"])).count()

    r = db.query(BillingReport).filter(
        BillingReport.year == year, BillingReport.month == month).first()
    if r:
        r.cnt_join         = counts.get("협회가입", r.cnt_join or 0)
        r.cnt_transfer     = counts.get("양도", r.cnt_transfer or 0)
        r.cnt_cross        = counts.get("타도", r.cnt_cross or 0)
        r.cnt_close        = counts.get("폐지", r.cnt_close or 0)
        r.cnt_quit         = counts.get("탈퇴", r.cnt_quit or 0)
        r.cnt_delivery_new = counts.get("택배신규", r.cnt_delivery_new or 0)
        r.cnt_mgmt_close   = counts.get("관리비폐지", r.cnt_mgmt_close or 0)
        r.cnt_age70        = counts.get("70세", r.cnt_age70 or 0)
        r.cnt_base   = _billing_count_int(counts.get("협회기본대수"))
        r.cnt_total  = _billing_count_int(counts.get("총부과대수"))
        r.cnt_delivery = _billing_count_int(counts.get("택배관리"))
        r.source_file = file.filename
        r.raw_data = json.dumps(result, ensure_ascii=False, default=str)[:65000]
        r.upload_type = "file"
    else:
        r = BillingReport(
            year=year, month=month,
            cnt_join         = counts.get("협회가입", 0),
            cnt_transfer     = counts.get("양도", 0),
            cnt_cross        = counts.get("타도", 0),
            cnt_close        = counts.get("폐지", 0),
            cnt_quit         = counts.get("탈퇴", 0),
            cnt_delivery_new = counts.get("택배신규", 0),
            cnt_mgmt_close   = counts.get("관리비폐지", 0),
            cnt_age70        = counts.get("70세", 0),
            cnt_base   = _billing_count_int(counts.get("협회기본대수")),
            cnt_total  = _billing_count_int(counts.get("총부과대수")),
            cnt_delivery = _billing_count_int(counts.get("택배관리")),
            source_file = file.filename,
            raw_data = json.dumps(result, ensure_ascii=False, default=str)[:65000],
            upload_type = "file",
            created_by = user.id,
        )
        db.add(r)
    db.flush()  # r.id 확보

    # 기존 처리대기 항목 제거 후 재생성 (재업로드 시)
    db.query(BillingPerson).filter(
        BillingPerson.billing_report_id == r.id,
        BillingPerson.reflect_status == "처리대기",
    ).delete()

    created = 0
    existing_keys = _existing_pending_keys_for_billing(db)

    for p in persons:
        if p["process_type"] not in BILLING_WORK_TYPES:
            continue

        pending_key = _norm_pending_key(
            p.get("process_type", ""),
            p.get("vehicle_no", ""),
            p.get("name", "")
        )
        if pending_key in existing_keys:
            continue
        existing_keys.add(pending_key)

        db.add(BillingPerson(
            billing_report_id = r.id,
            source_file       = file.filename,
            source_sheet      = p.get("source_sheet", ""),
            source_row        = p.get("source_row", 0),
            raw_data          = json.dumps(p.get("raw_data", {}), ensure_ascii=False),
            year              = year,
            month             = month,
            process_type      = p["process_type"],
            account           = p.get("account", ""),
            name              = p.get("name", ""),
            vehicle_no        = p.get("vehicle_no", ""),
            region            = p.get("region", ""),
            from_status       = p.get("from_status", "정상"),
            to_status         = p.get("to_status", ""),
            reflect_status    = "처리대기",
        ))
        created += 1

    db.commit()
    _invalidate_snap(db, "dashboard")
    add_log(db, user.id, "부과대수업로드", f"{year}년{month}월 {file.filename} 개인항목{created}건")
    return RedirectResponse(
        f"/billing-report?year={year}&month={month}&msg=업로드완료 (항목{created}건 처리대기 생성)",
        status_code=302)


@app.post("/billing-person/{pid}/reflect")
def billing_person_reflect(pid: int, db: Session = Depends(get_db),
                            user: User = Depends(require_user)):
    bp = db.query(BillingPerson).filter(BillingPerson.id == pid).first()
    if not bp: raise HTTPException(404)
    bp.reflect_status = "반영완료"
    bp.reflected_by = user.id
    bp.reflected_at = datetime.now()
    db.commit()
    _invalidate_snap(db, "dashboard")
    return RedirectResponse(
        f"/billing-report?year={bp.year}&month={bp.month}&msg=반영완료",
        status_code=302)


@app.post("/billing-person/{pid}/hold")
def billing_person_hold(pid: int, db: Session = Depends(get_db),
                        user: User = Depends(require_user)):
    bp = db.query(BillingPerson).filter(BillingPerson.id == pid).first()
    if not bp: raise HTTPException(404)
    bp.reflect_status = "보류"
    db.commit()
    return RedirectResponse(
        f"/billing-report?year={bp.year}&month={bp.month}&msg=보류처리",
        status_code=302)


@app.post("/billing-person/{pid}/exclude")
def billing_person_exclude(pid: int, db: Session = Depends(get_db),
                            user: User = Depends(require_user)):
    bp = db.query(BillingPerson).filter(BillingPerson.id == pid).first()
    if not bp: raise HTTPException(404)
    bp.reflect_status = "제외"
    db.commit()
    return RedirectResponse(
        f"/billing-report?year={bp.year}&month={bp.month}&msg=제외처리",
        status_code=302)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT","8080")))


# =========================================================
# 보고/집계 엑셀 추출 라우트
# - N월 부과대수 엑셀 추출
# - N월 입금추출 엑셀 추출
# =========================================================
try:
    from report_export_routes import router as report_export_router
    app.include_router(report_export_router)
    print("INFO: report export routes loaded")
except Exception as e:
    print("WARN: report export routes not loaded:", e)


# =========================================================
# ??? / ??? ??
# - BankTransaction.match_status = ??? / ??? ?? ??
# - ?? ?? ??
# - ?? ????
# - ???? ???? ?? ??
# =========================================================

def _tx_amount_for_income(tx):
    import re as _re
    for attr in ["amount", "deposit_amount", "in_amount", "paid_amount", "txn_amount", "money"]:
        if hasattr(tx, attr):
            v = getattr(tx, attr)
            try:
                if v is None:
                    continue
                if isinstance(v, (int, float)):
                    return int(v)
                s = str(v).replace(",", "").replace("₩", "").replace("�", "").strip()
                m = _re.search(r"-?\d+", s)
                if m:
                    return int(m.group(0))
            except Exception:
                pass
    return 0


def _tx_date_for_income(tx):
    for attr in ["txn_date", "date", "paid_date", "deposit_date", "transaction_date"]:
        if hasattr(tx, attr):
            v = getattr(tx, attr)
            if v:
                return str(v)[:10]
    return ""


def _tx_memo_for_income(tx):
    for attr in ["memo", "description", "raw_data"]:
        if hasattr(tx, attr):
            v = getattr(tx, attr)
            if v:
                return str(v)
    return ""



def _ensure_income_ledger_details(db):
    """
    income_ledger_details 테이블 보장.
    SQLite/PostgreSQL 모두 안전하게 처리.
    실패 시 rollback 해서 이후 쿼리까지 죽지 않게 함.
    """
    try:
        from sqlalchemy import text as _t2

        dialect = getattr(getattr(db, "bind", None), "dialect", None)
        name = getattr(dialect, "name", "")

        if name == "postgresql":
            ddl = """
            CREATE TABLE IF NOT EXISTS income_ledger_details (
                id SERIAL PRIMARY KEY,
                bank_transaction_id INTEGER,
                income_type VARCHAR(20),
                work_type VARCHAR(50),
                pending_target VARCHAR(20) DEFAULT '없음',
                related_vehicle_no VARCHAR(100),
                related_name VARCHAR(100),
                note TEXT,
                next_billing_date VARCHAR(20),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP
            )
            """
        else:
            ddl = """
            CREATE TABLE IF NOT EXISTS income_ledger_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_transaction_id INTEGER,
                income_type VARCHAR(20),
                work_type VARCHAR(50),
                pending_target VARCHAR(20) DEFAULT '없음',
                related_vehicle_no VARCHAR(100),
                related_name VARCHAR(100),
                note TEXT,
                next_billing_date VARCHAR(20),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP
            )
            """

        db.execute(_t2(ddl))
        db.commit()

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("income_ledger_details ensure error:", repr(e))


def _calc_pending_target(income_type, work_type):
    ASSOC_KW = ["가입비", "특별회비", "협회가입", "협회비", "입회"]
    MGMT_KW  = ["자격증명발급", "자격증명", "택배신규", "신규관리", "관리비"]
    wt = str(work_type or "")
    it = str(income_type or "")
    if it == "가수금" and any(k in wt for k in ASSOC_KW):
        return "협회비"
    if it == "잡수입" and any(k in wt for k in MGMT_KW):
        return "관리비"
    return "없음"


def _upsert_income_ledger_detail_from_bank(
    db,
    tx,
    income_type="",
    work_type="",
    related_vehicle_no="",
    related_name="",
    note="",
    billing_date=""
):
    """
    통장매칭에서 잡수입/가수금으로 분류하는 순간
    income_ledger_details에 구조화 상세를 자동 저장한다.
    실패해도 화면 전체가 죽지 않게 False만 반환.
    """
    try:
        from sqlalchemy import text as _t

        _ensure_income_ledger_details(db)

        income_type = str(income_type or "").strip()
        work_type = str(work_type or "").strip()
        related_vehicle_no = str(related_vehicle_no or "").strip()
        related_name = str(related_name or "").strip()
        note = str(note or "").strip()
        billing_date = str(billing_date or "").strip()

        # 사유가 비어 있으면 match_reason/memo에서 키워드 추정
        if not work_type:
            raw = (str(getattr(tx, "match_reason", "") or "") + " " +
                   str(getattr(tx, "memo", "") or ""))
            for kw in [
                "자격증명발급", "자격증명", "택배신규", "신규관리", "관리비",
                "가입비", "특별회비", "협회가입", "협회비", "입회",
                "대폐차", "예금이자", "상가임대료", "기타"
            ]:
                if kw in raw:
                    work_type = kw
                    break

        pending_target = _calc_pending_target(income_type, work_type)

        txn_date = str(getattr(tx, "txn_date", "") or "")[:10]
        next_billing_date = ""
        if pending_target in ["협회비", "관리비"]:
            next_billing_date = billing_date or _calc_next_billing_date(txn_date)

        row = db.execute(_t("""
            SELECT id
            FROM income_ledger_details
            WHERE bank_transaction_id = :tid
            LIMIT 1
        """), {"tid": getattr(tx, "id", None)}).mappings().first()

        params = {
            "tid": getattr(tx, "id", None),
            "income_type": income_type,
            "work_type": work_type,
            "pending_target": pending_target,
            "related_vehicle_no": related_vehicle_no,
            "related_name": related_name,
            "note": note,
            "next_billing_date": next_billing_date,
        }

        if row:
            params["id"] = row["id"]
            db.execute(_t("""
                UPDATE income_ledger_details
                SET income_type = :income_type,
                    work_type = :work_type,
                    pending_target = :pending_target,
                    related_vehicle_no = :related_vehicle_no,
                    related_name = :related_name,
                    note = :note,
                    next_billing_date = :next_billing_date,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
            """), params)
        else:
            db.execute(_t("""
                INSERT INTO income_ledger_details
                (bank_transaction_id, income_type, work_type, pending_target,
                 related_vehicle_no, related_name, note, next_billing_date)
                VALUES
                (:tid, :income_type, :work_type, :pending_target,
                 :related_vehicle_no, :related_name, :note, :next_billing_date)
            """), params)

        return True

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("income detail upsert from bank error:", repr(e))
        return False



@app.get("/income-ledger", response_class=HTMLResponse)
def income_ledger_page(
    request: Request,
    kind: str = "",
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text as _t3
    MISC = "잡수입"
    SUSP = "가수금"

    if kind not in [MISC, SUSP]:
        kind = MISC

    _ensure_income_ledger_details(db)

    bq = db.query(BankTransaction).filter(BankTransaction.match_status == kind)
    if q:
        like = f"%{q}%"
        bq = bq.filter(BankTransaction.memo.ilike(like))
    txs = bq.order_by(BankTransaction.id.desc()).all()
    total_amount = sum(_tx_amount_for_income(tx) for tx in txs)

    # IncomeLedgerDetail 맵 (bank_transaction_id -> detail)
    detail_map = {}
    try:
        all_details = db.query(IncomeLedgerDetail).all()
        for d in all_details:
            if d.bank_transaction_id:
                detail_map[d.bank_transaction_id] = d
    except Exception:
        pass

    # 상세 없는 tx는 match_reason에서 자동 생성
    rows = []
    for tx in txs:
        d = detail_map.get(tx.id)
        if d:
            rows.append({
                "tx": tx,
                "id": tx.id,
                "txn_date": str(getattr(tx, "txn_date", "") or "")[:10],
                "amount": int(getattr(tx, "deposit_amount", 0) or 0),
                "memo": str(getattr(tx, "memo", "") or ""),
                "income_type": d.income_type or kind,
                "work_type": d.work_type or "",
                "pending_target": d.pending_target or "없음",
                "related_vehicle_no": d.related_vehicle_no or "",
                "related_name": d.related_name or "",
                "next_billing_date": d.next_billing_date or "",
                "note": d.note or "",
                "has_detail": True,
            })
        else:
            parsed = _parse_income_reason(str(getattr(tx, "match_reason", "") or ""))
            memo = str(getattr(tx, "memo", "") or "")
            combined = str(getattr(tx, "match_reason", "") or "") + " " + memo
            wt = parsed.get("work_reason") or ""
            if not wt:
                for kw in ["가입비","특별회비","협회가입","협회비","자격증명발급","자격증명","택배신규","신규관리","예금이자","상가임대료","대폐차"]:
                    if kw in combined:
                        wt = kw; break
            pt = _calc_pending_target(kind, wt)
            txn_date = str(getattr(tx, "txn_date", "") or "")[:10]
            rows.append({
                "tx": tx,
                "id": tx.id,
                "txn_date": txn_date,
                "amount": int(getattr(tx, "deposit_amount", 0) or 0),
                "memo": str(getattr(tx, "memo", "") or ""),
                "income_type": kind,
                "work_type": wt,
                "pending_target": pt,
                "related_vehicle_no": parsed.get("vehicle_no") or "",
                "related_name": parsed.get("name") or "",
                "next_billing_date": _calc_next_billing_date(txn_date) if pt != "없음" else "",
                "note": parsed.get("note") or "",
                "has_detail": False,
            })

    return templates.TemplateResponse(request, "income_ledger.html", {
        "request": request,
        "user": user,
        "kind": kind,
        "q": q,
        "rows": rows,
        "total_count": len(rows),
        "total_amount": total_amount,
        "fmt_amt": fmt_amt,
    })


@app.get("/income-ledger/export")
def income_ledger_export(
    kind: str = "",
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    """
    잡수입·가수금 관리 엑셀 다운로드.
    업무용으로 필요한 컬럼만 출력:
    성명 / 차량번호 / 업무사유 / 입금액
    """
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy import text as _t

    MISC = "잡수입"
    SUSP = "가수금"

    if kind not in [MISC, SUSP]:
        kind = MISC

    try:
        _ensure_income_ledger_details(db)
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("income export ensure detail error:", repr(e))

    bq = db.query(BankTransaction).filter(BankTransaction.match_status == kind)

    if q:
        like = f"%{q}%"
        try:
            bq = bq.filter(BankTransaction.memo.ilike(like))
        except Exception:
            pass

    txs = bq.order_by(BankTransaction.id.desc()).all()

    # income_ledger_details 매핑
    detail_map = {}
    try:
        rows = db.execute(_t("""
            SELECT bank_transaction_id, work_type, pending_target,
                   related_vehicle_no, related_name, note, next_billing_date
            FROM income_ledger_details
        """)).mappings().all()
        for r in rows:
            detail_map[r["bank_transaction_id"]] = r
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("income export detail load error:", repr(e))
        detail_map = {}

    wb = Workbook()
    ws = wb.active
    ws.title = kind

    headers = ["성명", "차량번호", "업무사유", "입금액"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="FFEAF3")
    header_font = Font(bold=True, color="C2185B")
    thin = Side(style="thin", color="E8C7D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for tx in txs:
        d = detail_map.get(getattr(tx, "id", None))

        if d:
            name = d.get("related_name") or ""
            vehicle_no = d.get("related_vehicle_no") or ""
            work_type = d.get("work_type") or ""
        else:
            parsed = {}
            try:
                parsed = _parse_income_reason(str(getattr(tx, "match_reason", "") or ""))
            except Exception:
                parsed = {}

            raw = (str(getattr(tx, "match_reason", "") or "") + " " +
                   str(getattr(tx, "memo", "") or ""))

            name = parsed.get("name") or ""
            vehicle_no = parsed.get("vehicle_no") or ""
            work_type = parsed.get("work_reason") or ""

            if not work_type:
                for kw in [
                    "자격증명발급", "자격증명", "택배신규", "신규관리", "관리비",
                    "가입비", "특별회비", "협회가입", "협회비", "입회",
                    "대폐차", "예금이자", "상가임대료", "기타"
                ]:
                    if kw in raw:
                        work_type = kw
                        break

        amount = 0
        try:
            amount = int(getattr(tx, "deposit_amount", 0) or getattr(tx, "amount", 0) or 0)
        except Exception:
            amount = 0

        ws.append([name, vehicle_no, work_type, amount])

    # 보기 좋게 서식
    widths = {
        "A": 16,   # 성명
        "B": 18,   # 차량번호
        "C": 22,   # 업무사유
        "D": 14,   # 입금액
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        row[3].number_format = '#,##0'

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "income_ledger_clean.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@app.post("/billing-counts/generate-next-arrears")
def generate_next_month_arrears_from_billing_counts(
    year: int = Form(...),
    month: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    from sqlalchemy import text
    import datetime as _dt
    import re as _re

    FEE_ASSOC = 10000
    FEE_MGMT = 5000

    ACC_ASSOC = "\ud611\ud68c\ube44"   # ???
    ACC_MGMT = "\uad00\ub9ac\ube44"    # ???

    # ??? ??
    if month == 12:
        bill_year = year + 1
        bill_month = 1
    else:
        bill_year = year
        bill_month = month + 1

    # ?? ?? ??? ???
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS billing_generated_arrears (
            member_id INTEGER NOT NULL,
            bill_year INTEGER NOT NULL,
            bill_month INTEGER NOT NULL,
            account VARCHAR(50) NOT NULL,
            amount INTEGER NOT NULL,
            source_year INTEGER NOT NULL,
            source_month INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (member_id, bill_year, bill_month, account)
        )
    """))

    def _s(v):
        return str(v or "").strip()

    def _compact(v):
        return _re.sub(r"\s+", "", _s(v))

    def _get(m, names):
        for n in names:
            if hasattr(m, n):
                v = getattr(m, n)
                if v not in [None, ""]:
                    return v
        return None

    def _parse_date(v):
        if not v:
            return None
        if isinstance(v, _dt.datetime):
            return v.date()
        if isinstance(v, _dt.date):
            return v
        s = str(v).strip()
        m = _re.search(r"(\d{4})[-./?\s]*(\d{1,2})[-./?\s]*(\d{1,2})?", s)
        if not m:
            return None
        y = int(m.group(1))
        mo = int(m.group(2))
        d = int(m.group(3) or 1)
        try:
            return _dt.date(y, mo, d)
        except Exception:
            return None

    def _next_month_ym(d):
        if not d:
            return None
        if d.month == 12:
            return (d.year + 1, 1)
        return (d.year, d.month + 1)

    def _ym_le(a, b):
        return a[0] < b[0] or (a[0] == b[0] and a[1] <= b[1])

    def _vehicle(m):
        return _s(_get(m, [
            "vehicle_no", "vehicle_number", "car_no", "plate_number",
            "car_number", "truck_no"
        ]))

    def _account(m):
        return _s(_get(m, ["account", "account_type", "fee_type", "acct"]))

    def _status_text(m):
        parts = []
        for n in [
            "status", "member_status", "process_type", "process_status",
            "note", "memo", "remark", "remarks", "bigo", "description"
        ]:
            if hasattr(m, n):
                v = getattr(m, n)
                if v:
                    parts.append(str(v))
        return _compact(" ".join(parts))

    def _is_excluded(m):
        txt = _status_text(m)
        bad_words = [
            "\ud3d0\uc5c5",  # ??
            "\ud3d0\uc9c0",  # ??
            "\uc591\ub3c4",  # ??
            "\uc774\uad00",  # ??
            "\ud0c8\ud1f4",  # ??
            "\uc0ac\ub9dd",  # ??
            "\uc911\uc9c0",  # ??
        ]
        return any(w in txt for w in bad_words)

    def _is_delivery_vehicle(m):
        veh = _vehicle(m)
        return "\ubc30" in veh  # ?

    def _delivery_charge_allowed(m):
        # ??? ???? + ???????? ? ? ??? ????? ??
        approval = _parse_date(_get(m, [
            "approval_date", "authorized_date", "permit_date",
            "approval_dt", "inga_date"
        ]))
        cert = _parse_date(_get(m, [
            "certificate_issue_date", "cert_issue_date", "license_issue_date",
            "qualification_issue_date", "jagyeok_date"
        ]))

        # ?? ??? ?? ? ?? ?? DB? ??? ?? ??
        has_approval_col = any(hasattr(m, n) for n in [
            "approval_date", "authorized_date", "permit_date", "approval_dt", "inga_date"
        ])
        has_cert_col = any(hasattr(m, n) for n in [
            "certificate_issue_date", "cert_issue_date", "license_issue_date",
            "qualification_issue_date", "jagyeok_date"
        ])

        if not has_approval_col and not has_cert_col:
            return True

        if not approval or not cert:
            return False

        start_base = approval if approval >= cert else cert
        start_ym = _next_month_ym(start_base)
        target_ym = (bill_year, bill_month)
        return _ym_le(start_ym, target_ym)

    def _target_account_and_amount(m):
        acc = _account(m)
        acc_c = _compact(acc)

        if "\ud611" in acc_c:  # ?
            return ACC_ASSOC, FEE_ASSOC

        if "\uad00" in acc_c:  # ?
            return ACC_MGMT, FEE_MGMT

        return None, 0

    members = db.query(Member).all()

    created_assoc = 0
    created_mgmt = 0
    skipped_duplicate = 0
    skipped_excluded = 0
    skipped_no_account = 0
    skipped_delivery_wait = 0
    total_amount = 0

    for m in members:
        if _is_excluded(m):
            skipped_excluded += 1
            continue

        if _is_delivery_vehicle(m) and not _delivery_charge_allowed(m):
            skipped_delivery_wait += 1
            continue

        account, amount = _target_account_and_amount(m)
        if not account or amount <= 0:
            skipped_no_account += 1
            continue

        exists = db.execute(text("""
            SELECT 1
            FROM billing_generated_arrears
            WHERE member_id = :member_id
              AND bill_year = :bill_year
              AND bill_month = :bill_month
              AND account = :account
        """), {
            "member_id": m.id,
            "bill_year": bill_year,
            "bill_month": bill_month,
            "account": account,
        }).first()

        if exists:
            skipped_duplicate += 1
            continue

        db.execute(text("""
            INSERT INTO billing_generated_arrears
            (member_id, bill_year, bill_month, account, amount, source_year, source_month)
            VALUES
            (:member_id, :bill_year, :bill_month, :account, :amount, :source_year, :source_month)
        """), {
            "member_id": m.id,
            "bill_year": bill_year,
            "bill_month": bill_month,
            "account": account,
            "amount": amount,
            "source_year": year,
            "source_month": month,
        })

        # /arrears ?????? ?? ????? Member.excel_arrears ??
        old_arr = getattr(m, "excel_arrears", 0) or 0
        try:
            old_arr = int(old_arr)
        except Exception:
            old_arr = 0

        m.excel_arrears = old_arr + amount

        if account == ACC_ASSOC:
            created_assoc += 1
        else:
            created_mgmt += 1

        total_amount += amount
        db.add(m)

    db.commit()

    msg = (
        f"{bill_year}\ub144 {bill_month}\uc6d4 \ubbf8\uc218\uae08 \uc0dd\uc131 \uc644\ub8cc / "
        f"\ud611\ud68c\ube44 {created_assoc}\uac74 / "
        f"\uad00\ub9ac\ube44 {created_mgmt}\uac74 / "
        f"\uc911\ubcf5\uc81c\uc678 {skipped_duplicate}\uac74 / "
        f"\uc0dd\uc131\uae08\uc561 {total_amount:,}\uc6d0"
    )

    return RedirectResponse(
        f"/billing-counts?year={year}&month={month}&msg=" + quote(msg),
        status_code=302
    )


@app.get("/bank/{tid}/mark-income", response_class=HTMLResponse)
def bank_mark_income_form(
    tid: int,
    request: Request,
    kind: str = "misc",
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        raise HTTPException(404)

    kind_code = str(kind or "misc").strip().lower()

    return templates.TemplateResponse(request, "income_mark_form.html", {
        "request": request,
        "user": user,
        "tx": tx,
        "kind_code": kind_code,
        "fmt_amt": fmt_amt,
    })


@app.post("/bank/{tid}/mark-income")
def bank_mark_income_save(
    tid: int,
    kind: str = Form(...),
    reason: str = Form(""),
    billing_date: str = Form(""),
    related_vehicle_no: str = Form(""),
    related_name: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    from sqlalchemy import text
    import datetime as _dt

    MISC = "\uc7a1\uc218\uc785"          # ???
    SUSP = "\uac00\uc218\uae08"          # ???
    DONE = "\ubc18\uc601\uc644\ub8cc"    # ????
    PENDING = "\ubc18\uc601\ub300\uae30" # ????

    raw_kind = str(kind or "").strip().lower()

    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        raise HTTPException(404)

    if str(getattr(tx, "match_status", "") or "") == DONE:
        return RedirectResponse(
            "/bank?msg=" + quote("?? ????? ??? ??? ? ???? ???."),
            status_code=302
        )

    parts = []
    if reason:
        parts.append("\ube44\uace0: " + reason.strip())
    if related_vehicle_no:
        parts.append("\uc218\ub3d9\ubd84\ub958: " + related_vehicle_no.strip())
    if related_name:
        parts.append("\uc218\ub3d9\ubd84\ub958: " + related_name.strip())
    if note:
        parts.append("\ube44\uace0: " + note.strip())

    memo = str(getattr(tx, "memo", "") or "")
    amount = getattr(tx, "deposit_amount", None) or getattr(tx, "amount", None) or 0
    txn_date = str(getattr(tx, "txn_date", "") or "")[:10]

    # 1) ???
    if raw_kind in ["misc", "misc_income"]:
        new_status = MISC
        old_reason = getattr(tx, "match_reason", "") or ""
        add_reason = "\uc218\ub3d9\ubd84\ub958: " + new_status
        if parts:
            add_reason += " / " + " / ".join(parts)

        tx.match_status = new_status
        tx.matched_member_id = None
        tx.match_reason = (old_reason + ", " if old_reason else "") + add_reason
        _upsert_income_ledger_detail_from_bank(
            db=db,
            tx=tx,
            income_type=new_status,
            work_type=locals().get("reason", locals().get("work_type", "")),
            related_vehicle_no=locals().get("related_vehicle_no", ""),
            related_name=locals().get("related_name", ""),
            note=locals().get("note", ""),
        )

        db.add(tx)
        db.commit()

        return RedirectResponse(
            "/income-ledger?kind=" + quote(new_status),
            status_code=302
        )

    # 2) ???
    if raw_kind in ["suspense", "temporary", "deposit"]:
        new_status = SUSP
        old_reason = getattr(tx, "match_reason", "") or ""
        add_reason = "\uc218\ub3d9\ubd84\ub958: " + new_status
        if parts:
            add_reason += " / " + " / ".join(parts)

        tx.match_status = new_status
        tx.matched_member_id = None
        tx.match_reason = (old_reason + ", " if old_reason else "") + add_reason
        _upsert_income_ledger_detail_from_bank(
            db=db,
            tx=tx,
            income_type=new_status,
            work_type=locals().get("reason", locals().get("work_type", "")),
            related_vehicle_no=locals().get("related_vehicle_no", ""),
            related_name=locals().get("related_name", ""),
            note=locals().get("note", ""),
        )

        db.add(tx)
        db.commit()

        return RedirectResponse(
            "/income-ledger?kind=" + quote(new_status),
            status_code=302
        )

    # 3) ???? - ??? / ???
    if raw_kind in ["pending_assoc", "pending_mgmt"]:
        pending_type = "???" if raw_kind == "pending_assoc" else "???"

        # ?? ??? ????? ??
        old_reason = getattr(tx, "match_reason", "") or ""
        add_reason = pending_type
        if parts:
            add_reason += " / " + " / ".join(parts)

        tx.match_status = PENDING
        tx.matched_member_id = None
        tx.match_reason = (old_reason + ", " if old_reason else "") + add_reason
        _upsert_income_ledger_detail_from_bank(
            db=db,
            tx=tx,
            income_type=new_status,
            work_type=locals().get("reason", locals().get("work_type", "")),
            related_vehicle_no=locals().get("related_vehicle_no", ""),
            related_name=locals().get("related_name", ""),
            note=locals().get("note", ""),
            billing_date=locals().get("billing_date", locals().get("next_billing_date", "")),
        )

        db.add(tx)

        # ???? ?? ??? ??
        # ?? WorkQueue ??? ????? ??, ???? ????? ?? ??
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS bank_income_pending_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_transaction_id INTEGER,
                process_type VARCHAR(50),
                income_kind VARCHAR(50),
                txn_date VARCHAR(20),
                amount INTEGER,
                memo TEXT,
                related_vehicle_no VARCHAR(100),
                related_name VARCHAR(100),
                reason TEXT,
                note TEXT,
                status VARCHAR(50) DEFAULT '반영대기',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))

        db.execute(text("""
            INSERT INTO bank_income_pending_queue
            (bank_transaction_id, process_type, income_kind, txn_date, amount, memo,
             related_vehicle_no, related_name, reason, note, status)
            VALUES
            (:bank_transaction_id, :process_type, :income_kind, :txn_date, :amount, :memo,
             :related_vehicle_no, :related_name, :reason, :note, :status)
        """), {
            "bank_transaction_id": tid,
            "process_type": "????",
            "income_kind": pending_type,
            "txn_date": txn_date,
            "amount": int(amount or 0),
            "memo": memo,
            "related_vehicle_no": related_vehicle_no,
            "related_name": related_name,
            "reason": reason,
            "note": note,
            "status": "????",
        })

        db.commit()

        return RedirectResponse(
            "/work?status=" + quote("????") + "&msg=" + quote(pending_type + " ????? ?????."),
            status_code=302
        )

    # ???? ???
    tx.match_status = SUSP
    tx.matched_member_id = None
    tx.match_reason = "\uc218\ub3d9\ubd84\ub958: " + SUSP
    db.add(tx)
    db.commit()

    return RedirectResponse(
        "/income-ledger?kind=" + quote(SUSP),
        status_code=302
    )


# =========================================================
# ???/??? ??: ???? ??? ??? ????? ???
# =========================================================
@app.post("/income-ledger/{tid}/delete")
def income_ledger_delete(
    tid: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote

    UNMATCHED = "\ubbf8\ub9e4\uce6d"  # ???

    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        raise HTTPException(404)

    tx.match_status = UNMATCHED
    tx.matched_member_id = None
    tx.match_reason = "???/??? ?? ?? - ????? ??"

    db.add(tx)
    db.commit()

    return RedirectResponse(
        "/bank?status=" + quote(UNMATCHED) + "&msg=" + quote("???/??? ??? ?????? ???????."),
        status_code=302
    )


# =========================================================
# ???? ??: ???? ? ? ????? ??
# =========================================================
@app.post("/bank/{tid}/delete")
def bank_transaction_delete(
    tid: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote

    DONE = "\ubc18\uc601\uc644\ub8cc"  # ????

    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        raise HTTPException(404)

    if getattr(tx, "applied", False) or str(getattr(tx, "match_status", "") or "") == DONE:
        return RedirectResponse(
            "/bank?msg=" + quote("????? ????? ??? ? ????."),
            status_code=302
        )

    db.delete(tx)
    db.commit()

    return RedirectResponse(
        "/bank?msg=" + quote("????? ??????."),
        status_code=302
    )


def _ensure_pending_income_table(db):
    from sqlalchemy import text
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS bank_income_pending_queue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_transaction_id INTEGER,
            process_type VARCHAR(50),
            income_kind VARCHAR(50),
            txn_date VARCHAR(20),
            amount INTEGER,
            memo TEXT,
            related_vehicle_no VARCHAR(100),
            related_name VARCHAR(100),
            reason TEXT,
            note TEXT,
            status VARCHAR(50) DEFAULT '반영대기',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """))


# =========================================================
# ?????? - ???? ?? ??
# =========================================================
@app.get("/work/pending-income", response_class=HTMLResponse)
def pending_income_page(
    request: Request,
    q: str = "",
    status: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text

    _ensure_pending_income_table(db)

    where = []
    params = {}

    if q:
        where.append("""
            (
                memo LIKE :q OR
                related_vehicle_no LIKE :q OR
                related_name LIKE :q OR
                reason LIKE :q OR
                note LIKE :q OR
                income_kind LIKE :q
            )
        """)
        params["q"] = f"%{q}%"

    if status:
        where.append("status = :status")
        params["status"] = status

    sql = """
        SELECT *
        FROM bank_income_pending_queue
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"

    rows = db.execute(text(sql), params).mappings().all()

    return templates.TemplateResponse(request, "pending_income.html", {
        "request": request,
        "user": user,
        "rows": rows,
        "q": q,
        "status": status,
        "fmt_amt": fmt_amt,
    })


# =========================================================
# ?????? - ???? ?? ?? ??
# =========================================================
@app.get("/work/pending-income/add", response_class=HTMLResponse)
def pending_income_add_form(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    return templates.TemplateResponse(request, "pending_income_add.html", {
        "request": request,
        "user": user,
    })


@app.post("/work/pending-income/add")
def pending_income_add_save(
    income_kind: str = Form(...),
    txn_date: str = Form(""),
    amount: int = Form(0),
    memo: str = Form(""),
    related_vehicle_no: str = Form(""),
    related_name: str = Form(""),
    reason: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    from sqlalchemy import text

    _ensure_pending_income_table(db)

    allowed = [
        "???",
        "???",
    ]
    if income_kind not in allowed:
        income_kind = "???"

    db.execute(text("""
        INSERT INTO bank_income_pending_queue
        (bank_transaction_id, process_type, income_kind, txn_date, amount, memo,
         related_vehicle_no, related_name, reason, note, status)
        VALUES
        (NULL, '????', :income_kind, :txn_date, :amount, :memo,
         :related_vehicle_no, :related_name, :reason, :note, '????')
    """), {
        "income_kind": income_kind,
        "txn_date": txn_date,
        "amount": int(amount or 0),
        "memo": memo,
        "related_vehicle_no": related_vehicle_no,
        "related_name": related_name,
        "reason": reason,
        "note": note,
    })

    db.commit()

    return RedirectResponse(
        "/work/pending-income?msg=" + quote("???? ??? ??????."),
        status_code=302
    )


# =========================================================
# ?????? - ???? ??
# =========================================================
@app.post("/work/pending-income/{pid}/delete")
def pending_income_delete(
    pid: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    from sqlalchemy import text

    _ensure_pending_income_table(db)

    db.execute(text("DELETE FROM bank_income_pending_queue WHERE id = :id"), {"id": pid})
    db.commit()

    return RedirectResponse(
        "/work/pending-income?msg=" + quote("???? ??? ??????."),
        status_code=302
    )


# =========================================================
# ?????? - ???? ??
# ??? ??? ????? ??.
# ?? ???/??? ??? ?? ??? ?? ???? Member ?? ?? ??? ?.
# =========================================================
@app.post("/work/pending-income/{pid}/apply")
def pending_income_apply(
    pid: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    from sqlalchemy import text

    _ensure_pending_income_table(db)

    db.execute(text("""
        UPDATE bank_income_pending_queue
        SET status = '????'
        WHERE id = :id
    """), {"id": pid})

    db.commit()

    return RedirectResponse(
        "/work/pending-income?status=????&msg=" + quote("???? ??????."),
        status_code=302
    )


@app.post("/work/{wid}/apply")
@app.post("/work-queue/{wid}/apply")
@app.post("/workqueue/{wid}/apply")
def apply_work_queue_item(
    wid: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    import datetime as _dt

    w = db.query(WorkQueue).filter(WorkQueue.id == wid).first()
    if not w:
        raise HTTPException(404)

    # ???? ??
    process_type = str(
        getattr(w, "process_type", None)
        or getattr(w, "type", None)
        or getattr(w, "work_type", None)
        or ""
    ).strip()

    # ?? ?? ??
    member_id = (
        getattr(w, "member_id", None)
        or getattr(w, "mid", None)
        or getattr(w, "target_member_id", None)
    )

    m = None
    if member_id:
        m = db.query(Member).filter(Member.id == member_id).first()

    # member_id? ??? ??/????? ????
    if not m:
        name = str(getattr(w, "name", "") or getattr(w, "member_name", "") or "").strip()
        vehicle_no = str(getattr(w, "vehicle_no", "") or getattr(w, "car_no", "") or "").strip()

        q = db.query(Member)
        if vehicle_no:
            m = q.filter(Member.vehicle_no == vehicle_no).first() if hasattr(Member, "vehicle_no") else None
        if not m and name:
            m = q.filter(Member.name == name).first() if hasattr(Member, "name") else None

    exclude_types = [
        "??",
        "??",
        "??",
        "??",
        "??",
        "??",
        "??",
    ]

    restore_types = [
        "????",
        "??",
    ]

    # 1) ??/??/??/??/??/??? ?????? ?? ???? ??
    #    ???? ?? ??
    if any(x in process_type for x in exclude_types):
        if m:
            # ?? ???? ???? ? ???? 0 ??
            if hasattr(m, "excel_arrears"):
                m.excel_arrears = 0

            # ??/?? ?? ??? ??? ?? ??
            for attr in ["status", "member_status", "billing_status", "arrears_status"]:
                if hasattr(m, attr):
                    try:
                        setattr(m, attr, process_type)
                    except Exception:
                        pass

            # ??/??? ?? ?? ??
            note_text = f"{_dt.date.today()} {process_type} ?? - ???? ??"
            for attr in ["note", "memo", "remark", "remarks", "bigo"]:
                if hasattr(m, attr):
                    old = str(getattr(m, attr) or "")
                    try:
                        setattr(m, attr, (old + " / " if old else "") + note_text)
                        break
                    except Exception:
                        pass

            db.add(m)

        # workqueue ??? ????
        if hasattr(w, "status"):
            w.status = "????"
        if hasattr(w, "applied_at"):
            w.applied_at = _dt.datetime.now()
        if hasattr(w, "result"):
            w.result = process_type + " ?? - ???? ??"

        db.add(w)
        db.commit()

        return RedirectResponse(
            "/work?msg=" + quote(process_type + " ????: ????? ??? ?? ?????? ??????."),
            status_code=302
        )

    # 2) ????? ?? ?????? ??
    if any(x in process_type for x in restore_types):
        if m:
            for attr in ["status", "member_status", "billing_status", "arrears_status"]:
                if hasattr(m, attr):
                    try:
                        setattr(m, attr, "??")
                    except Exception:
                        pass

            note_text = f"{_dt.date.today()} ???? ?? - ???? ??"
            for attr in ["note", "memo", "remark", "remarks", "bigo"]:
                if hasattr(m, attr):
                    old = str(getattr(m, attr) or "")
                    try:
                        setattr(m, attr, (old + " / " if old else "") + note_text)
                        break
                    except Exception:
                        pass

            db.add(m)

        if hasattr(w, "status"):
            w.status = "????"
        if hasattr(w, "applied_at"):
            w.applied_at = _dt.datetime.now()
        if hasattr(w, "result"):
            w.result = "???? ?? - ???? ??"

        db.add(w)
        db.commit()

        return RedirectResponse(
            "/work?msg=" + quote("???? ????: ?????? ??????."),
            status_code=302
        )

    # 3) ? ? ??? ??? ????
    if hasattr(w, "status"):
        w.status = "????"
    if hasattr(w, "applied_at"):
        w.applied_at = _dt.datetime.now()
    if hasattr(w, "result"):
        w.result = process_type + " ????"

    db.add(w)
    db.commit()

    return RedirectResponse(
        "/work?msg=" + quote(process_type + " ????"),
        status_code=302
    )


@app.post("/bank/{tid}/apply")
@app.post("/bank/{tid}/confirm")
@app.post("/bank/{tid}/apply-payment")
@app.post("/bank/{tid}/confirm-apply")
async def bank_apply_payment_safe(
    tid: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    from fastapi.responses import JSONResponse
    import datetime as _dt
    import re as _re

    DONE = "\ubc18\uc601\uc644\ub8cc"      # ????
    UNMATCHED = "\ubbf8\ub9e4\uce6d"      # ???

    def _num(v):
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).replace(",", "").replace("₩", "").replace("�", "").strip()
        m = _re.search(r"-?\d+", s)
        return int(m.group(0)) if m else 0

    def _get(obj, names):
        for n in names:
            if hasattr(obj, n):
                v = getattr(obj, n)
                if v not in [None, ""]:
                    return v
        return None

    def _wants_json():
        h = request.headers
        return (
            "application/json" in (h.get("accept") or "")
            or "application/json" in (h.get("content-type") or "")
            or (h.get("x-requested-with") or "").lower() == "xmlhttprequest"
        )

    async def _read_payload():
        data = {}
        try:
            if "application/json" in (request.headers.get("content-type") or ""):
                data = await request.json()
            else:
                form = await request.form()
                data = dict(form)
        except Exception:
            data = {}
        return data

    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        if _wants_json():
            return JSONResponse({"ok": False, "message": "????? ?? ? ????."}, status_code=404)
        raise HTTPException(404)

    if str(getattr(tx, "match_status", "") or "") == DONE or getattr(tx, "applied", False):
        msg = "?? ????? ?????."
        if _wants_json():
            return JSONResponse({"ok": False, "message": msg}, status_code=400)
        return RedirectResponse("/bank?msg=" + quote(msg), status_code=302)

    payload = await _read_payload()
    memo_note = str(payload.get("note") or payload.get("memo") or "").strip()

    amount = _num(
        _get(tx, [
            "deposit_amount", "amount", "paid_amount",
            "in_amount", "txn_amount", "money"
        ])
    )

    member_id = _get(tx, [
        "matched_member_id", "member_id",
        "target_member_id", "mid"
    ])

    if not member_id:
        msg = "??? ???? ?? ??? ? ????."
        if _wants_json():
            return JSONResponse({"ok": False, "message": msg}, status_code=400)
        return RedirectResponse("/bank?status=" + quote(UNMATCHED) + "&msg=" + quote(msg), status_code=302)

    member = db.query(Member).filter(Member.id == int(member_id)).first()
    if not member:
        msg = "?? ??? ?? ? ????."
        if _wants_json():
            return JSONResponse({"ok": False, "message": msg}, status_code=404)
        return RedirectResponse("/bank?msg=" + quote(msg), status_code=302)

    old_arrears = _num(getattr(member, "excel_arrears", 0))
    new_arrears = max(0, old_arrears - amount)

    if hasattr(member, "excel_arrears"):
        member.excel_arrears = new_arrears

    # ?? ??? ?? ??? ??? ??? ??
    today = _dt.date.today()
    tx_date = _get(tx, ["txn_date", "deposit_date", "paid_date", "date"])
    for attr in ["last_paid_date", "last_payment_date", "paid_date"]:
        if hasattr(member, attr):
            try:
                setattr(member, attr, tx_date or today)
            except Exception:
                pass

    # ?? ?? ????
    if hasattr(tx, "match_status"):
        tx.match_status = DONE

    if hasattr(tx, "applied"):
        tx.applied = True

    if hasattr(tx, "applied_at"):
        tx.applied_at = _dt.datetime.now()

    old_reason = str(getattr(tx, "match_reason", "") or "")
    add_reason = f"????: {amount:,}? / ??? {old_arrears:,}? / ??? {new_arrears:,}?"
    if memo_note:
        add_reason += " / ??: " + memo_note

    if hasattr(tx, "match_reason"):
        tx.match_reason = (old_reason + " / " if old_reason else "") + add_reason

    db.add(member)
    db.add(tx)
    db.commit()

    msg = f"?? {amount:,}? ????. ??? {old_arrears:,}? ? {new_arrears:,}?"

    if _wants_json():
        return JSONResponse({
            "ok": True,
            "message": msg,
            "amount": amount,
            "before": old_arrears,
            "after": new_arrears,
            "status": DONE,
        })

    referer = request.headers.get("referer") or "/bank"
    return RedirectResponse(referer, status_code=302)


@app.get("/income-ledger/{tid}/edit", response_class=HTMLResponse)
def income_ledger_edit_form(
    tid: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        raise HTTPException(404)

    _ensure_income_ledger_details(db)

    MISC = "잡수입"
    SUSP = "가수금"

    current_status = str(getattr(tx, "match_status", "") or "")
    kind_code = "misc" if current_status == MISC else "suspense"

    detail = None
    try:
        detail = db.query(IncomeLedgerDetail).filter(IncomeLedgerDetail.bank_transaction_id == tid).first()
    except Exception as e:
        print("income detail load error:", e)
        detail = None

    parsed = {}
    try:
        parsed = _parse_income_reason(str(getattr(tx, "match_reason", "") or ""))
    except Exception:
        parsed = {}

    work_type = ""
    pending_target = "없음"
    related_vehicle_no = ""
    related_name = ""
    note = ""
    next_billing_date = ""

    if detail:
        work_type = str(getattr(detail, "work_type", "") or "")
        pending_target = str(getattr(detail, "pending_target", "") or "없음")
        related_vehicle_no = str(getattr(detail, "related_vehicle_no", "") or "")
        related_name = str(getattr(detail, "related_name", "") or "")
        note = str(getattr(detail, "note", "") or "")
        next_billing_date = str(getattr(detail, "next_billing_date", "") or "")
    else:
        work_type = parsed.get("work_reason") or ""
        related_vehicle_no = parsed.get("vehicle_no") or ""
        related_name = parsed.get("name") or ""
        note = parsed.get("note") or ""
        pending_target = _calc_pending_target(current_status, work_type)

        txn_date = str(getattr(tx, "txn_date", "") or "")[:10]
        if pending_target in ["협회비", "관리비"]:
            next_billing_date = _calc_next_billing_date(txn_date)

    return templates.TemplateResponse(request, "income_edit_form.html", {
        "request": request,
        "user": user,
        "tx": tx,
        "kind_code": kind_code,
        "work_type": work_type,
        "pending_target": pending_target,
        "related_vehicle_no": related_vehicle_no,
        "related_name": related_name,
        "note": note,
        "next_billing_date": next_billing_date,
        "fmt_amt": fmt_amt,
    })


@app.post("/income-ledger/{tid}/edit")
def income_ledger_edit_save(
    tid: int,
    kind: str = Form(...),
    work_type: str = Form(""),
    pending_target: str = Form("없음"),
    related_vehicle_no: str = Form(""),
    related_name: str = Form(""),
    next_billing_date: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    from sqlalchemy import text as _t

    _ensure_income_ledger_details(db)

    MISC = "잡수입"
    SUSP = "가수금"

    raw_kind = str(kind or "").strip().lower()
    new_status = MISC if raw_kind in ["misc", "misc_income", "잡수입"] else SUSP

    tx = db.query(BankTransaction).filter(BankTransaction.id == tid).first()
    if not tx:
        raise HTTPException(404)

    work_type = str(work_type or "").strip()
    pending_target = str(pending_target or "없음").strip()
    if pending_target not in ["없음", "협회비", "관리비"]:
        pending_target = _calc_pending_target(new_status, work_type)

    txn_date = str(getattr(tx, "txn_date", "") or "")[:10]
    if pending_target in ["협회비", "관리비"] and not next_billing_date:
        next_billing_date = _calc_next_billing_date(txn_date)
    if pending_target == "없음":
        next_billing_date = ""

    tx.match_status = new_status
    if hasattr(tx, "matched_member_id"):
        tx.matched_member_id = None

    parts = ["수동분류 수정: " + new_status]
    if work_type:
        parts.append("업무사유: " + work_type)
    if pending_target:
        parts.append("처리대상: " + pending_target)
    if related_vehicle_no:
        parts.append("관련차량: " + related_vehicle_no.strip())
    if related_name:
        parts.append("관련성명: " + related_name.strip())
    if next_billing_date:
        parts.append("다음부과일: " + next_billing_date.strip())
    if note:
        parts.append("비고: " + note.strip())
    tx.match_reason = " / ".join(parts)
    _upsert_income_ledger_detail_from_bank(
        db=db,
        tx=tx,
        income_type=new_status,
        work_type=locals().get("reason", locals().get("work_type", "")),
        related_vehicle_no=locals().get("related_vehicle_no", ""),
        related_name=locals().get("related_name", ""),
        note=locals().get("note", ""),
        billing_date=locals().get("billing_date", locals().get("next_billing_date", "")),
    )

    detail = None
    try:
        detail = db.query(IncomeLedgerDetail).filter(IncomeLedgerDetail.bank_transaction_id == tid).first()
    except Exception:
        detail = None

    if detail:
        detail.income_type = new_status
        detail.work_type = work_type
        detail.pending_target = pending_target
        detail.related_vehicle_no = related_vehicle_no
        detail.related_name = related_name
        detail.note = note
        detail.next_billing_date = next_billing_date
        db.add(detail)
    else:
        db.execute(_t("""
            INSERT INTO income_ledger_details
            (bank_transaction_id, income_type, work_type, pending_target,
             related_vehicle_no, related_name, note, next_billing_date)
            VALUES
            (:tid, :income_type, :work_type, :pending_target,
             :related_vehicle_no, :related_name, :note, :next_billing_date)
        """), {
            "tid": tid,
            "income_type": new_status,
            "work_type": work_type,
            "pending_target": pending_target,
            "related_vehicle_no": related_vehicle_no,
            "related_name": related_name,
            "note": note,
            "next_billing_date": next_billing_date,
        })

    db.add(tx)
    db.commit()

    return RedirectResponse(
        "/income-ledger?kind=" + quote(new_status),
        status_code=302
    )


def _ensure_pending_board_table(db):
    from sqlalchemy import text
    dialect = getattr(getattr(db, "bind", None), "dialect", None)
    name = getattr(dialect, "name", "")

    if name == "postgresql":
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS bank_income_pending_queue (
                id SERIAL PRIMARY KEY,
                bank_transaction_id INTEGER,
                process_type VARCHAR(50),
                income_kind VARCHAR(50),
                txn_date VARCHAR(20),
                amount INTEGER,
                memo TEXT,
                related_vehicle_no VARCHAR(100),
                related_name VARCHAR(100),
                reason TEXT,
                note TEXT,
                status VARCHAR(50) DEFAULT '반영대기',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
    else:
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS bank_income_pending_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_transaction_id INTEGER,
                process_type VARCHAR(50),
                income_kind VARCHAR(50),
                txn_date VARCHAR(20),
                amount INTEGER,
                memo TEXT,
                related_vehicle_no VARCHAR(100),
                related_name VARCHAR(100),
                reason TEXT,
                note TEXT,
                status VARCHAR(50) DEFAULT '반영대기',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))




def _parse_income_reason(reason):
    import re as _re
    result = {'work_reason': '', 'vehicle_no': '', 'name': '', 'note': ''}
    if not reason: return result
    LABELS = [
        (r'사유', 'work_reason'),
        (r'업무사유', 'work_reason'),
        (r'관련차량', 'vehicle_no'),
        (r'차량', 'vehicle_no'),
        (r'관련성명', 'name'),
        (r'성명', 'name'),
        (r'비고', 'note'),
    ]
    parts = [p.strip() for p in str(reason).split('/')]
    for part in parts:
        for label, key in LABELS:
            m = _re.match(r'.*' + label + r'[:\s]+(.+)', part)
            if m and not result[key]:
                result[key] = m.group(1).strip()
                break
    return result

def _calc_next_billing_date(base_date_str):
    import re as _re
    from datetime import date as _date
    if not base_date_str: return ''
    s = str(base_date_str)[:10].strip()
    m = _re.match(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', s)
    if not m: return ''
    try:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if mo == 12: ny, nm = y+1, 1
        else: ny, nm = y, mo+1
        import calendar
        max_d = calendar.monthrange(ny, nm)[1]
        return '%04d-%02d-%02d' % (ny, nm, min(d, max_d))
    except: return ''

ALL_PROCESS_TYPES = [
    "폐업", "폐지", "양도", "이관", "탈퇴",
    "사망", "말소", "협회가입", "택배신규",
    "관리비폐지", "현역복구", "70세",
    "협회비", "관리비",
]
ALL_TABS = ["전체", "반영대기",
    "폐업", "폐지", "양도", "이관", "탈퇴",
    "사망", "말소", "협회가입", "택배신규",
    "관리비폐지", "현역복구", "70세",
    "협회비", "관리비", "반영완료"]
PROCESS_NORM = {
    "타도": "이관",
    "폐업": "폐업",
    "폐지": "폐지",
    "포업": "폐업",
    "포지": "폐지",
    "협회가입원": "협회가입",
    "반영대기-협회비": "협회비",
    "반영대기-관리비": "관리비",
}
# BillingPerson process_type -> 반영대기 탭용 최종 처리구분 변환

def _norm_pending_key(process_type="", vehicle_no="", name=""):
    pt = str(process_type or "").strip()
    pt = PROCESS_NORM.get(pt, pt)
    pt = BILLING_TO_PENDING_PT.get(pt, pt)

    v = str(vehicle_no or "").strip()
    v = v.replace(" ", "").replace("-", "").replace("호", "")

    n = str(name or "").strip().replace(" ", "")

    return (pt, v, n)


def _existing_pending_keys_for_billing(db):
    """
    처리대기목록에 이미 있는 사람 키 수집.
    통장매칭/잡수입·가수금으로 만든 협회비/관리비 후보를 우선 보호한다.
    """
    keys = set()

    try:
        # 기존 BillingPerson
        for bp in db.query(BillingPerson).all():
            keys.add(_norm_pending_key(
                getattr(bp, "process_type", "") or "",
                getattr(bp, "vehicle_no", "") or "",
                getattr(bp, "name", "") or "",
            ))
    except Exception as e:
        print("existing pending keys BillingPerson error:", e)

    try:
        _ensure_income_ledger_details(db)
        details = db.query(IncomeLedgerDetail).filter(
            IncomeLedgerDetail.pending_target.in_(["협회비", "관리비"])
        ).all()

        for d in details:
            keys.add(_norm_pending_key(
                getattr(d, "pending_target", "") or "",
                getattr(d, "related_vehicle_no", "") or "",
                getattr(d, "related_name", "") or "",
            ))
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print("existing pending keys IncomeLedgerDetail error:", e)

    try:
        rows = db.execute(_text("SELECT * FROM bank_income_pending_queue")).mappings().all()
        for r in rows:
            keys.add(_norm_pending_key(
                r.get("process_type", "") or r.get("income_kind", "") or "",
                r.get("related_vehicle_no", "") or "",
                r.get("related_name", "") or "",
            ))
    except Exception:
        pass

    # 빈 키 제거
    keys = {k for k in keys if k[1] or k[2]}
    return keys



BILLING_TO_PENDING_PT = {
    "협회가입": "협회비",
    "협회가입원": "협회비",
    "택배신규": "관리비",
    "자격증명발급": "관리비",
    "신규관리": "관리비",
}

def _safe_json_dict(v):
    try:
        import json
        if not v:
            return {}
        if isinstance(v, dict):
            return v
        return json.loads(v)
    except Exception:
        return {}


def _extract_korean_date_text(v):
    """
    비고/원문에서 날짜만 추출.
    26.04.29. / 2026-04-29 / 18. 5. 1. / 94.3.8. 대응.
    """
    import re
    t = str(v or "")
    if not t:
        return ""

    m = re.search(r"(20\d{2}|19\d{2})[.\-/년\s]+(\d{1,2})[.\-/월\s]+(\d{1,2})", t)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"

    m = re.search(r"(?<!\d)(\d{2})\s*[.\-/]\s*(\d{1,2})\s*[.\-/]\s*(\d{1,2})", t)
    if m:
        yy, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = 2000 + yy if yy <= 30 else 1900 + yy
        return f"{y:04d}-{mo:02d}-{d:02d}"

    return ""


def _bp_request_and_next_dates(bp, final_process_type):
    """
    부과대수 엑셀 행의 업무 기준 날짜 계산.
    관리비: 인가일자 우선, 없으면 자격증명발급일자
    협회비: 가입일자
    제외/폐지류: 비고 날짜
    """
    raw = _safe_json_dict(getattr(bp, "raw_data", "") or "")
    raw_data = raw.get("raw_data") if isinstance(raw.get("raw_data"), dict) else {}

    def pick(*keys):
        for k in keys:
            v = raw.get(k, "")
            if v:
                return str(v)
        return ""

    def pick_raw_contains(*names):
        for _, v in raw_data.items():
            txt = str(v or "")
            for name in names:
                if name in txt:
                    d = _extract_korean_date_text(txt)
                    if d:
                        return d
        return ""

    pt = str(final_process_type or "").strip()

    request_date = ""
    next_billing_date = ""

    if pt == "관리비":
        request_date = (
            _extract_korean_date_text(pick("permit_date", "인가일자", "허가일자"))
            or _extract_korean_date_text(pick("cert_issue_date", "자격증명발급일자", "자격증명일자"))
            or pick_raw_contains("인가", "허가", "자격증명")
        )
        if request_date:
            next_billing_date = _calc_next_billing_date(request_date)

    elif pt == "협회비":
        request_date = (
            _extract_korean_date_text(pick("join_date", "가입일자", "협회가입일"))
            or pick_raw_contains("가입")
        )
        if request_date:
            next_billing_date = _calc_next_billing_date(request_date)

    else:
        note = pick("note", "비고", "메모")
        request_date = (
            _extract_korean_date_text(note)
            or pick_raw_contains("폐지", "폐업", "양도", "이관", "탈퇴", "말소")
        )
        next_billing_date = ""

    if not request_date:
        request_date = str(getattr(bp, "created_at", "") or "")[:10]

    return request_date, next_billing_date



INCOME_KEYWORDS_ASSOC = [
    "가입비", "특별회비", "협회가입", "협회비", "가입", "입회"
]
MGMT_KEYWORDS = [
    "자격증명", "자격증명발급", "택배신규", "신규관리", "관리비"
]

@app.get("/work", response_class=HTMLResponse)
@app.get("/work/pending-board", response_class=HTMLResponse)
def pending_board_page(
    request: Request,
    tab: str = "전체",
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text as _text
    from urllib.parse import quote
    _ensure_pending_board_table(db)

    # SAFE TAB NORMALIZE
    tab = str(tab or "전체").strip()
    if not tab:
        tab = "전체"
    rows = []

    # 1) BillingPerson
    try:
        bps = (
            db.query(BillingPerson)
            .filter(BillingPerson.reflect_status.in_(["처리대기", "반영대기", "반영완료", "보류"]))
            .order_by(BillingPerson.id.desc())
            .all()
        )
        for bp in bps:
            pt_raw = str(getattr(bp, "process_type", "") or "")
            pt_norm = PROCESS_NORM.get(pt_raw, pt_raw)
            # 협회가입/택배신규 등은 반영대기 탭에서 협회비/관리비로 표시
            pt = BILLING_TO_PENDING_PT.get(pt_norm, pt_norm)
            st = str(getattr(bp, "reflect_status", "") or "반영대기")
            if st in ["처리대기", "부과대수상세"] and pt in ["협회비", "관리비"]:
                st = "반영대기"
            elif st == "처리대기":
                st = "반영대기"
            rd, nb_date = _bp_request_and_next_dates(bp, pt)
            acct_map = {"협회비": "협", "관리비": "관"}
            acct = acct_map.get(pt, getattr(bp, "account", "") or "")
            rows.append({
                "row_type": "bp", "id": bp.id, "status": st, "process_type": pt,
                "region": getattr(bp, "region", "") or "",
                "name": getattr(bp, "name", "") or "",
                "vehicle_no": getattr(bp, "vehicle_no", "") or "",
                "account": acct,
                "before_arrears": None,
                "request_date": rd,
                "next_billing_date": nb_date,
                "source": "부과대수업로드",
                "source_sheet": getattr(bp, "source_sheet", "") or "",
                "reason": getattr(bp, "note", "") or "", "note": pt_raw,
            })
    except Exception as e:
        print("BillingPerson error:", e)

    # 2) WorkQueue
    try:
        wqs = db.query(WorkQueue).order_by(WorkQueue.id.desc()).all()
        for w in wqs:
            m = db.query(Member).filter(Member.id == w.member_id).first() if w.member_id else None
            pt_raw = str(getattr(w, "process_type", "") or "")
            pt = PROCESS_NORM.get(pt_raw, pt_raw)
            st = str(getattr(w, "status", "") or "반영대기")
            rd = str(getattr(w, "submitted_at", "") or "")[:10]
            rows.append({
                "row_type": "wq", "id": w.id, "status": st, "process_type": pt,
                "region": getattr(m, "region", "") if m else "",
                "name": getattr(m, "name", "") if m else "",
                "vehicle_no": getattr(m, "vehicle_no", "") if m else "",
                "account": getattr(m, "account", "") if m else "",
                "before_arrears": int(getattr(w, "arrears_at_submit", 0) or 0),
                "request_date": rd,
                "next_billing_date": "",
                "source": "미수금명단",
                "source_sheet": getattr(w, "source_screen", "") or "",
                "reason": getattr(w, "reason", "") or "",
                "note": getattr(w, "note", "") or "",
            })
    except Exception as e:
        print("WorkQueue error:", e)

    # 3) IncomeLedgerDetail -> 협회비/관리비 반영대기 후보
    try:
        _ensure_income_ledger_details(db)
        details = db.query(IncomeLedgerDetail).filter(
            IncomeLedgerDetail.pending_target.in_(["협회비", "관리비"])
        ).order_by(IncomeLedgerDetail.id.desc()).all()
        for d in details:
            pt = d.pending_target or ""
            if pt not in ["협회비", "관리비"]:
                continue
            tx = d.bank_transaction
            txn_date = str(getattr(d, "next_billing_date", "") or "")
            req_date = ""
            if tx:
                req_date = str(getattr(tx, "txn_date", "") or "")[:10]
                if not txn_date:
                    txn_date = _calc_next_billing_date(req_date)
            acct = "협" if pt == "협회비" else "관"
            rsn = ("가입비/특별회비 확인 후 협회비 부과대상 반영 필요"
                   if pt == "협회비" else
                   "자격증명발급 확인 후 관리비 부과대상 반영 필요")
            rows.append({
                "row_type": "income", "id": d.id, "status": "반영대기",
                "process_type": pt, "region": "",
                "name": d.related_name or "",
                "vehicle_no": d.related_vehicle_no or "",
                "account": acct,
                "before_arrears": None,
                "request_date": req_date,
                "next_billing_date": txn_date,
                "source": d.income_type or "",
                "source_sheet": "",
                "reason": rsn,
                "note": d.work_type or "",
            })
        # 상세 없는 BankTransaction도 보조로 읽기 (detail 없는 것만)
        detail_tx_ids = {d.bank_transaction_id for d in details if d.bank_transaction_id}
        btxs = db.query(BankTransaction).filter(
            BankTransaction.match_status.in_(["잡수입", "가수금"])
        ).order_by(BankTransaction.id.desc()).all()
        for tx in btxs:
            if tx.id in detail_tx_ids:
                continue
            ms = str(getattr(tx, "match_status", "") or "")
            reason = str(getattr(tx, "match_reason", "") or "")
            memo = str(getattr(tx, "memo", "") or "")
            txn_date = str(getattr(tx, "txn_date", "") or "")[:10]
            parsed = _parse_income_reason(reason)
            combined = reason + " " + memo
            pt = None
            if ms == "가수금":
                if any(kw in combined for kw in INCOME_KEYWORDS_ASSOC):
                    pt = "협회비"
            elif ms == "잡수입":
                if any(kw in combined for kw in MGMT_KEYWORDS):
                    pt = "관리비"
            if not pt:
                continue
            acct = "협" if pt == "협회비" else "관"
            rsn = ("가입비/특별회비 확인 후 협회비 부과대상 반영 필요"
                   if pt == "협회비" else
                   "자격증명발급 확인 후 관리비 부과대상 반영 필요")
            rows.append({
                "row_type": "income", "id": tx.id, "status": "반영대기",
                "process_type": pt, "region": "",
                "name": parsed.get("name") or "",
                "vehicle_no": parsed.get("vehicle_no") or memo,
                "account": acct,
                "before_arrears": None,
                "request_date": txn_date,
                "next_billing_date": _calc_next_billing_date(txn_date),
                "source": ms, "source_sheet": "",
                "reason": rsn,
                "note": parsed.get("work_reason") or "",
            })
    except Exception as e:
        print("income/suspense error:", e)

    # 4) bank_income_pending_queue
    try:
        raw_queue = db.execute(_text("SELECT * FROM bank_income_pending_queue ORDER BY id DESC")).mappings().all()
        for r in raw_queue:
            rd2 = dict(r)
            pt_raw = str(rd2.get("process_type", "") or rd2.get("income_kind", "") or "")
            pt = PROCESS_NORM.get(pt_raw, pt_raw)
            rdate = str(rd2.get("txn_date", "") or "")[:10]
            rows.append({
                "row_type": "queue", "id": rd2.get("id", 0),
                "status": str(rd2.get("status", "") or "반영대기"),
                "process_type": pt, "region": "",
                "name": str(rd2.get("related_name", "") or ""),
                "vehicle_no": str(rd2.get("related_vehicle_no", "") or ""),
                "account": "",
                "before_arrears": None,
                "request_date": rdate,
                "next_billing_date": _calc_next_billing_date(rdate),
                "source": "통장입금", "source_sheet": "",
                "reason": str(rd2.get("reason", "") or ""),
                "note": str(rd2.get("note", "") or ""),
            })
    except Exception:
        pass

    def row_text(r):
        return " ".join(str(r.get(k) or "") for k in [
            "status","process_type","region","name","vehicle_no","account","source","reason","note"])

    filtered = rows
    if tab and tab != "전체":
        if tab == "반영대기":
            filtered = [r for r in rows if r["status"] == "반영대기"
                        and r["process_type"] in ["협회비", "관리비"]]
        elif tab == "반영완료":
            filtered = [r for r in rows if r["status"] == "반영완료"]
        else:
            filtered = [r for r in rows if r["process_type"] == tab]
    if q:
        filtered = [r for r in filtered if q in row_text(r)]

    counts = {"전체": len(rows)}
    for t in ALL_TABS:
        if t == "전체": continue
        if t == "반영대기":
            counts[t] = sum(1 for r in rows if r["status"] == "반영대기"
                            and r["process_type"] in ["협회비", "관리비"])
        elif t == "반영완료":
            counts[t] = sum(1 for r in rows if r["status"] == "반영완료")
        else:
            counts[t] = sum(1 for r in rows if r["process_type"] == t)

    return templates.TemplateResponse(request, "pending_board.html", {
        "request": request, "user": user, "rows": filtered,
        "tabs": ALL_TABS, "tab": tab, "q": q, "counts": counts,
        "fmt_amt": fmt_amt, "msg": request.query_params.get("msg", ""),
    })

@app.get("/work/pending-board/add", response_class=HTMLResponse)
def pending_board_add_form(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    return templates.TemplateResponse(request, "pending_board_add.html", {
        "request": request,
        "user": user,
    })


@app.post("/work/pending-board/add")
def pending_board_add_save(
    income_kind: str = Form(...),
    txn_date: str = Form(""),
    amount: int = Form(0),
    memo: str = Form(""),
    related_vehicle_no: str = Form(""),
    related_name: str = Form(""),
    reason: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text
    from urllib.parse import quote

    _ensure_pending_board_table(db)

    allowed = ["????-???", "????-???"]
    if income_kind not in allowed:
        income_kind = "????-???"

    db.execute(text("""
        INSERT INTO bank_income_pending_queue
        (bank_transaction_id, process_type, income_kind, txn_date, amount, memo,
         related_vehicle_no, related_name, reason, note, status)
        VALUES
        (NULL, '????', :income_kind, :txn_date, :amount, :memo,
         :related_vehicle_no, :related_name, :reason, :note, '????')
    """), {
        "income_kind": income_kind,
        "txn_date": txn_date,
        "amount": int(amount or 0),
        "memo": memo,
        "related_vehicle_no": related_vehicle_no,
        "related_name": related_name,
        "reason": reason,
        "note": note,
    })

    db.commit()

    return RedirectResponse(
        "/work?tab=" + quote("????"),
        status_code=302
    )


# =========================================================
# ????: ???? ????
# =========================================================
@app.post("/bank/apply-auto")
@app.post("/bank/auto-apply")
@app.post("/bank/apply-all")
@app.post("/bank/apply-auto-all")
def bank_apply_auto_all_safe(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote
    import datetime as _dt
    import re as _re

    AUTO = "\uc790\ub3d9\ub9e4\uce6d"      # ????
    DONE = "\ubc18\uc601\uc644\ub8cc"      # ????

    def _num(v):
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).replace(",", "").replace("₩", "").replace("�", "").strip()
        m = _re.search(r"-?\d+", s)
        return int(m.group(0)) if m else 0

    def _get(obj, names):
        for n in names:
            if hasattr(obj, n):
                v = getattr(obj, n)
                if v not in [None, ""]:
                    return v
        return None

    txs = db.query(BankTransaction).all()

    applied_count = 0
    skipped_count = 0
    total_amount = 0

    for tx in txs:
        status = str(getattr(tx, "match_status", "") or "")
        if status != AUTO:
            continue

        if getattr(tx, "applied", False) or status == DONE:
            skipped_count += 1
            continue

        member_id = _get(tx, ["matched_member_id", "member_id", "target_member_id", "mid"])
        if not member_id:
            skipped_count += 1
            continue

        m = db.query(Member).filter(Member.id == int(member_id)).first()
        if not m:
            skipped_count += 1
            continue

        amount = _num(_get(tx, ["deposit_amount", "amount", "paid_amount", "in_amount", "txn_amount", "money"]))
        old_arr = _num(getattr(m, "excel_arrears", 0))
        new_arr = max(0, old_arr - amount)

        if hasattr(m, "excel_arrears"):
            m.excel_arrears = new_arr

        for attr in ["last_paid_date", "last_payment_date", "paid_date"]:
            if hasattr(m, attr):
                try:
                    setattr(m, attr, _get(tx, ["txn_date", "deposit_date", "date"]) or _dt.date.today())
                except Exception:
                    pass

        tx.match_status = DONE
        if hasattr(tx, "applied"):
            tx.applied = True
        if hasattr(tx, "applied_at"):
            tx.applied_at = _dt.datetime.now()

        old_reason = str(getattr(tx, "match_reason", "") or "")
        add_reason = f"[자동반영] 입금 {amount:,}원 / 전 {old_arr:,}원 / 후 {new_arr:,}원"
        if hasattr(tx, "match_reason"):
            tx.match_reason = (old_reason + " / " if old_reason else "") + add_reason

        db.add(m)
        db.add(tx)

        applied_count += 1
        total_amount += amount

    db.commit()

    msg = f"자동반영 {applied_count}건 완료 / 합계 {total_amount:,}원 / 제외 {skipped_count}건"
    return RedirectResponse(
        "/bank?status=" + quote(DONE) + "&msg=" + quote(msg),
        status_code=302
    )


# =========================================================
# ????: ???? ???
# ??? ???? ??. include_applied/on ?? ? ?????? ???.
# =========================================================
@app.post("/bank/reset-match-status")
@app.post("/bank/reset-matches")
@app.post("/bank/reset-match")
@app.post("/bank/reset-status")
@app.post("/bank/reset")
def bank_reset_match_status_safe(
    include_applied: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from urllib.parse import quote

    UNMATCHED = "\ubbf8\ub9e4\uce6d"      # ???
    DONE = "\ubc18\uc601\uc644\ub8cc"     # ????

    include_done = str(include_applied or "").lower() in ["on", "true", "1", "yes", "y"]

    txs = db.query(BankTransaction).all()

    reset_count = 0
    protected_count = 0

    for tx in txs:
        status = str(getattr(tx, "match_status", "") or "")

        if not include_done and (getattr(tx, "applied", False) or status == DONE):
            protected_count += 1
            continue

        if hasattr(tx, "match_status"):
            tx.match_status = UNMATCHED
        if hasattr(tx, "matched_member_id"):
            tx.matched_member_id = None
        if hasattr(tx, "match_reason"):
            tx.match_reason = "매칭상태 초기화"

        # ?????? ?? ??? ???? applied? ???.
        # ?, ?? ??? ??? ??? ???? ? ??? ?? ??.
        if include_done:
            if hasattr(tx, "applied"):
                tx.applied = False
            if hasattr(tx, "applied_at"):
                tx.applied_at = None

        db.add(tx)
        reset_count += 1

    db.commit()

    msg = f"매칭상태 초기화 {reset_count}건 / 보호 {protected_count}건"
    return RedirectResponse(
        "/bank?status=" + quote(UNMATCHED) + "&msg=" + quote(msg),
        status_code=302
    )



# ── FORCE DEBUG: billing upload stats ─────────────────────────────
@app.get("/api/debug/billing-upload-stats-force")
def api_debug_billing_upload_stats_force(
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):
    from sqlalchemy import text, inspect

    result = {
        "ok": True,
        "db": "",
        "tables": [],
        "target_table": None,
        "columns": [],
        "total": 0,
        "by_month_sheet_process": [],
        "by_process": [],
        "error": None,
    }

    try:
        bind = db.get_bind()
        result["db"] = getattr(getattr(bind, "dialect", None), "name", "")

        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result["tables"] = [t for t in tables if "billing" in t.lower() or "work" in t.lower()]

        target = None
        for cand in [
            "billing_people",
            "billing_persons",
            "billing_person",
            "work_queue",
            "work_queues",
            "billing_generated_arrears"
        ]:
            if cand in tables:
                target = cand
                break

        if not target:
            result["ok"] = False
            result["error"] = "billing/work 관련 대상 테이블을 찾지 못했습니다."
            return result

        result["target_table"] = target

        cols = [c["name"] for c in inspector.get_columns(target)]
        result["columns"] = cols

        result["total"] = db.execute(text(f"SELECT COUNT(*) FROM {target}")).scalar() or 0

        source_year_col = "source_year" if "source_year" in cols else "''"
        source_month_col = "source_month" if "source_month" in cols else "''"
        source_sheet_col = "source_sheet" if "source_sheet" in cols else "''"
        process_col = "process_type" if "process_type" in cols else ("account" if "account" in cols else "''")

        rows = db.execute(text(f"""
            SELECT 
                COALESCE(CAST({source_year_col} AS TEXT),'') AS source_year,
                COALESCE(CAST({source_month_col} AS TEXT),'') AS source_month,
                COALESCE(CAST({source_sheet_col} AS TEXT),'') AS source_sheet,
                COALESCE(CAST({process_col} AS TEXT),'') AS process_type,
                COUNT(*) AS cnt
            FROM {target}
            GROUP BY source_year, source_month, source_sheet, process_type
            ORDER BY source_year, source_month, source_sheet, process_type
        """)).mappings().all()

        result["by_month_sheet_process"] = [dict(r) for r in rows]

        rows2 = db.execute(text(f"""
            SELECT COALESCE(CAST({process_col} AS TEXT),'') AS process_type, COUNT(*) AS cnt
            FROM {target}
            GROUP BY process_type
            ORDER BY process_type
        """)).mappings().all()

        result["by_process"] = [dict(r) for r in rows2]

        return result

    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        result["ok"] = False
        result["error"] = repr(e)
        return result


