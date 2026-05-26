
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from io import BytesIO
from datetime import datetime, date
from calendar import monthrange
import re
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/reports", tags=["reports-export"])


def _get_db():
    """
    프로젝트마다 DB 연결 방식이 조금 달라서 최대한 호환되게 처리.
    database.py 안에 SessionLocal이 있으면 그걸 우선 사용.
    """
    try:
        from database import SessionLocal
        db = SessionLocal()
        try:
            yield db
        finally:
            db.close()
    except Exception:
        yield None


def _safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def _parse_date(v):
    if v is None or v == "":
        return None

    # datetime은 date보다 먼저 처리해야 함.
    # datetime도 date의 하위 타입이라 순서가 중요함.
    if isinstance(v, datetime):
        return v.date()

    if isinstance(v, date):
        return v

    s = str(v).strip()
    if not s:
        return None

    s = s.replace(".", "-").replace("/", "-")
    s = re.sub(r"\s+.*$", "", s)

    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y%m%d", "%Y.%m.%d"):
        try:
            d = datetime.strptime(s, fmt)
            return d.date()
        except Exception:
            pass

    m = re.search(r"(20\d{2}|19\d{2})[-.]?(\d{1,2})[-.]?(\d{1,2})?", s)
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        da = int(m.group(3) or 1)
        try:
            return date(y, mo, da)
        except Exception:
            return None

    return None


def _month_range(year, month):
    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])
    return start, end


def _next_month_after(d):
    if not d:
        return None
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _norm_col(name):
    return str(name or "").strip().lower().replace(" ", "").replace("_", "")


def _find_col(columns, candidates):
    norm_map = {_norm_col(c): c for c in columns}
    for cand in candidates:
        key = _norm_col(cand)
        if key in norm_map:
            return norm_map[key]
    for c in columns:
        nc = _norm_col(c)
        for cand in candidates:
            if _norm_col(cand) in nc:
                return c
    return None


def _get_tables_and_rows(db, mode="all", limit=3000):
    """
    DB 전체를 무작정 다 읽으면 Railway에서 500/timeout이 날 수 있어서
    mode에 따라 필요한 테이블만 제한적으로 읽음.
    mode="deposit"이면 입금/통장/거래/납부 관련 테이블만 읽음.
    """
    if db is None:
        return []

    try:
        bind = db.get_bind()
        from sqlalchemy import inspect, text
        inspector = inspect(bind)
        tables = inspector.get_table_names()
        result = []

        deposit_words = [
            "payment", "deposit", "bank", "transaction", "paid", "pay",
            "입금", "통장", "거래", "수납", "납부"
        ]

        for t in tables:
            try:
                low_t = t.lower()
                if low_t.startswith("alembic"):
                    continue

                cols = [c["name"] for c in inspector.get_columns(t)]
                if not cols:
                    continue

                joined_cols = " ".join(cols).lower()

                if mode == "deposit":
                    if not any(w in low_t or w in joined_cols for w in deposit_words):
                        continue

                # 너무 많은 데이터를 한 번에 엑셀화하지 않도록 제한
                sql = text(f'SELECT * FROM "{t}" LIMIT {int(limit)}')
                rows = db.execute(sql).mappings().all()
                result.append({"table": t, "columns": cols, "rows": [dict(r) for r in rows]})
            except Exception as e:
                print("WARN: table read skipped", t, e)
                continue
        return result
    except Exception as e:
        print("WARN: get tables failed", e)
        return []


def _style_sheet(ws):
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAF3FF")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        max_len = 8
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 35)


def _safe_sheet_title(title):
    title = str(title or "sheet")
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        title = title.replace(ch, '_')
    title = title.replace(" ", "_")
    title = title[:31] or "sheet"
    return title


def _add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=_safe_sheet_title(title))
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    _style_sheet(ws)
    return ws


def _is_vehicle_like_row(row):
    keys = list(row.keys())
    return bool(
        _find_col(keys, ["차량번호", "vehicle_number", "car_no", "차량", "번호"]) or
        _find_col(keys, ["성명", "이름", "name", "owner_name"])
    )


def _row_to_member(row):
    keys = list(row.keys())

    col_region = _find_col(keys, ["지역", "region", "시군", "관할"])
    col_vehicle = _find_col(keys, ["차량번호", "vehicle_number", "car_no", "차량"])
    col_name = _find_col(keys, ["성명", "이름", "name", "owner_name"])
    col_phone = _find_col(keys, ["전화번호", "핸드폰", "휴대폰", "phone", "mobile"])
    col_approval = _find_col(keys, ["인가일자", "approval_date", "허가일자"])
    col_join = _find_col(keys, ["가입일자", "join_date", "협회가입일자"])
    col_cert = _find_col(keys, ["자격증명발급일자", "certificate_issue_date", "자격증명일자", "자격증명 발급일자"])
    col_cert_no = _find_col(keys, ["자격증명발급번호", "certificate_no", "자격증명번호"])
    col_status = _find_col(keys, ["상태", "status", "처리구분", "구분"])
    col_process = _find_col(keys, ["처리일자", "process_date", "변경일자", "폐업일자"])
    col_note = _find_col(keys, ["비고", "메모", "note", "remark"])

    return {
        "지역": _safe_str(row.get(col_region)),
        "차량번호": _safe_str(row.get(col_vehicle)),
        "성명": _safe_str(row.get(col_name)),
        "전화번호": _safe_str(row.get(col_phone)),
        "인가일자": _safe_str(row.get(col_approval)),
        "가입일자": _safe_str(row.get(col_join)),
        "자격증명발급일자": _safe_str(row.get(col_cert)),
        "자격증명발급번호": _safe_str(row.get(col_cert_no)),
        "처리구분": _safe_str(row.get(col_status)),
        "처리일자": _safe_str(row.get(col_process)),
        "상태": _safe_str(row.get(col_status)),
        "비고": _safe_str(row.get(col_note)),
        "_approval_date": _parse_date(row.get(col_approval)),
        "_join_date": _parse_date(row.get(col_join)),
        "_cert_date": _parse_date(row.get(col_cert)),
        "_process_date": _parse_date(row.get(col_process)),
    }


def _is_delivery_vehicle(m):
    car = m.get("차량번호", "")
    note = m.get("비고", "")
    status = m.get("상태", "")
    txt = f"{car} {note} {status}"
    return ("배" in car) or ("택배" in txt)


def _is_removed_status(m):
    txt = f"{m.get('처리구분','')} {m.get('상태','')} {m.get('비고','')}"
    removed_words = ["폐업", "폐지", "양도", "이관", "탈퇴", "타도", "사망"]
    return any(w in txt for w in removed_words)


def _is_joined(m):
    return bool(m.get("가입일자"))


def _billing_count_workbook(year, month, tables):
    start, end = _month_range(year, month)
    기준월 = f"{year}-{month:02d}"

    members = []
    for table in tables:
        for row in table["rows"]:
            if _is_vehicle_like_row(row):
                m = _row_to_member(row)
                if m.get("차량번호") or m.get("성명"):
                    members.append(m)

    # 중복 제거: 차량번호 + 성명 기준
    unique = {}
    for m in members:
        key = (m.get("차량번호"), m.get("성명"))
        unique[key] = m
    members = list(unique.values())

    active = [m for m in members if not _is_removed_status(m)]
    joined = [m for m in active if _is_joined(m)]
    delivery = [m for m in active if _is_delivery_vehicle(m)]

    # N월 택배관리: 인가일자 + 자격증명발급일자 둘 다 있고, 다음달부터 반영
    delivery_billable = []
    delivery_new = []
    for m in delivery:
        approval = m.get("_approval_date")
        cert = m.get("_cert_date")
        if approval and cert:
            base = max(approval, cert)
            bill_start = _next_month_after(base)
            if bill_start and bill_start <= start:
                delivery_billable.append(m)
            if bill_start and bill_start.year == year and bill_start.month == month:
                delivery_new.append(m)

    removed = [m for m in members if _is_removed_status(m)]
    transfer = [m for m in removed if "양도" in f"{m.get('처리구분','')} {m.get('상태','')} {m.get('비고','')}"]
    other_area = [m for m in removed if ("타도" in f"{m.get('처리구분','')} {m.get('상태','')} {m.get('비고','')}" or "이관" in f"{m.get('처리구분','')} {m.get('상태','')} {m.get('비고','')}")]
    closed = [m for m in removed if ("폐업" in f"{m.get('처리구분','')} {m.get('상태','')} {m.get('비고','')}" or "폐지" in f"{m.get('처리구분','')} {m.get('상태','')} {m.get('비고','')}")]
    withdrawn = [m for m in removed if "탈퇴" in f"{m.get('처리구분','')} {m.get('상태','')} {m.get('비고','')}"]

    # 70세는 주민번호/생년월일 컬럼이 있을 경우에만 잡힘. 없으면 0.
    age70 = []
    for m in active:
        txt = " ".join(str(v) for v in m.values())
        # 1956년 기준 2026년에 70세 도달. 연도별 자동.
        target_birth_year = year - 70
        if str(target_birth_year) in txt:
            age70.append(m)

    협회기본대수 = len(joined)
    n월택배관리 = len(delivery_billable)
    총부과대수 = 협회기본대수 + n월택배관리

    wb = Workbook()
    ws = wb.active
    ws.title = "부과대수 요약"

    headers = [
        "기준월", "협회기본대수", "협회가입", "양도", "타도", "폐업",
        "탈퇴", "택배신규", "관리비폐지", "70세", "N월 택배관리",
        "총부과대수", "비고"
    ]

    ws.append(headers)
    ws.append([
        기준월,
        협회기본대수,
        len(joined),
        len(transfer),
        len(other_area),
        len(closed),
        len(withdrawn),
        len(delivery_new),
        0,
        len(age70),
        n월택배관리,
        총부과대수,
        "자동 산정. 택배는 인가일자+자격증명발급일자 둘 다 있는 경우 다음 달부터 반영."
    ])
    _style_sheet(ws)

    detail_headers = [
        "지역", "차량번호", "성명", "전화번호", "인가일자", "가입일자",
        "자격증명발급일자", "자격증명발급번호", "처리구분", "처리일자", "상태", "비고"
    ]

    _add_sheet(wb, "협회가입 상세", detail_headers, joined)
    _add_sheet(wb, "양도 상세", detail_headers, transfer)
    _add_sheet(wb, "타도이관 상세", detail_headers, other_area)
    _add_sheet(wb, "폐업 상세", detail_headers, closed)
    _add_sheet(wb, "탈퇴 상세", detail_headers, withdrawn)
    _add_sheet(wb, "택배신규 상세", detail_headers, delivery_new)
    _add_sheet(wb, "관리비폐지 상세", detail_headers, [])
    _add_sheet(wb, "70세 전환 상세", detail_headers, age70)
    _add_sheet(wb, "N월 택배관리 상세", detail_headers, delivery_billable)

    return wb


def _is_payment_like_table(table_name, columns):
    t = table_name.lower()
    text = " ".join(columns).lower()
    words = ["payment", "deposit", "bank", "transaction", "입금", "통장", "거래", "수납", "납부"]
    return any(w in t or w in text for w in words)


def _payment_rows_for_month(year, month, tables):
    start, end = _month_range(year, month)
    out = []

    for table in tables:
        cols = table["columns"]
        if not _is_payment_like_table(table["table"], cols):
            continue

        col_date = _find_col(cols, ["입금일자", "거래일자", "납부일자", "date", "paid_at", "created_at", "transaction_date"])
        col_amount = _find_col(cols, ["입금액", "금액", "amount", "deposit_amount", "paid_amount"])
        col_memo = _find_col(cols, ["입금자명", "보낸분", "이체메모", "메모", "memo", "description", "내용"])
        col_name = _find_col(cols, ["성명", "이름", "name", "member_name"])
        col_vehicle = _find_col(cols, ["차량번호", "vehicle_number", "car_no"])
        col_account = _find_col(cols, ["계정", "구분", "account", "type", "category"])
        col_status = _find_col(cols, ["상태", "매칭상태", "status", "match_status"])

        if not col_date and not col_amount:
            continue

        for row in table["rows"]:
            d = _parse_date(row.get(col_date)) if col_date else None
            if d and not (start <= d <= end):
                continue

            # 날짜 컬럼이 없으면 일단 포함
            out.append({
                "기준월": f"{year}-{month:02d}",
                "자료테이블": table["table"],
                "입금일자": _safe_str(row.get(col_date)),
                "입금액": _safe_str(row.get(col_amount)),
                "입금자명/메모": _safe_str(row.get(col_memo)),
                "성명": _safe_str(row.get(col_name)),
                "차량번호": _safe_str(row.get(col_vehicle)),
                "계정": _safe_str(row.get(col_account)),
                "매칭상태": _safe_str(row.get(col_status)),
                "비고": "",
            })

    return out



def _error_workbook(title, error_message):
    wb = Workbook()
    ws = wb.active
    ws.title = "오류내용"
    ws.append(["구분", "내용"])
    ws.append(["오류", str(error_message)])
    ws.append(["안내", "서버 오류로 다운로드가 실패하지 않도록 오류 내용을 엑셀로 출력했습니다."])
    _style_sheet(ws)
    return wb



def _payment_rows_for_month_direct(db, year, month):
    """
    N? ???? ??.
    bank_transactions ????? ??? ??? ?? ????.
    LIMIT ?? ?? ? ??? ????.
    """
    from sqlalchemy import inspect, text as sql_text
    from datetime import datetime

    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    if "bank_transactions" not in tables:
        raise Exception("bank_transactions ???? ????.")

    cols = [c["name"] for c in inspector.get_columns("bank_transactions")]

    col_date = _find_col(cols, [
        "????", "????", "????",
        "transaction_date", "paid_at", "payment_date", "deposit_date",
        "date", "created_at"
    ])
    col_amount = _find_col(cols, [
        "???", "??", "amount", "deposit_amount", "paid_amount"
    ])
    col_memo = _find_col(cols, [
        "????", "???", "????", "??", "memo",
        "description", "??", "sender", "depositor", "payer_name"
    ])
    col_name = _find_col(cols, ["??", "??", "name", "member_name"])
    col_vehicle = _find_col(cols, ["????", "vehicle_number", "car_no"])
    col_account = _find_col(cols, ["??", "??", "account", "type", "category"])
    col_status = _find_col(cols, ["??", "????", "status", "match_status"])

    if not col_date:
        raise Exception("bank_transactions?? ????/????/created_at ??? ?? ?????.")

    start_dt = datetime(year, month, 1)
    if month == 12:
        end_dt = datetime(year + 1, 1, 1)
    else:
        end_dt = datetime(year, month + 1, 1)

    query = (
        'SELECT * FROM bank_transactions '
        'WHERE "' + col_date + '" >= :start_dt '
        'AND "' + col_date + '" < :end_dt '
        'ORDER BY "' + col_date + '" ASC'
    )

    rows = db.execute(
        sql_text(query),
        {"start_dt": start_dt, "end_dt": end_dt}
    ).mappings().all()

    out = []
    for row in rows:
        row = dict(row)
        out.append({
            "???": f"{year}-{month:02d}",
            "?????": "bank_transactions",
            "????": _safe_str(row.get(col_date)),
            "???": _safe_str(row.get(col_amount)),
            "????/??": _safe_str(row.get(col_memo)),
            "??": _safe_str(row.get(col_name)),
            "????": _safe_str(row.get(col_vehicle)),
            "??": _safe_str(row.get(col_account)),
            "????": _safe_str(row.get(col_status)),
            "??": "",
        })

    return out


def _deposit_workbook(year, month, tables):
    rows = _payment_rows_for_month(year, month, tables)

    wb = Workbook()
    ws = wb.active
    ws.title = "N월 입금추출"

    headers = [
        "기준월", "자료테이블", "입금일자", "입금액", "입금자명/메모",
        "성명", "차량번호", "계정", "매칭상태", "비고"
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    _style_sheet(ws)

    # 요약 시트
    ws2 = wb.create_sheet("입금요약", 0)
    total = 0
    for r in rows:
        amt = str(r.get("입금액", "")).replace(",", "").replace("원", "").strip()
        try:
            total += int(float(amt))
        except Exception:
            pass

    ws2.append(["기준월", "입금건수", "입금합계", "비고"])
    ws2.append([f"{year}-{month:02d}", len(rows), total, "DB 내 입금/통장/거래/납부 관련 테이블에서 자동 추출"])
    _style_sheet(ws2)

    return wb


def _excel_response(wb, filename):
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    # 한글 파일명은 HTTP 헤더에서 반드시 URL 인코딩해야 함
    safe_filename = quote(filename)
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"
    }
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.get("/billing-count/export")
def export_billing_count_excel(
    year: int = Query(..., description="기준년도"),
    month: int = Query(..., ge=1, le=12, description="기준월")
):
    try:
        db_gen = _get_db()
        db = next(db_gen)
        tables = _get_tables_and_rows(db, mode="all", limit=5000)
        wb = _billing_count_workbook(year, month, tables)
        filename = f"{year}년_{month:02d}월_부과대수.xlsx"
        return _excel_response(wb, filename)
    except Exception as e:
        print("ERROR: billing count export failed:", repr(e))
        wb = _error_workbook("부과대수 오류", repr(e))
        filename = f"{year}년_{month:02d}월_부과대수_오류.xlsx"
        return _excel_response(wb, filename)




def _find_recent_payment_rows(db, year, month):
    """
    N? ????:
    /arrears ??? ?? ???? ?????? YYYY-MM? ??? ?? ???.
    ?? ??? ???? / ?? / ?? / ??? ????.
    """
    from sqlalchemy import inspect, text as sql_text

    ym = f"{year}-{month:02d}"

    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    result = []

    recent_date_candidates = [
        "?????", "?? ???", "??????", "??? ???",
        "last_payment_date", "recent_payment_date", "latest_payment_date",
        "last_paid_at", "paid_at", "payment_date"
    ]

    vehicle_candidates = [
        "????", "?? ??", "vehicle_number", "car_number", "car_no", "plate_number"
    ]

    name_candidates = [
        "??", "??", "???", "name", "member_name", "owner_name"
    ]

    account_candidates = [
        "??", "??", "????", "account", "account_type", "category", "fee_type"
    ]

    amount_candidates = [
        "??", "???", "????", "?????", "?? ???",
        "amount", "payment_amount", "paid_amount", "deposit_amount",
        "last_payment_amount", "recent_payment_amount"
    ]

    # arrears ?? ??? ??
    preferred_words = [
        "arrears", "member", "dues", "fee", "misu", "??", "??", "??", "??"
    ]

    sorted_tables = sorted(
        tables,
        key=lambda t: 0 if any(w in t.lower() for w in preferred_words) else 1
    )

    for table in sorted_tables:
        try:
            cols = [c["name"] for c in inspector.get_columns(table)]
            if not cols:
                continue

            col_recent = _find_col(cols, recent_date_candidates)
            if not col_recent:
                continue

            col_vehicle = _find_col(cols, vehicle_candidates)
            col_name = _find_col(cols, name_candidates)
            col_account = _find_col(cols, account_candidates)
            col_amount = _find_col(cols, amount_candidates)

            # ??? ????? + ??/??/?? ? ???? ??? ?
            if not (col_vehicle or col_name or col_amount):
                continue

            q = sql_text(
                'SELECT * FROM "' + table + '" '
                'WHERE CAST("' + col_recent + '" AS TEXT) LIKE :ym '
                'ORDER BY "' + col_recent + '" ASC'
            )

            rows = db.execute(q, {"ym": ym + "%"}).mappings().all()

            for row in rows:
                row = dict(row)
                result.append({
                    "????": _safe_str(row.get(col_vehicle)),
                    "??": _safe_str(row.get(col_name)),
                    "??": _safe_str(row.get(col_account)),
                    "??": _safe_str(row.get(col_amount)),
                    "_table": table,
                    "_recent_col": col_recent,
                    "_recent_value": _safe_str(row.get(col_recent)),
                })

        except Exception as e:
            print("WARN: recent payment table skipped", table, repr(e))
            continue

    # ?? ??: ???? + ?? + ?? + ?? ??
    unique = {}
    for r in result:
        key = (
            r.get("????", ""),
            r.get("??", ""),
            r.get("??", ""),
            r.get("??", ""),
        )
        unique[key] = r

    return list(unique.values())


def _make_simple_deposit_workbook(year, month, rows):
    wb = Workbook()

    ws = wb.active
    ws.title = "summary"

    total = 0
    for r in rows:
        amt = str(r.get("??", "")).replace(",", "").replace("?", "").strip()
        try:
            total += int(float(amt))
        except Exception:
            pass

    ws.append(["???", "????", "????", "??"])
    ws.append([
        f"{year}-{month:02d}",
        len(rows),
        total,
        "?????? ???? ???? ??? ?? ? ?? ??"
    ])
    _style_sheet(ws)

    ws2 = wb.create_sheet(title="deposit_list")
    headers = ["????", "??", "??", "??"]
    ws2.append(headers)

    for r in rows:
        ws2.append([
            r.get("????", ""),
            r.get("??", ""),
            r.get("??", ""),
            r.get("??", ""),
        ])

    _style_sheet(ws2)
    return wb




def _find_monthly_deposit_export_rows_v4(db, year, month):
    """
    N? ???? ?? ??:
    - ?????/?????? ?? ??? ?? ???? ?? ??
    - YYYY-MM?? ???? ? ??? ?? ?? ???? ??
    - ??? '?? ???? 624?'? ?? ??? ?? ?? ??
    - ??? ???? / ?? / ?? / ?? 4??
    """
    from sqlalchemy import inspect, text as sql_text

    ym = f"{year}-{month:02d}"
    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    recent_candidates = [
        "?????", "?? ???",
        "??????", "??? ???",
        "?????", "?? ???",
        "last_payment_date", "recent_payment_date", "latest_payment_date",
        "last_paid_at", "last_deposit_date", "recent_deposit_date",
        "paid_at", "payment_date"
    ]

    vehicle_candidates = [
        "????", "?? ??",
        "vehicle_number", "vehicle_no", "car_number", "car_no", "plate_number"
    ]

    name_candidates = [
        "??", "??", "???",
        "name", "member_name", "owner_name", "person_name"
    ]

    account_candidates = [
        "??", "??", "????",
        "account", "account_type", "category", "fee_type", "dues_type"
    ]

    amount_candidates = [
        "??", "???", "????", "?????", "?? ???",
        "??????", "?? ????",
        "amount", "payment_amount", "paid_amount", "deposit_amount",
        "last_payment_amount", "recent_payment_amount", "last_paid_amount"
    ]

    candidates = []

    for table in tables:
        try:
            cols = [c["name"] for c in inspector.get_columns(table)]
            if not cols:
                continue

            col_recent = _find_col(cols, recent_candidates)
            if not col_recent:
                continue

            col_vehicle = _find_col(cols, vehicle_candidates)
            col_name = _find_col(cols, name_candidates)
            col_account = _find_col(cols, account_candidates)
            col_amount = _find_col(cols, amount_candidates)

            if not col_vehicle and not col_name:
                continue

            q_count = sql_text(
                'SELECT COUNT(*) FROM "' + table + '" '
                'WHERE CAST("' + col_recent + '" AS TEXT) LIKE :ym'
            )
            cnt = db.execute(q_count, {"ym": ym + "%"}).scalar() or 0

            if int(cnt) <= 0:
                continue

            candidates.append({
                "table": table,
                "count": int(cnt),
                "recent": col_recent,
                "vehicle": col_vehicle,
                "name": col_name,
                "account": col_account,
                "amount": col_amount,
            })

        except Exception as e:
            print("WARN: deposit v4 scan skipped", table, repr(e))
            continue

    if not candidates:
        return [], {
            "source_table": "",
            "recent_col": "",
            "count": 0,
            "message": f"{ym} ????? ??? ??"
        }

    # ??: ??? ?? ????? ??? ?? ??? ?? ?? ???? ??
    candidates.sort(key=lambda x: x["count"], reverse=True)
    src = candidates[0]

    q_rows = sql_text(
        'SELECT * FROM "' + src["table"] + '" '
        'WHERE CAST("' + src["recent"] + '" AS TEXT) LIKE :ym '
        'ORDER BY CAST("' + src["recent"] + '" AS TEXT) ASC'
    )

    db_rows = db.execute(q_rows, {"ym": ym + "%"}).mappings().all()

    rows = []
    for row in db_rows:
        row = dict(row)
        rows.append({
            "????": _safe_str(row.get(src["vehicle"])) if src["vehicle"] else "",
            "??": _safe_str(row.get(src["name"])) if src["name"] else "",
            "??": _safe_str(row.get(src["account"])) if src["account"] else "",
            "??": _safe_str(row.get(src["amount"])) if src["amount"] else "",
        })

    return rows, {
        "source_table": src["table"],
        "recent_col": src["recent"],
        "count": len(rows),
        "all_candidates": candidates[:10],
    }


def _make_monthly_deposit_export_workbook_v4(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "deposit_export"

    headers = ["????", "??", "??", "??"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("????", ""),
            r.get("??", ""),
            r.get("??", ""),
            r.get("??", ""),
        ])

    _style_sheet(ws)
    return wb




def _find_monthly_deposit_export_rows_v5(db, year, month):
    """
    Monthly deposit export.
    Uses unicode escapes for Korean labels so Excel headers never become ????.
    """
    from sqlalchemy import inspect, text as sql_text

    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
    NAME = "\uc774\ub984"                 # ??
    ACCOUNT = "\uacc4\uc815"              # ??
    AMOUNT = "\uae08\uc561"               # ??

    ym = f"{year}-{month:02d}"
    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    recent_candidates = [
        "\ucd5c\uadfc\uc785\uae08\uc77c",              # ?????
        "\ucd5c\uadfc \uc785\uae08\uc77c",             # ?? ???
        "\ub9c8\uc9c0\ub9c9\uc785\uae08\uc77c",        # ??????
        "\ub9c8\uc9c0\ub9c9 \uc785\uae08\uc77c",       # ??? ???
        "\ucd5c\uadfc\ub0a9\ubd80\uc77c",              # ?????
        "\ucd5c\uadfc \ub0a9\ubd80\uc77c",             # ?? ???
        "last_payment_date", "recent_payment_date", "latest_payment_date",
        "last_paid_at", "last_deposit_date", "recent_deposit_date",
        "paid_at", "payment_date"
    ]

    vehicle_candidates = [
        "\ucc28\ub7c9\ubc88\ud638", "\ucc28\ub7c9 \ubc88\ud638",
        "vehicle_number", "vehicle_no", "car_number", "car_no", "plate_number"
    ]

    name_candidates = [
        "\uc131\uba85", "\uc774\ub984", "\ud68c\uc6d0\uba85",
        "name", "member_name", "owner_name", "person_name"
    ]

    account_candidates = [
        "\uacc4\uc815", "\uad6c\ubd84", "\ud68c\ube44\uad6c\ubd84",
        "account", "account_type", "category", "fee_type", "dues_type"
    ]

    amount_candidates = [
        "\uae08\uc561", "\uc785\uae08\uc561", "\ub0a9\ubd80\uae08\uc561",
        "\ucd5c\uadfc\uc785\uae08\uc561", "\ucd5c\uadfc \uc785\uae08\uc561",
        "\ucd5c\uadfc\ub0a9\ubd80\uae08\uc561", "\ucd5c\uadfc \ub0a9\ubd80\uae08\uc561",
        "amount", "payment_amount", "paid_amount", "deposit_amount",
        "last_payment_amount", "recent_payment_amount", "last_paid_amount"
    ]

    candidates = []

    for table in tables:
        try:
            cols = [c["name"] for c in inspector.get_columns(table)]
            if not cols:
                continue

            col_recent = _find_col(cols, recent_candidates)
            if not col_recent:
                continue

            col_vehicle = _find_col(cols, vehicle_candidates)
            col_name = _find_col(cols, name_candidates)
            col_account = _find_col(cols, account_candidates)
            col_amount = _find_col(cols, amount_candidates)

            if not col_vehicle and not col_name:
                continue

            q_count = sql_text(
                'SELECT COUNT(*) FROM "' + table + '" '
                'WHERE CAST("' + col_recent + '" AS TEXT) LIKE :ym'
            )
            cnt = db.execute(q_count, {"ym": ym + "%"}).scalar() or 0

            if int(cnt) <= 0:
                continue

            candidates.append({
                "table": table,
                "count": int(cnt),
                "recent": col_recent,
                "vehicle": col_vehicle,
                "name": col_name,
                "account": col_account,
                "amount": col_amount,
            })

        except Exception as e:
            print("WARN: deposit v5 scan skipped", table, repr(e))
            continue

    if not candidates:
        return [], {
            "source_table": "",
            "recent_col": "",
            "count": 0,
            "message": f"{ym} data not found"
        }

    # Choose the table with the largest monthly payer count.
    candidates.sort(key=lambda x: x["count"], reverse=True)
    src = candidates[0]

    q_rows = sql_text(
        'SELECT * FROM "' + src["table"] + '" '
        'WHERE CAST("' + src["recent"] + '" AS TEXT) LIKE :ym '
        'ORDER BY CAST("' + src["recent"] + '" AS TEXT) ASC'
    )

    db_rows = db.execute(q_rows, {"ym": ym + "%"}).mappings().all()

    rows = []
    for row in db_rows:
        row = dict(row)
        rows.append({
            VEHICLE: _safe_str(row.get(src["vehicle"])) if src["vehicle"] else "",
            NAME: _safe_str(row.get(src["name"])) if src["name"] else "",
            ACCOUNT: _safe_str(row.get(src["account"])) if src["account"] else "",
            AMOUNT: _safe_str(row.get(src["amount"])) if src["amount"] else "",
        })

    return rows, {
        "source_table": src["table"],
        "recent_col": src["recent"],
        "count": len(rows),
        "all_candidates": candidates[:10],
    }


def _make_monthly_deposit_export_workbook_v5(rows):
    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
    NAME = "\uc774\ub984"                 # ??
    ACCOUNT = "\uacc4\uc815"              # ??
    AMOUNT = "\uae08\uc561"               # ??

    wb = Workbook()
    ws = wb.active
    ws.title = "deposit_export"

    headers = [VEHICLE, NAME, ACCOUNT, AMOUNT]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get(VEHICLE, ""),
            r.get(NAME, ""),
            r.get(ACCOUNT, ""),
            r.get(AMOUNT, ""),
        ])

    _style_sheet(ws)

    # N? ???? ?? ??? ??
    ws.column_dimensions["A"].width = 18  # ????
    ws.column_dimensions["B"].width = 16  # ??
    ws.column_dimensions["C"].width = 14  # ??
    ws.column_dimensions["D"].width = 14  # ??

    return wb




def _extract_name_from_memo_for_export(memo):
    import re
    s = str(memo or "").strip()
    m = re.search(r"[?-?]{2,5}", s)
    return m.group(0) if m else ""


def _extract_vehicle_tail_from_memo_for_export(memo):
    import re
    s = str(memo or "")
    # ??80?1629, ??81?2589 ?? ???? ?? ??
    m = re.search(r"[?-?]{2}\d{2}[?-?]\d{3,4}", s)
    if m:
        return m.group(0)

    # ?? ?/?? ?? 3~4???? ?? ?????? ??
    m = re.search(r"\d{3,4}", s)
    return m.group(0) if m else ""


def _find_monthly_bank_deposit_rows_4cols(db, year, month):
    """
    ???? ??? ?? ???? ??:
    bank_transactions?? ????? YYYY-MM? ?? ????? ????.
    ?? ??? ???? / ?? / ?? / ?? 4??.
    """
    from sqlalchemy import inspect, text as sql_text
    from datetime import datetime

    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
    NAME = "\uc774\ub984"                 # ??
    ACCOUNT = "\uacc4\uc815"              # ??
    AMOUNT = "\uae08\uc561"               # ??

    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    if "bank_transactions" not in tables:
        raise Exception("bank_transactions ???? ????.")

    cols = [c["name"] for c in inspector.get_columns("bank_transactions")]

    col_date = _find_col(cols, [
        "\uc785\uae08\uc77c\uc790", "\uac70\ub798\uc77c\uc790", "\ub0a9\ubd80\uc77c\uc790",
        "transaction_date", "deposit_date", "payment_date", "paid_at", "date", "created_at"
    ])
    col_amount = _find_col(cols, [
        "\uc785\uae08\uc561", "\uae08\uc561", "amount", "deposit_amount", "paid_amount"
    ])
    col_memo = _find_col(cols, [
        "\uc774\uccb4\uba54\ubaa8", "\uba54\ubaa8", "\uc785\uae08\uc790\uba85", "\ubcf4\ub0b8\ubd84",
        "memo", "description", "sender", "depositor", "payer_name"
    ])
    col_target = _find_col(cols, [
        "\ub9e4\uce6d\ub300\uc0c1", "\ub300\uc0c1", "matched_target", "match_target"
    ])
    col_account = _find_col(cols, [
        "\uacc4\uc815", "\uad6c\ubd84", "account", "account_type", "category", "fee_type"
    ])

    if not col_date:
        raise Exception("bank_transactions?? ???? ??? ?? ?????.")

    start_dt = datetime(year, month, 1)
    if month == 12:
        end_dt = datetime(year + 1, 1, 1)
    else:
        end_dt = datetime(year, month + 1, 1)

    q = sql_text(
        'SELECT * FROM "bank_transactions" '
        'WHERE "' + col_date + '" >= :start_dt '
        'AND "' + col_date + '" < :end_dt '
        'ORDER BY "' + col_date + '" ASC'
    )

    db_rows = db.execute(q, {"start_dt": start_dt, "end_dt": end_dt}).mappings().all()

    rows = []
    for row in db_rows:
        row = dict(row)

        memo = row.get(col_memo) if col_memo else ""
        target = str(row.get(col_target) or "") if col_target else ""

        vehicle = ""
        name = ""

        # ????? ??/??? ??? ?? ??
        if target and target not in ["-", "?", "None"]:
            vehicle = _extract_vehicle_tail_from_memo_for_export(target)
            name = _extract_name_from_memo_for_export(target)

        # ??? ???????? ??
        if not vehicle:
            vehicle = _extract_vehicle_tail_from_memo_for_export(memo)
        if not name:
            name = _extract_name_from_memo_for_export(memo)

        rows.append({
            VEHICLE: vehicle,
            NAME: name,
            ACCOUNT: _safe_str(row.get(col_account)) if col_account else "",
            AMOUNT: _safe_str(row.get(col_amount)) if col_amount else "",
        })

    return rows, {
        "source_table": "bank_transactions",
        "date_col": col_date,
        "amount_col": col_amount or "",
        "memo_col": col_memo or "",
        "count": len(rows),
    }


def _make_deposit_export_bank4_workbook(rows):
    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"
    NAME = "\uc774\ub984"
    ACCOUNT = "\uacc4\uc815"
    AMOUNT = "\uae08\uc561"

    wb = Workbook()
    ws = wb.active
    ws.title = "deposit_export"

    headers = [VEHICLE, NAME, ACCOUNT, AMOUNT]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get(VEHICLE, ""),
            r.get(NAME, ""),
            r.get(ACCOUNT, ""),
            r.get(AMOUNT, ""),
        ])

    _style_sheet(ws)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    return wb




def _extract_korean_name_for_export(memo):
    import re
    s = str(memo or "")
    m = re.search(r"[?-?]{2,5}", s)
    return m.group(0) if m else ""


def _extract_vehicle_for_export(memo):
    import re
    s = str(memo or "")

    # ??80?1629 / ??81?2589 ?? ?? ????
    m = re.search(r"[?-?]{2}\d{2}[?-?]\d{3,4}", s)
    if m:
        return m.group(0)

    # ?? ?? ?? ?? ?? ???
    m = re.search(r"\d{3,4}", s)
    if m:
        return m.group(0)

    return ""


def _find_real_monthly_deposit_rows(db, year, month):
    """
    ??? ?? ???? ???? ?? ?? ?? ???? ??.
    ?? ??? ?? ?? CAST(date AS TEXT) LIKE 'YYYY-MM%' ? ????.
    """
    from sqlalchemy import inspect, text as sql_text

    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
    NAME = "\uc774\ub984"                 # ??
    ACCOUNT = "\uacc4\uc815"              # ??
    AMOUNT = "\uae08\uc561"               # ??

    ym = f"{year}-{month:02d}"

    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    date_candidates = [
        "\uc785\uae08\uc77c\uc790",        # ????
        "\uac70\ub798\uc77c\uc790",        # ????
        "\ub0a9\ubd80\uc77c\uc790",        # ????
        "transaction_date", "deposit_date", "payment_date",
        "paid_at", "date", "created_at"
    ]

    memo_candidates = [
        "\uc774\uccb4\uba54\ubaa8",        # ????
        "\uba54\ubaa8",                    # ??
        "\uc785\uae08\uc790\uba85",        # ????
        "\ubcf4\ub0b8\ubd84",              # ???
        "memo", "description", "sender", "depositor", "payer_name"
    ]

    amount_candidates = [
        "\uc785\uae08\uc561",              # ???
        "\uae08\uc561",                    # ??
        "amount", "deposit_amount", "paid_amount", "payment_amount"
    ]

    account_candidates = [
        "\uacc4\uc815",                    # ??
        "\uad6c\ubd84",                    # ??
        "account", "account_type", "category", "fee_type", "type"
    ]

    target_candidates = [
        "\ub9e4\uce6d\ub300\uc0c1",        # ????
        "\ub300\uc0c1",                    # ??
        "matched_target", "match_target"
    ]

    possible_sources = []

    for table in tables:
        try:
            cols = [c["name"] for c in inspector.get_columns(table)]
            if not cols:
                continue

            col_date = _find_col(cols, date_candidates)
            col_memo = _find_col(cols, memo_candidates)
            col_amount = _find_col(cols, amount_candidates)

            if not col_date or not col_amount:
                continue

            # memo? ??? ????/?? ??? ?? ? ??? memo ??? ?? ???? ??
            col_account = _find_col(cols, account_candidates)
            col_target = _find_col(cols, target_candidates)

            q_count = sql_text(
                'SELECT COUNT(*) FROM "' + table + '" '
                'WHERE CAST("' + col_date + '" AS TEXT) LIKE :ym'
            )
            cnt = db.execute(q_count, {"ym": ym + "%"}).scalar() or 0

            if int(cnt) <= 0:
                continue

            possible_sources.append({
                "table": table,
                "count": int(cnt),
                "date": col_date,
                "memo": col_memo,
                "amount": col_amount,
                "account": col_account,
                "target": col_target,
            })

        except Exception as e:
            print("WARN: real deposit source scan skipped", table, repr(e))
            continue

    if not possible_sources:
        return [], {
            "source_table": "",
            "date_col": "",
            "count": 0,
            "sources": [],
        }

    # ??? 624?? ??? ?? ? ?? ? ?? ?? ?? ??? ??
    possible_sources.sort(key=lambda x: x["count"], reverse=True)
    src = possible_sources[0]

    q_rows = sql_text(
        'SELECT * FROM "' + src["table"] + '" '
        'WHERE CAST("' + src["date"] + '" AS TEXT) LIKE :ym '
        'ORDER BY CAST("' + src["date"] + '" AS TEXT) ASC'
    )

    db_rows = db.execute(q_rows, {"ym": ym + "%"}).mappings().all()

    rows = []
    for row in db_rows:
        row = dict(row)

        memo = _safe_str(row.get(src["memo"])) if src["memo"] else ""
        target = _safe_str(row.get(src["target"])) if src["target"] else ""

        base_text = target if target and target not in ["-", "?", "None"] else memo

        rows.append({
            VEHICLE: _extract_vehicle_for_export(base_text),
            NAME: _extract_korean_name_for_export(base_text),
            ACCOUNT: _safe_str(row.get(src["account"])) if src["account"] else "",
            AMOUNT: _safe_str(row.get(src["amount"])) if src["amount"] else "",
        })

    return rows, {
        "source_table": src["table"],
        "date_col": src["date"],
        "memo_col": src["memo"] or "",
        "amount_col": src["amount"] or "",
        "account_col": src["account"] or "",
        "count": len(rows),
        "sources": possible_sources[:10],
    }


def _make_real_monthly_deposit_workbook(rows):
    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"
    NAME = "\uc774\ub984"
    ACCOUNT = "\uacc4\uc815"
    AMOUNT = "\uae08\uc561"

    wb = Workbook()
    ws = wb.active
    ws.title = "deposit_export"

    headers = [VEHICLE, NAME, ACCOUNT, AMOUNT]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get(VEHICLE, ""),
            r.get(NAME, ""),
            r.get(ACCOUNT, ""),
            r.get(AMOUNT, ""),
        ])

    _style_sheet(ws)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    return wb




def _member_monthly_deposit_rows(db, year, month):
    """
    /arrears ??? ?? ???? ??:
    Member.last_paid_date ? YYYY-MM ?? ???? ?? ?? ??.
    ?? ??: ???? / ?? / ?? / ??
    """
    from sqlalchemy import cast, String
    from models import Member

    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
    NAME = "\uc774\ub984"                 # ??
    ACCOUNT = "\uacc4\uc815"              # ??
    AMOUNT = "\uae08\uc561"               # ??

    ym = f"{year}-{month:02d}"

    q = db.query(Member).filter(
        cast(Member.last_paid_date, String).like(ym + "%")
    )

    members = q.order_by(Member.region, Member.name).all()

    rows = []
    for m in members:
        vehicle = (
            getattr(m, "vehicle_number", None)
            or getattr(m, "car_number", None)
            or getattr(m, "plate_number", None)
            or getattr(m, "vehicle_no", None)
            or ""
        )

        name = getattr(m, "name", "") or ""

        account = (
            getattr(m, "account", None)
            or getattr(m, "account_type", None)
            or getattr(m, "fee_type", None)
            or getattr(m, "category", None)
            or ""
        )

        # ??? ????? ??? ??? ?? ??, ??? ???/????? ?? ??
        amount = (
            getattr(m, "last_paid_amount", None)
            or getattr(m, "recent_payment_amount", None)
            or getattr(m, "paid_amount", None)
            or getattr(m, "payment_amount", None)
            or getattr(m, "monthly_fee", None)
            or getattr(m, "excel_arrears", None)
            or ""
        )

        rows.append({
            VEHICLE: vehicle,
            NAME: name,
            ACCOUNT: account,
            AMOUNT: amount,
        })

    return rows


def _make_member_monthly_deposit_workbook(rows):
    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"
    NAME = "\uc774\ub984"
    ACCOUNT = "\uacc4\uc815"
    AMOUNT = "\uae08\uc561"

    wb = Workbook()
    ws = wb.active
    ws.title = "deposit_export"

    headers = [VEHICLE, NAME, ACCOUNT, AMOUNT]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get(VEHICLE, ""),
            r.get(NAME, ""),
            r.get(ACCOUNT, ""),
            r.get(AMOUNT, ""),
        ])

    _style_sheet(ws)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    return wb




def _arrears_current_rows_for_deposit_export(db):
    """
    /arrears ??? ?? ??? ??? ??.
    ?? total_count? ??? ?? Member.excel_arrears > 0 ???? ?? ??.
    ?? ??: ???? / ?? / ?? / ??
    """
    from models import Member

    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
    NAME = "\uc774\ub984"                 # ??
    ACCOUNT = "\uacc4\uc815"              # ??
    AMOUNT = "\uae08\uc561"               # ??

    q = db.query(Member).filter(Member.excel_arrears > 0)

    members = q.order_by(Member.region, Member.name).all()

    rows = []
    for m in members:
        vehicle = (
            getattr(m, "vehicle_number", None)
            or getattr(m, "car_number", None)
            or getattr(m, "plate_number", None)
            or getattr(m, "vehicle_no", None)
            or ""
        )

        name = getattr(m, "name", "") or ""

        account = (
            getattr(m, "account", None)
            or getattr(m, "account_type", None)
            or getattr(m, "fee_type", None)
            or getattr(m, "category", None)
            or ""
        )

        amount = getattr(m, "excel_arrears", "") or ""

        rows.append({
            VEHICLE: vehicle,
            NAME: name,
            ACCOUNT: account,
            AMOUNT: amount,
        })

    return rows


def _make_arrears_current_export_workbook(rows):
    VEHICLE = "\ucc28\ub7c9\ubc88\ud638"
    NAME = "\uc774\ub984"
    ACCOUNT = "\uacc4\uc815"
    AMOUNT = "\uae08\uc561"

    wb = Workbook()
    ws = wb.active
    ws.title = "deposit_export"

    headers = [VEHICLE, NAME, ACCOUNT, AMOUNT]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get(VEHICLE, ""),
            r.get(NAME, ""),
            r.get(ACCOUNT, ""),
            r.get(AMOUNT, ""),
        ])

    _style_sheet(ws)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    return wb



@router.get("/deposit/rematch-name-last4")
def rematch_deposit_name_last4():
    try:
        db_gen = _get_db()
        db = next(db_gen)
        result = _rematch_confirm_needed_name_last4(db)
        return result
    except Exception as e:
        print("ERROR: rematch failed:", repr(e))
        return {
            "status": "error",
            "message": repr(e),
        }




# =========================================================
# ???? ???? ??? v2
# - DB ???? ??? '????'? ???? ?? ???/????? ??
# - ????? ?? + ???? ?4??? ?? ??? ???? ??
# =========================================================
def _find_confirm_needed_transaction_source(db):
    from sqlalchemy import inspect, text as sql_text

    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    status_words = ["????", "?"]
    memo_candidates = ["????", "??", "????", "???", "memo", "description", "sender", "depositor", "payer_name"]
    amount_candidates = ["???", "??", "amount", "deposit_amount", "paid_amount"]
    status_candidates = ["??", "????", "status", "match_status", "matching_status"]

    best = None

    for t in tables:
        try:
            cols = [c["name"] for c in inspector.get_columns(t)]
            if not cols:
                continue

            memo_col = _find_col(cols, memo_candidates)
            amount_col = _find_col(cols, amount_candidates)

            # ??/?? ??? ??? ?? ???? ???? ??
            if not memo_col:
                continue

            # ?? ?? ?? ???? ???? ??
            possible_status_cols = []
            for c in cols:
                cn = _norm_col(c)
                if any(_norm_col(x) in cn for x in status_candidates):
                    possible_status_cols.append(c)

            # ??? ??? ?? ??? ?? ??? ??
            if not possible_status_cols:
                possible_status_cols = cols

            for status_col in possible_status_cols:
                try:
                    q = sql_text(
                        'SELECT COUNT(*) AS cnt FROM "' + t + '" '
                        'WHERE CAST("' + status_col + '" AS TEXT) LIKE :kw1 '
                        'OR CAST("' + status_col + '" AS TEXT) LIKE :kw2'
                    )
                    cnt = db.execute(q, {"kw1": "%????%", "kw2": "%?%"}).scalar() or 0

                    if cnt > 0:
                        if best is None or cnt > best["count"]:
                            best = {
                                "table": t,
                                "status_col": status_col,
                                "memo_col": memo_col,
                                "amount_col": amount_col,
                                "count": int(cnt),
                                "columns": cols,
                            }
                except Exception:
                    continue

        except Exception as e:
            print("WARN: confirm source scan skipped", t, repr(e))
            continue

    return best


def _rematch_confirm_needed_name_last4_v2(db):
    from sqlalchemy import inspect, text as sql_text

    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    source = _find_confirm_needed_transaction_source(db)

    if not source:
        return {
            "status": "error",
            "message": "DB?? ????? ???? ?? ???/??? ?? ?????.",
            "checked": 0,
            "updated": 0,
            "examples": [],
        }

    tx_table = source["table"]
    col_status = source["status_col"]
    col_memo = source["memo_col"]
    col_amount = source["amount_col"]

    tx_cols = source["columns"]

    tx_pk = None
    try:
        pk_cols = inspector.get_pk_constraint(tx_table).get("constrained_columns") or []
        if pk_cols:
            tx_pk = pk_cols[0]
    except Exception:
        pass

    if not tx_pk:
        tx_pk = _find_col(tx_cols, ["id", "transaction_id", "bank_transaction_id", "payment_id"])

    if not tx_pk:
        raise Exception(f"{tx_table} primary key not found")

    col_target = _find_col(tx_cols, ["????", "??", "matched_target", "match_target"])
    col_reason = _find_col(tx_cols, ["??", "????", "reason", "match_reason"])

    # ?? ?? ??
    member_candidates = []

    for t in tables:
        if t == tx_table:
            continue

        try:
            cols = [c["name"] for c in inspector.get_columns(t)]
            col_name = _find_col(cols, ["??", "??", "???", "name", "member_name", "owner_name"])
            col_vehicle = _find_col(cols, ["????", "?? ??", "vehicle_number", "car_number", "car_no", "plate_number"])
            col_account = _find_col(cols, ["??", "??", "????", "account", "account_type", "category", "fee_type"])

            if not col_name or not col_vehicle:
                continue

            q = sql_text(
                'SELECT "' + col_name + '" AS name, "' + col_vehicle + '" AS vehicle'
                + (', "' + col_account + '" AS account' if col_account else ', NULL AS account')
                + ' FROM "' + t + '"'
            )

            for r in db.execute(q).mappings().all():
                name = _safe_str(r.get("name"))
                vehicle = _safe_str(r.get("vehicle"))
                last4 = _vehicle_last4(vehicle)

                if name and last4:
                    member_candidates.append({
                        "name": name,
                        "vehicle": vehicle,
                        "last4": last4,
                        "account": _safe_str(r.get("account")),
                        "table": t,
                    })

        except Exception as e:
            print("WARN: member candidate table skipped", t, repr(e))
            continue

    # ?? ?? ??
    unique = {}
    for c in member_candidates:
        unique[(c["name"], c["vehicle"], c["last4"], c["account"])] = c
    member_candidates = list(unique.values())

    # ?? ???? ?? ??
    q = sql_text(
        'SELECT * FROM "' + tx_table + '" '
        'WHERE CAST("' + col_status + '" AS TEXT) LIKE :kw1 '
        'OR CAST("' + col_status + '" AS TEXT) LIKE :kw2'
    )
    tx_rows = db.execute(q, {"kw1": "%????%", "kw2": "%?%"}).mappings().all()

    updated = 0
    skipped_none = 0
    skipped_multi = 0
    examples = []

    for row in tx_rows:
        row = dict(row)
        memo = _compact_text(row.get(col_memo))

        matched = []
        for c in member_candidates:
            name_c = _compact_text(c["name"])
            last4 = c["last4"]

            if name_c and last4 and name_c in memo and last4 in memo:
                matched.append(c)

        # ? 1??? ???? ??? ?? ??
        if len(matched) == 1:
            c = matched[0]

            set_parts = ['"' + col_status + '" = :new_status']
            params = {
                "new_status": "????",
                "pk": row.get(tx_pk),
            }

            if col_target:
                set_parts.append('"' + col_target + '" = :target')
                params["target"] = f'{c["name"]} {c["vehicle"]}'.strip()

            if col_reason:
                old_reason = _safe_str(row.get(col_reason))
                set_parts.append('"' + col_reason + '" = :reason')
                params["reason"] = (old_reason + ", " if old_reason else "") + "??+?????4????"

            update_q = sql_text(
                'UPDATE "' + tx_table + '" SET '
                + ", ".join(set_parts)
                + ' WHERE "' + tx_pk + '" = :pk'
            )

            db.execute(update_q, params)
            updated += 1

            if len(examples) < 30:
                examples.append({
                    "memo": _safe_str(row.get(col_memo)),
                    "matched": f'{c["name"]} {c["vehicle"]}',
                    "amount": _safe_str(row.get(col_amount)) if col_amount else "",
                })

        elif len(matched) > 1:
            skipped_multi += 1
        else:
            skipped_none += 1

    db.commit()

    return {
        "status": "ok",
        "source_table": tx_table,
        "status_col": col_status,
        "memo_col": col_memo,
        "checked": len(tx_rows),
        "updated": updated,
        "skipped_none": skipped_none,
        "skipped_multi": skipped_multi,
        "examples": examples,
    }


@router.get("/deposit/rematch-name-last4-v2")
def rematch_deposit_name_last4_v2():
    try:
        db_gen = _get_db()
        db = next(db_gen)
        return _rematch_confirm_needed_name_last4_v2(db)
    except Exception as e:
        print("ERROR: rematch v2 failed:", repr(e))
        return {
            "status": "error",
            "message": repr(e),
        }




# =========================================================
# ???? ??? v3
# - ??? ???? 81? ?? ?? ??
# - ????? ????+??????? ?? ???????+????? ?? ??? ????
# - ?: ???1780, ???2970, 1347???, ??80?1629???
# =========================================================
def _memo_has_korean_name_with_tail_number(memo):
    memo = _compact_text(memo)

    # ?? 2~5? + ?? 3~4??
    # ?: ???1780, ???304, ???1756
    if re.search(r"[?-?]{2,5}\d{3,4}", memo):
        return True

    # ?? 3~4?? + ?? 2~5?
    # ?: 1347???, 1442???
    if re.search(r"\d{3,4}[?-?]{2,5}", memo):
        return True

    # ???? ?? + ??
    # ?: ??80?1629???, ??81?2589?
    if re.search(r"[?-?]{2}\d{2}[?-?]\d{3,4}[?-?]{1,5}", memo):
        return True

    return False


def _rematch_confirm_needed_by_visible_memo_pattern(db):
    from sqlalchemy import inspect, text as sql_text

    bind = db.get_bind()
    inspector = inspect(bind)
    tables = inspector.get_table_names()

    # ?? ????? ???? ???/?? ????? v2 ?? ???
    source = _find_confirm_needed_transaction_source(db)

    if not source:
        return {
            "status": "error",
            "message": "????? ???? ?? ???/??? ?? ?????.",
            "checked": 0,
            "updated": 0,
            "examples": [],
        }

    tx_table = source["table"]
    col_status = source["status_col"]
    col_memo = source["memo_col"]
    col_amount = source["amount_col"]
    tx_cols = source["columns"]

    tx_pk = None
    try:
        pk_cols = inspector.get_pk_constraint(tx_table).get("constrained_columns") or []
        if pk_cols:
            tx_pk = pk_cols[0]
    except Exception:
        pass

    if not tx_pk:
        tx_pk = _find_col(tx_cols, ["id", "transaction_id", "bank_transaction_id", "payment_id"])

    if not tx_pk:
        raise Exception(f"{tx_table} primary key not found")

    col_reason = _find_col(tx_cols, ["??", "????", "reason", "match_reason"])

    q = sql_text(
        'SELECT * FROM "' + tx_table + '" '
        'WHERE CAST("' + col_status + '" AS TEXT) LIKE :kw1 '
        'OR CAST("' + col_status + '" AS TEXT) LIKE :kw2'
    )

    rows = db.execute(q, {"kw1": "%????%", "kw2": "%?%"}).mappings().all()

    checked = len(rows)
    updated = 0
    skipped = 0
    examples = []

    for row in rows:
        row = dict(row)
        memo = _safe_str(row.get(col_memo))
        reason = _safe_str(row.get(col_reason)) if col_reason else ""

        # ????? ?? ? ? ???,
        # ??? ??+??? ??? ??? ????
        # ???? ??? ?? ??? ?? ????
        can_update = False

        if "????" in reason and _memo_has_korean_name_with_tail_number(memo):
            can_update = True

        if "?????????" in reason and _memo_has_korean_name_with_tail_number(memo):
            can_update = True

        if not can_update:
            skipped += 1
            continue

        set_parts = ['"' + col_status + '" = :new_status']
        params = {
            "new_status": "????",
            "pk": row.get(tx_pk),
        }

        if col_reason:
            old_reason = _safe_str(row.get(col_reason))
            set_parts.append('"' + col_reason + '" = :reason')
            params["reason"] = (old_reason + ", " if old_reason else "") + "??? ??+??????? ????"

        update_q = sql_text(
            'UPDATE "' + tx_table + '" SET '
            + ", ".join(set_parts)
            + ' WHERE "' + tx_pk + '" = :pk'
        )

        db.execute(update_q, params)
        updated += 1

        if len(examples) < 50:
            examples.append({
                "memo": memo,
                "amount": _safe_str(row.get(col_amount)) if col_amount else "",
                "reason": reason,
            })

    db.commit()

    return {
        "status": "ok",
        "source_table": tx_table,
        "status_col": col_status,
        "memo_col": col_memo,
        "checked": checked,
        "updated": updated,
        "skipped": skipped,
        "examples": examples,
    }


@router.get("/deposit/rematch-visible-pattern")
def rematch_deposit_visible_pattern():
    try:
        db_gen = _get_db()
        db = next(db_gen)
        return _rematch_confirm_needed_by_visible_memo_pattern(db)
    except Exception as e:
        print("ERROR: rematch visible pattern failed:", repr(e))
        return {
            "status": "error",
            "message": repr(e),
        }


@router.get("/deposit/export")
def export_deposit_excel(
    year: int = Query(..., description="year"),
    month: int = Query(..., ge=1, le=12, description="month")
):
    """
    N? ???? ???.
    /arrears ??? ?? ??? ?? ???? Member ????? ?? ??.
    ?? ??? ???? / ?? / ?? / ?? 4??.
    """
    try:
        from models import Member
        from openpyxl import Workbook

        db_gen = _get_db()
        db = next(db_gen)

        VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
        NAME = "\uc774\ub984"                 # ??
        ACCOUNT = "\uacc4\uc815"              # ??
        AMOUNT = "\uae08\uc561"               # ??

        # /arrears? ?? ??? ?? ??
        q = db.query(Member).filter(Member.excel_arrears > 0)
        members = q.order_by(Member.region, Member.name).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "deposit_export"
        ws.append([VEHICLE, NAME, ACCOUNT, AMOUNT])

        for m in members:
            vehicle = (
                getattr(m, "vehicle_number", None)
                or getattr(m, "car_number", None)
                or getattr(m, "plate_number", None)
                or getattr(m, "vehicle_no", None)
                or ""
            )

            name = getattr(m, "name", "") or ""

            account = (
                getattr(m, "account", None)
                or getattr(m, "account_type", None)
                or getattr(m, "fee_type", None)
                or getattr(m, "category", None)
                or ""
            )

            amount = getattr(m, "excel_arrears", "") or ""

            ws.append([vehicle, name, account, amount])

        _style_sheet(ws)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14

        filename = f"{year}_{month:02d}_deposit_export.xlsx"
        return _excel_response(wb, filename)

    except Exception as e:
        print("ERROR: final deposit export failed:", repr(e))
        wb = _error_workbook("deposit export error", repr(e))
        filename = f"{year}_{month:02d}_deposit_export_error.xlsx"
        return _excel_response(wb, filename)


# =========================================================
# N? ???? ? ???
# ?? /deposit/export ?? ???
# ??: /api/reports/deposit/current-members-export
# ??: ???? / ?? / ?? / ??
# ??: /arrears ?? ?? Member.excel_arrears > 0
# =========================================================
@router.get("/deposit/current-members-export")
def export_current_members_deposit_excel(
    year: int = Query(..., description="year"),
    month: int = Query(..., ge=1, le=12, description="month")
):
    try:
        from models import Member
        from openpyxl import Workbook

        db_gen = _get_db()
        db = next(db_gen)

        VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
        NAME = "\uc774\ub984"                 # ??
        ACCOUNT = "\uacc4\uc815"              # ??
        AMOUNT = "\uae08\uc561"               # ??

        q = db.query(Member).filter(Member.excel_arrears > 0)
        members = q.order_by(Member.region, Member.name).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "deposit_export"
        ws.append([VEHICLE, NAME, ACCOUNT, AMOUNT])

        for m in members:
            vehicle = (
                getattr(m, "vehicle_number", None)
                or getattr(m, "car_number", None)
                or getattr(m, "plate_number", None)
                or getattr(m, "vehicle_no", None)
                or getattr(m, "car_no", None)
                or ""
            )

            name = getattr(m, "name", "") or ""

            account = (
                getattr(m, "account", None)
                or getattr(m, "account_type", None)
                or getattr(m, "fee_type", None)
                or getattr(m, "category", None)
                or ""
            )

            amount = getattr(m, "excel_arrears", "") or ""

            ws.append([vehicle, name, account, amount])

        _style_sheet(ws)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14

        ws_debug = wb.create_sheet(title="debug_source")
        ws_debug.append(["item", "value"])
        ws_debug.append(["route", "/api/reports/deposit/current-members-export"])
        ws_debug.append(["basis", "Member.excel_arrears > 0"])
        ws_debug.append(["count", len(members)])
        ws_debug.sheet_state = "hidden"

        filename = f"{year}_{month:02d}_deposit_current_members.xlsx"
        return _excel_response(wb, filename)

    except Exception as e:
        print("ERROR: current members deposit export failed:", repr(e))
        wb = _error_workbook("deposit export error", repr(e))
        filename = f"{year}_{month:02d}_deposit_current_members_error.xlsx"
        return _excel_response(wb, filename)


# =========================================================
# ?? ??? ?? ???
# ??: MonthlyLedger.year/month + paid_amount > 0
# ???? "?? ????"? ?? ??
# ??: ???? / ?? / ?? / ??
# =========================================================
@router.get("/deposit/monthly-ledger-export")
def export_monthly_ledger_deposit_excel(
    year: int = Query(..., description="year"),
    month: int = Query(..., ge=1, le=12, description="month")
):
    try:
        from sqlalchemy import func
        from models import Member, MonthlyLedger
        from openpyxl import Workbook

        db_gen = _get_db()
        db = next(db_gen)

        VEHICLE = "\ucc28\ub7c9\ubc88\ud638"  # ????
        NAME = "\uc774\ub984"                 # ??
        ACCOUNT = "\uacc4\uc815"              # ??
        AMOUNT = "\uae08\uc561"               # ??

        # ???? ?? ????? ?? ??:
        # MonthlyLedger.year/month + paid_amount > 0 + member_id? ??
        paid_rows = (
            db.query(
                MonthlyLedger.member_id.label("member_id"),
                func.sum(MonthlyLedger.paid_amount).label("paid_sum")
            )
            .filter(MonthlyLedger.year == year)
            .filter(MonthlyLedger.month == month)
            .filter(MonthlyLedger.paid_amount > 0)
            .group_by(MonthlyLedger.member_id)
            .all()
        )

        member_ids = [r.member_id for r in paid_rows if r.member_id]
        members = {}
        if member_ids:
            for m in db.query(Member).filter(Member.id.in_(member_ids)).all():
                members[m.id] = m

        wb = Workbook()
        ws = wb.active
        ws.title = "deposit_export"
        ws.append([VEHICLE, NAME, ACCOUNT, AMOUNT])

        total_amount = 0

        for r in paid_rows:
            m = members.get(r.member_id)
            paid_sum = int(r.paid_sum or 0)
            total_amount += paid_sum

            if m:
                vehicle = (
                    getattr(m, "vehicle_no", None)
                    or getattr(m, "vehicle_number", None)
                    or getattr(m, "car_no", None)
                    or getattr(m, "plate_number", None)
                    or ""
                )
                name = getattr(m, "name", "") or ""
                account = getattr(m, "account", "") or ""
            else:
                vehicle = ""
                name = ""
                account = ""

            ws.append([vehicle, name, account, paid_sum])

        _style_sheet(ws)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14

        ws_debug = wb.create_sheet(title="debug_source")
        ws_debug.append(["item", "value"])
        ws_debug.append(["basis", "MonthlyLedger.year/month + paid_amount > 0"])
        ws_debug.append(["year", year])
        ws_debug.append(["month", month])
        ws_debug.append(["count", len(paid_rows)])
        ws_debug.append(["total_amount", total_amount])
        ws_debug.sheet_state = "hidden"

        filename = f"{year}_{month:02d}_monthly_deposit_export.xlsx"
        return _excel_response(wb, filename)

    except Exception as e:
        print("ERROR: monthly ledger deposit export failed:", repr(e))
        wb = _error_workbook("monthly ledger deposit export error", repr(e))
        filename = f"{year}_{month:02d}_monthly_deposit_export_error.xlsx"
        return _excel_response(wb, filename)


# =========================================================
# ???? ??
# 1) ?? ?? ???? ?? ??
# 2) ??? ????? ???? ??? ??
# =========================================================
@router.get("/deposit/cleanup-bank-status")
def cleanup_bank_status():
    try:
        import re
        from models import BankTransaction
        from sqlalchemy import or_

        db_gen = _get_db()
        db = next(db_gen)

        def memo_text(tx):
            return str(getattr(tx, "memo", "") or "")

        def has_korean_name(s):
            return bool(re.search(r"[?-?]{2,5}", s or ""))

        def has_vehicle_like(s):
            s = str(s or "")
            return bool(
                re.search(r"[?-?]{2}\d{2}[?-?]\d{3,4}", s)
                or re.search(r"\d{3,4}", s)
            )

        def is_noise_memo(s):
            s = str(s or "").strip()
            compact = re.sub(r"\s+", "", s)

            # ?? ??? ?? ?? ?? ??
            if compact in ["??", "??", "???"]:
                return True

            # ??/??/??? ?? ??
            if re.fullmatch(r"[A-Za-z0-9\-_\(\)]+", compact):
                return True

            # ??? + ??/??? ?? ??/?? ??? ?? ??
            bank_words = ["??", "??", "??", "??", "??", "???", "???", "??"]
            if any(w in compact for w in bank_words):
                if not has_korean_name(compact) and not re.search(r"[?-?]{2}\d{2}[?-?]\d{3,4}", compact):
                    return True

            # ??? ?? ???? ??? ??? ??
            if not has_korean_name(compact) and not re.search(r"[?-?]{2}\d{2}[?-?]\d{3,4}", compact):
                return True

            return False

        excluded = 0
        upgraded = 0
        checked = 0
        examples_excluded = []
        examples_upgraded = []

        txs = (
            db.query(BankTransaction)
            .filter(BankTransaction.match_status.in_(["???", "????", "? ????"]))
            .all()
        )

        for tx in txs:
            checked += 1
            memo = memo_text(tx)
            reason = str(getattr(tx, "match_reason", "") or "")

            # 1) ?? ?? ???? ?? ??
            if str(tx.match_status or "") == "???" and is_noise_memo(memo):
                tx.match_status = "??"
                tx.match_reason = (reason + ", " if reason else "") + "????: ??/???? ???? ??"
                db.add(tx)
                excluded += 1
                if len(examples_excluded) < 20:
                    examples_excluded.append(memo)
                continue

            # 2) ???? ? ???? ?? ? ???? ??? ??
            strong_reason = (
                "??=?????" in reason
                or "???????" in reason
            )

            name_tail_pattern = bool(
                re.search(r"[?-?]{2,5}\d{3,4}", memo)
                or re.search(r"\d{3,4}[?-?]{2,5}", memo)
                or re.search(r"[?-?]{2}\d{2}[?-?]\d{3,4}[?-?]{1,5}", memo)
            )

            if "????" in str(tx.match_status or ""):
                if strong_reason or name_tail_pattern:
                    tx.match_status = "????"
                    tx.match_reason = (reason + ", " if reason else "") + "??: ???? ??"
                    db.add(tx)
                    upgraded += 1
                    if len(examples_upgraded) < 20:
                        examples_upgraded.append(memo)

        db.commit()

        return {
            "status": "ok",
            "checked": checked,
            "excluded_to_??": excluded,
            "confirm_to_????": upgraded,
            "examples_excluded": examples_excluded,
            "examples_upgraded": examples_upgraded,
        }

    except Exception as e:
        print("ERROR: cleanup bank status failed:", repr(e))
        return {"status": "error", "message": repr(e)}


# =========================================================
# ?? ?? ???
# ?: ??? ??? / ???? 2974 / ?? "??, ? ? ?"
#     ???? "???2974??" -> ????? ????
# =========================================================
@router.get("/deposit/rematch-alias-note")
def rematch_bank_by_alias_note():
    try:
        import re
        from models import Member, BankTransaction

        db_gen = _get_db()
        db = next(db_gen)

        def compact(s):
            return re.sub(r"\s+", "", str(s or "")).strip()

        def digits(s):
            return re.sub(r"\D", "", str(s or ""))

        def vehicle_last4(v):
            d = digits(v)
            return d[-4:] if len(d) >= 4 else d

        def extract_aliases_from_note(note):
            """
            ???? ??? ?? ?? ??.
            '??, ? ? ?' -> ['???']
            """
            raw = str(note or "")
            aliases = set()

            # ??/???/?? ???? ??
            parts = re.split(r"[,/()\[\]{}:;|]+", raw)

            stop = {
                "??", "??", "??", "??", "??", "??", "??",
                "??", "??", "???", "???", "??", "??",
                "??", "??", "??", "??", "??"
            }

            for part in parts:
                p = compact(part)
                if not p or p in stop:
                    continue

                # ?? 2~5? ??? ??? ??
                found = re.findall(r"[?-?]{2,5}", p)
                for f in found:
                    f = compact(f)
                    if f and f not in stop and 2 <= len(f) <= 5:
                        aliases.add(f)

            return list(aliases)

        def member_note_text(m):
            texts = []
            for attr in [
                "note", "memo", "remark", "remarks", "raw_note",
                "bigo", "??", "etc", "description"
            ]:
                if hasattr(m, attr):
                    v = getattr(m, attr)
                    if v:
                        texts.append(str(v))
            return " ".join(texts)

        # 1) ??? ?? ?? ???
        candidates = []

        members = db.query(Member).all()
        for m in members:
            vehicle = (
                getattr(m, "vehicle_no", None)
                or getattr(m, "vehicle_number", None)
                or getattr(m, "car_no", None)
                or getattr(m, "plate_number", None)
                or ""
            )
            last4 = vehicle_last4(vehicle)
            if not last4:
                continue

            aliases = set()

            # ?? ?? ??? ??? ??
            name = compact(getattr(m, "name", "") or "")
            if name:
                aliases.add(name)

            # ?? ? ?? ??
            for a in extract_aliases_from_note(member_note_text(m)):
                aliases.add(a)

            for a in aliases:
                candidates.append({
                    "member": m,
                    "alias": a,
                    "last4": last4,
                    "vehicle": vehicle,
                })

        checked = 0
        updated = 0
        skipped_none = 0
        skipped_multi = 0
        examples = []

        txs = (
            db.query(BankTransaction)
            .filter(BankTransaction.match_status.in_(["???", "????", "? ????"]))
            .all()
        )

        for tx in txs:
            checked += 1
            memo = compact(getattr(tx, "memo", "") or "")

            matched = []
            for c in candidates:
                if c["alias"] and c["last4"] and c["alias"] in memo and c["last4"] in memo:
                    matched.append(c)

            # ?? ?? ?? ?? ??
            uniq = {}
            for c in matched:
                uniq[c["member"].id] = c
            matched = list(uniq.values())

            if len(matched) == 1:
                c = matched[0]
                m = c["member"]

                tx.matched_member_id = m.id
                tx.match_status = "????"
                old_reason = getattr(tx, "match_reason", "") or ""
                add_reason = f"????+?????????: {c['alias']}+{c['last4']} -> {m.name}/{c['vehicle']}"
                tx.match_reason = (old_reason + ", " if old_reason else "") + add_reason

                db.add(tx)
                updated += 1

                if len(examples) < 30:
                    examples.append({
                        "memo": getattr(tx, "memo", "") or "",
                        "matched": f"{m.name} / {c['vehicle']}",
                        "alias": c["alias"],
                        "last4": c["last4"],
                    })

            elif len(matched) > 1:
                skipped_multi += 1
            else:
                skipped_none += 1

        db.commit()

        return {
            "status": "ok",
            "checked": checked,
            "updated_to_auto": updated,
            "skipped_none": skipped_none,
            "skipped_multi": skipped_multi,
            "examples": examples,
        }

    except Exception as e:
        print("ERROR: rematch alias note failed:", repr(e))
        return {"status": "error", "message": repr(e)}


# =========================================================
# ?? ?? 1? ?? ????
# - ???/???? ?? ?
# - ???? ? ?? ??? ???? ??? ??
# - ??? ? 1??? ?????? ??
# =========================================================
@router.get("/deposit/rematch-single-name")
def rematch_single_name_bank_transactions():
    try:
        import re
        from models import Member, BankTransaction

        db_gen = _get_db()
        db = next(db_gen)

        def compact(s):
            return re.sub(r"\s+", "", str(s or "")).strip()

        def extract_names_from_memo(memo):
            s = compact(memo)
            # ?? ?? ?? 2~5?
            names = re.findall(r"[?-?]{2,5}", s)
            stop = {
                "??", "??", "??", "??", "??", "???", "??",
                "??", "??", "??", "??", "??", "??", "??",
                "??", "??", "??", "??", "???", "???"
            }
            return [n for n in names if n not in stop]

        def get_amount(tx):
            for attr in ["amount", "deposit_amount", "in_amount", "paid_amount", "txn_amount", "money"]:
                if hasattr(tx, attr):
                    v = getattr(tx, attr)
                    try:
                        if v is not None:
                            return int(str(v).replace(",", "").replace("?", "").strip())
                    except Exception:
                        pass
            return 0

        # ?? ?? ?? ??? ??? ??
        members = (
            db.query(Member)
            .filter(Member.name != None, Member.name != "")
            .filter(Member.vehicle_no != None, Member.vehicle_no != "")
            .all()
        )

        name_map = {}
        for m in members:
            nk = compact(m.name)
            if not nk:
                continue
            name_map.setdefault(nk, []).append(m)

        txs = (
            db.query(BankTransaction)
            .filter(BankTransaction.match_status.in_(["???", "????", "? ????"]))
            .all()
        )

        checked = 0
        updated = 0
        skipped_none = 0
        skipped_multi = 0
        examples = []

        for tx in txs:
            checked += 1
            memo = getattr(tx, "memo", "") or ""
            names = extract_names_from_memo(memo)

            matched_members = {}

            for n in names:
                for m in name_map.get(compact(n), []):
                    matched_members[m.id] = m

            if len(matched_members) == 1:
                m = list(matched_members.values())[0]
                amount = get_amount(tx)

                tx.matched_member_id = m.id

                old_reason = getattr(tx, "match_reason", "") or ""
                add_reason = f"?? ?? 1? ?? ????: {m.name}/{m.vehicle_no}"

                # ???? 0??? ?? ????? ????? ???? ??
                if amount <= 0:
                    tx.match_status = "????"
                    add_reason += " / ???????"
                else:
                    tx.match_status = "????"

                tx.match_reason = (old_reason + ", " if old_reason else "") + add_reason

                db.add(tx)
                updated += 1

                if len(examples) < 30:
                    examples.append({
                        "memo": memo,
                        "matched": f"{m.name} / {m.vehicle_no}",
                        "amount": amount,
                        "status": tx.match_status,
                    })

            elif len(matched_members) > 1:
                skipped_multi += 1
            else:
                skipped_none += 1

        db.commit()

        return {
            "status": "ok",
            "checked": checked,
            "updated": updated,
            "skipped_none": skipped_none,
            "skipped_multi": skipped_multi,
            "examples": examples,
        }

    except Exception as e:
        print("ERROR: single name rematch failed:", repr(e))
        return {"status": "error", "message": repr(e)}

